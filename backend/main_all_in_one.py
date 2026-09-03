"""
CyberDojo backend — complete single-file FastAPI app.

Contains everything: database config, models, security helpers, user-facing
routes, and admin panel routes. Designed to be uploaded as one file via VNC.

Run with:
    uvicorn main_all_in_one:app --host 0.0.0.0 --port 8000

Or rename this file to main.py on the server and run:
    uvicorn main:app --host 0.0.0.0 --port 8000

Requires (pip install):
    fastapi uvicorn[standard] sqlalchemy psycopg2-binary pydantic[email]
    python-dotenv passlib[bcrypt] bcrypt PyJWT razorpay python-multipart

Environment (.env):
    DATABASE_URL=postgresql://<user>:<password>@localhost/cyberdojo
    SECRET_KEY=<long-random-string>
    RAZORPAY_KEY_ID=<your-razorpay-key-id>
    RAZORPAY_KEY_SECRET=<your-razorpay-secret>
    AWS_ACCESS_KEY_ID=<your-aws-access-key-id>
    AWS_SECRET_ACCESS_KEY=<your-aws-secret-access-key>
    SMTP_USER=<smtp-username>
    SMTP_PASSWORD=<smtp-app-password>

Migrations: run migrations.sql against the DB before starting the server.
Seed first superadmin manually (see README.md).
"""

# ============================================================================
# IMPORTS
# ============================================================================
import asyncio
import base64
import ipaddress
import json as _json
import os
import hmac
import hashlib
import random
import socket
import string
import urllib.error
import urllib.request
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime, timedelta, date, timezone
from typing import Any, Dict, List, Optional
from urllib.parse import urlparse

import boto3
import jwt
import razorpay
from Crypto.Cipher import PKCS1_v1_5
from Crypto.PublicKey import RSA
from dotenv import load_dotenv
from fastapi import BackgroundTasks, FastAPI, Depends, File, Form, HTTPException, Query, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import OAuth2PasswordBearer
from pydantic import BaseModel, EmailStr
from passlib.context import CryptContext
from sqlalchemy import (
    create_engine, Column, Integer, String, Boolean, ForeignKey,
    DateTime, Numeric, Text, UniqueConstraint, Index,
    or_, func, cast, Date,
)
from sqlalchemy.dialects.postgresql import JSONB
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import declarative_base, sessionmaker, Session, aliased
from sqlalchemy.sql import func as sqlfunc

load_dotenv()

# ============================================================================
# CONFIG
# ============================================================================
DATABASE_URL = os.getenv("DATABASE_URL", "")
if not DATABASE_URL:
    raise RuntimeError(
        "DATABASE_URL is not set. Add it to .env as "
        "postgresql://<user>:<password>@<host>/<database>"
    )
SECRET_KEY = os.getenv("SECRET_KEY", "CHANGE_ME_IN_PRODUCTION")
ALGORITHM = "HS256"
USER_TOKEN_DEFAULT_HOURS = 24
USER_TOKEN_REMEMBER_DAYS = 30
ADMIN_TOKEN_EXPIRE_HOURS = 12

# ============================================================================
# DATABASE
# ============================================================================
engine = create_engine(DATABASE_URL, pool_pre_ping=True)
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
Base = declarative_base()


def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


# ============================================================================
# MODELS — user-facing
# ============================================================================
class User(Base):
    __tablename__ = "users"
    id = Column(Integer, primary_key=True, index=True)
    email = Column(String, unique=True, index=True, nullable=False)
    full_name = Column(String)
    country = Column(String)
    phone = Column(String)
    hashed_password = Column(String, nullable=False)
    newsletter_opt_in = Column(Boolean, default=False)
    terms_agreed = Column(Boolean, default=False)
    available_credits = Column(Integer, default=0)
    referral_code = Column(String, unique=True)
    is_active = Column(Boolean, default=True)
    created_at = Column(DateTime(timezone=True), server_default=sqlfunc.now())


class Subscription(Base):
    __tablename__ = "subscriptions"
    id = Column(Integer, primary_key=True, index=True)
    user_id = Column(Integer, ForeignKey("users.id", ondelete="CASCADE"))
    plan_name = Column(String, nullable=False)
    billing_cycle = Column(String, nullable=False)
    start_date = Column(DateTime(timezone=True), server_default=sqlfunc.now())
    end_date = Column(DateTime(timezone=True), nullable=False)
    is_active = Column(Boolean, default=True)
    # Razorpay Subscriptions API (recurring) — nullable so legacy one-shot rows still load.
    razorpay_subscription_id = Column(String(64), index=True, nullable=True)
    razorpay_customer_id = Column(String(64), nullable=True)
    # 'manual' = legacy one-shot, 'pending' | 'active' | 'halted' |
    # 'cancelled_pending' (cancel queued, access until end_date) | 'cancelled' | 'completed'
    auto_renew_status = Column(String(20), default="manual")
    next_charge_at = Column(DateTime(timezone=True), nullable=True)
    current_period_end = Column(DateTime(timezone=True), nullable=True)


class SubscriptionEvent(Base):
    __tablename__ = "subscription_events"
    id = Column(Integer, primary_key=True, index=True)
    razorpay_event_id = Column(String(128), unique=True, index=True)
    razorpay_subscription_id = Column(String(64), index=True)
    event_type = Column(String(50), nullable=False)
    payload = Column(JSONB)
    received_at = Column(DateTime(timezone=True), server_default=sqlfunc.now())
    processed = Column(Boolean, default=False)


class CreditTransaction(Base):
    __tablename__ = "credit_transactions"
    id = Column(Integer, primary_key=True, index=True)
    user_id = Column(Integer, ForeignKey("users.id", ondelete="CASCADE"))
    plan_name = Column(String)
    credits_added = Column(Integer, nullable=False)
    amount_paid = Column(Numeric(10, 2), nullable=False)
    purchased_at = Column(DateTime(timezone=True), server_default=sqlfunc.now())


class CreditPack(Base):
    __tablename__ = "credit_packs"
    id = Column(Integer, primary_key=True, index=True)
    pack_name = Column(String(50), unique=True, nullable=False)
    credits = Column(Integer, nullable=False)
    price = Column(Numeric(10, 2), nullable=False)
    is_active = Column(Boolean, default=True, index=True)
    display_order = Column(Integer, default=0)
    description = Column(Text)
    updated_at = Column(DateTime(timezone=True), server_default=sqlfunc.now(), onupdate=sqlfunc.now())


class CourseCatalog(Base):
    __tablename__ = "course_catalog"
    id = Column(Integer, primary_key=True, index=True)
    slug = Column(String(50), unique=True, nullable=False, index=True)
    title = Column(String(200), nullable=False)
    tagline = Column(String(300))
    description = Column(Text)
    category = Column(String(50))
    difficulty = Column(String(20))   # CHECK enforced in SQL
    modules_count = Column(Integer)
    labs_count = Column(Integer)
    duration_hours = Column(Integer)
    price_inr = Column(Numeric(10, 2))
    currency = Column(String(8), default="INR")
    billing_label = Column(String(50))
    hero_image_url = Column(Text)
    accent_color = Column(String(20), default="red")
    audience = Column(JSONB, nullable=False, default=list)
    benefits = Column(JSONB, nullable=False, default=list)
    syllabus = Column(JSONB, nullable=False, default=list)
    is_active = Column(Boolean, default=True, index=True)
    display_order = Column(Integer, default=0)
    updated_at = Column(DateTime(timezone=True), server_default=sqlfunc.now(), onupdate=sqlfunc.now())


class CoursePurchase(Base):
    __tablename__ = "course_purchases"
    id = Column(Integer, primary_key=True, index=True)
    user_id = Column(Integer, ForeignKey("users.id", ondelete="CASCADE"), nullable=False, index=True)
    course_id = Column(Integer, ForeignKey("course_catalog.id"), nullable=False, index=True)
    course_slug = Column(String(50), nullable=False)
    course_title = Column(String(200))
    base_price = Column(Numeric(10, 2), nullable=False)
    gst_amount = Column(Numeric(10, 2), nullable=False)
    total_amount = Column(Numeric(10, 2), nullable=False)
    razorpay_order_id = Column(String(64), unique=True)
    razorpay_payment_id = Column(String(64))
    purchased_at = Column(DateTime(timezone=True), server_default=sqlfunc.now())
    
class BootcampCatalog(Base):
    __tablename__ = "bootcamp_catalog"
    id = Column(Integer, primary_key=True, index=True)
    slug = Column(String(50), unique=True, nullable=False, index=True)
    title = Column(String(200), nullable=False)
    tagline = Column(String(300))
    description = Column(Text)
    category = Column(String(50))
    difficulty = Column(String(20))   # 'Beginner' | 'Intermediate' | 'Advanced'
    modules_count = Column(Integer)
    labs_count = Column(Integer)
    duration_hours = Column(Integer)
    price_inr = Column(Numeric(10, 2))
    currency = Column(String(8), default="INR")
    billing_label = Column(String(50))
    hero_image_url = Column(Text)
    accent_color = Column(String(20), default="red")
    audience = Column(JSONB, nullable=False, default=list)
    benefits = Column(JSONB, nullable=False, default=list)
    syllabus = Column(JSONB, nullable=False, default=list)
    is_active = Column(Boolean, default=True, index=True)
    display_order = Column(Integer, default=0)
    updated_at = Column(DateTime(timezone=True), server_default=sqlfunc.now(), onupdate=sqlfunc.now())


class BootcampPurchase(Base):
    __tablename__ = "bootcamp_purchases"
    id = Column(Integer, primary_key=True, index=True)
    user_id = Column(Integer, ForeignKey("users.id", ondelete="CASCADE"), nullable=False, index=True)
    bootcamp_id = Column(Integer, ForeignKey("bootcamp_catalog.id"), nullable=False, index=True)
    bootcamp_slug = Column(String(50), nullable=False)
    bootcamp_title = Column(String(200))
    base_price = Column(Numeric(10, 2), nullable=False)
    gst_amount = Column(Numeric(10, 2), nullable=False)
    total_amount = Column(Numeric(10, 2), nullable=False)
    razorpay_order_id = Column(String(64), unique=True)
    razorpay_payment_id = Column(String(64))
    purchased_at = Column(DateTime(timezone=True), server_default=sqlfunc.now())


class Coupon(Base):
    __tablename__ = "coupons"
    id = Column(Integer, primary_key=True, index=True)
    code = Column(String(50), unique=True, nullable=False, index=True)
    discount_percent = Column(Integer, nullable=False)   # 1..100
    description = Column(Text)
    max_uses = Column(Integer, nullable=True)            # None = unlimited
    uses_count = Column(Integer, nullable=False, default=0)
    expires_at = Column(DateTime(timezone=True), nullable=True)
    is_active = Column(Boolean, nullable=False, default=True)
    created_at = Column(DateTime(timezone=True), server_default=sqlfunc.now())
    updated_at = Column(DateTime(timezone=True), server_default=sqlfunc.now(), onupdate=sqlfunc.now())


class CouponRedemption(Base):
    __tablename__ = "coupon_redemptions"
    id = Column(Integer, primary_key=True, index=True)
    coupon_id = Column(Integer, ForeignKey("coupons.id", ondelete="CASCADE"), index=True)
    user_id = Column(Integer, ForeignKey("users.id", ondelete="CASCADE"), index=True)
    credit_transaction_id = Column(Integer, ForeignKey("credit_transactions.id", ondelete="SET NULL"), nullable=True)
    discount_amount = Column(Numeric(10, 2), nullable=False)
    redeemed_at = Column(DateTime(timezone=True), server_default=sqlfunc.now())
    __table_args__ = (UniqueConstraint("coupon_id", "user_id", name="_coupon_user_unique"),)


class ContactMessage(Base):
    __tablename__ = "contact_messages"
    id = Column(Integer, primary_key=True, index=True)
    name = Column(String(100), nullable=False)
    email = Column(String(255), nullable=False)
    message = Column(Text, nullable=False)
    status = Column(String(50), default="unread")
    admin_note = Column(Text)
    created_at = Column(DateTime(timezone=True), server_default=sqlfunc.now())


class Referral(Base):
    __tablename__ = "referrals"
    id = Column(Integer, primary_key=True, index=True)
    referrer_id = Column(Integer, ForeignKey("users.id", ondelete="CASCADE"))
    referred_user_id = Column(Integer, ForeignKey("users.id", ondelete="CASCADE"))
    reward_credits = Column(Integer, default=10)
    status = Column(String(50), default="completed")
    created_at = Column(DateTime(timezone=True), server_default=sqlfunc.now())


class UserProgress(Base):
    __tablename__ = "user_progress"
    id = Column(Integer, primary_key=True, index=True)
    user_id = Column(Integer, ForeignKey("users.id", ondelete="CASCADE"))
    course_name = Column(String(255), nullable=False)
    module_name = Column(String(255), nullable=False)
    is_completed = Column(Boolean, default=True)
    completed_at = Column(DateTime(timezone=True), server_default=sqlfunc.now())
    __table_args__ = (
        UniqueConstraint("user_id", "course_name", "module_name", name="_user_course_module_uc"),
    )
    
class BootcampProgress(Base):
    __tablename__ = "bootcamp_progress"
    id = Column(Integer, primary_key=True, index=True)
    user_id = Column(Integer, ForeignKey("users.id", ondelete="CASCADE"))
    bootcamp_name = Column(String(255), nullable=False)
    module_name = Column(String(255), nullable=False)
    is_completed = Column(Boolean, default=True)
    completed_at = Column(DateTime(timezone=True), server_default=sqlfunc.now())
    __table_args__ = (
        UniqueConstraint("user_id", "bootcamp_name", "module_name", name="_user_bootcamp_module_uc"),
    )


# ============================================================================
# MODELS — University LMS
# ============================================================================
class University(Base):
    __tablename__ = "universities"
    id = Column(Integer, primary_key=True, index=True)
    name = Column(String(200), nullable=False)
    slug = Column(String(100), unique=True, nullable=False)
    domain = Column(String(100))
    enforce_domain = Column(Boolean, default=False)
    logo_url = Column(Text)
    description = Column(Text)
    credits_per_student = Column(Integer, default=0)
    is_active = Column(Boolean, default=True)
    display_order = Column(Integer, default=0)
    created_at = Column(DateTime(timezone=True), server_default=sqlfunc.now())

class UniversityProgram(Base):
    __tablename__ = "university_programs"
    id = Column(Integer, primary_key=True, index=True)
    university_id = Column(Integer, ForeignKey("universities.id", ondelete="CASCADE"), nullable=False, index=True)
    name = Column(String(200), nullable=False)
    slug = Column(String(100), nullable=False)
    description = Column(Text)
    credits_per_student = Column(Integer, default=0)
    is_active = Column(Boolean, default=True)
    display_order = Column(Integer, default=0)
    # Display-only stats shown in the LMS "About The Course" row. Stored as
    # short strings so admins control the exact rendering ("300 hrs", "100+").
    stat_duration = Column(String(20))
    stat_labs = Column(String(20))
    stat_modules = Column(String(20))
    created_at = Column(DateTime(timezone=True), server_default=sqlfunc.now())
    __table_args__ = (UniqueConstraint("university_id", "slug", name="_uni_prog_slug_uc"),)

class UniversitySemester(Base):
    __tablename__ = "university_semesters"
    id = Column(Integer, primary_key=True, index=True)
    university_id = Column(Integer, ForeignKey("universities.id", ondelete="CASCADE"), nullable=False, index=True)
    program_id = Column(Integer, ForeignKey("university_programs.id", ondelete="CASCADE"), index=True)
    semester_number = Column(Integer, nullable=False)
    name = Column(String(200), nullable=False)
    content = Column(JSONB, default=list)
    credits_grant = Column(Integer, default=0)
    # Cached result of the last link-reachability check, so the admin list can
    # show a status tag without re-fetching every external URL on page load.
    links_total = Column(Integer)
    links_ok = Column(Integer)
    links_warn = Column(Integer)
    links_checked_at = Column(DateTime(timezone=True))
    created_at = Column(DateTime(timezone=True), server_default=sqlfunc.now())

class UniversityStudent(Base):
    __tablename__ = "university_students"
    id = Column(Integer, primary_key=True, index=True)
    university_id = Column(Integer, ForeignKey("universities.id", ondelete="CASCADE"), nullable=False, index=True)
    program_id = Column(Integer, ForeignKey("university_programs.id", ondelete="SET NULL"), index=True)
    user_id = Column(Integer, ForeignKey("users.id", ondelete="CASCADE"), nullable=False, index=True)
    current_semester = Column(Integer, default=1)
    status = Column(String(20), default="pending")
    lms_paid = Column(Boolean, default=False)
    razorpay_order_id = Column(String(64))
    razorpay_payment_id = Column(String(64))
    enrolled_at = Column(DateTime(timezone=True), server_default=sqlfunc.now())
    promoted_at = Column(DateTime(timezone=True))
    __table_args__ = (UniqueConstraint("university_id", "user_id", name="_uni_user_uc"),)


# ============================================================================
# MODELS — assessments engine
# ============================================================================
class Assessment(Base):
    __tablename__ = "assessments"
    id = Column(Integer, primary_key=True, index=True)
    title = Column(String(200), nullable=False)
    slug = Column(String(100), unique=True, nullable=False, index=True)
    description = Column(Text)
    topic = Column(String(100))
    difficulty = Column(String(20), default="mixed")
    time_limit_minutes = Column(Integer, default=15)
    is_active = Column(Boolean, default=True)
    created_at = Column(DateTime(timezone=True), server_default=sqlfunc.now())


class AssessmentQuestion(Base):
    __tablename__ = "assessment_questions"
    id = Column(Integer, primary_key=True, index=True)
    assessment_id = Column(Integer, ForeignKey("assessments.id", ondelete="CASCADE"), nullable=False, index=True)
    question_text = Column(Text, nullable=False)
    options = Column(JSONB, nullable=False)
    correct_answers = Column(JSONB, nullable=False)
    multi_select = Column(Boolean, default=False)
    order_num = Column(Integer, default=0)


class AssessmentAttempt(Base):
    __tablename__ = "assessment_attempts"
    id = Column(Integer, primary_key=True, index=True)
    assessment_id = Column(Integer, ForeignKey("assessments.id"), nullable=False, index=True)
    user_id = Column(Integer, ForeignKey("users.id", ondelete="CASCADE"), nullable=False, index=True)
    started_at = Column(DateTime(timezone=True), server_default=sqlfunc.now())
    completed_at = Column(DateTime(timezone=True))
    score = Column(Integer)
    total_questions = Column(Integer)
    answers_payload = Column(JSONB)
    violations = Column(Integer, default=0)
    status = Column(String(20), default="in_progress")
    __table_args__ = (UniqueConstraint("assessment_id", "user_id", name="_asmt_user_uc"),)


class AssessmentViolation(Base):
    __tablename__ = "assessment_violations"
    id = Column(Integer, primary_key=True, index=True)
    attempt_id = Column(Integer, ForeignKey("assessment_attempts.id", ondelete="CASCADE"), nullable=False, index=True)
    user_id = Column(Integer, ForeignKey("users.id", ondelete="CASCADE"), nullable=False, index=True)
    violation_type = Column(String(50), default="tab_switch")
    occurred_at = Column(DateTime(timezone=True), server_default=sqlfunc.now())


# ============================================================================
# MODELS — admin panel
# ============================================================================
class Admin(Base):
    __tablename__ = "admins"
    id = Column(Integer, primary_key=True, index=True)
    email = Column(String(255), unique=True, nullable=False, index=True)
    hashed_password = Column(String(255), nullable=False)
    full_name = Column(String(255))
    role = Column(String(20), default="admin")
    is_active = Column(Boolean, default=True)
    created_at = Column(DateTime(timezone=True), server_default=sqlfunc.now())
    last_login_at = Column(DateTime(timezone=True), nullable=True)


class AdminAuditLog(Base):
    __tablename__ = "admin_audit_log"
    id = Column(Integer, primary_key=True, index=True)
    admin_id = Column(Integer, ForeignKey("admins.id", ondelete="CASCADE"), index=True)
    action = Column(String(50), nullable=False)
    entity_type = Column(String(50), nullable=False)
    entity_id = Column(String(100))
    before_payload = Column(JSONB)
    after_payload = Column(JSONB)
    created_at = Column(DateTime(timezone=True), server_default=sqlfunc.now())
    __table_args__ = (Index("idx_audit_entity", "entity_type", "entity_id"),)


class LabCatalog(Base):
    __tablename__ = "lab_catalog"
    lab_id = Column(String(100), primary_key=True)
    slug = Column(String(100), nullable=False)
    name = Column(String(255), nullable=False)
    category = Column(String(100), nullable=False, index=True)
    difficulty = Column(String(20), default="Medium")
    credits_cost = Column(Integer, nullable=False, default=1)
    duration_minutes = Column(Integer, nullable=False, default=60)
    os_type = Column(String(20), nullable=False, default="linux")
    description = Column(Text)
    is_active = Column(Boolean, default=True, index=True)
    created_at = Column(DateTime(timezone=True), server_default=sqlfunc.now())
    updated_at = Column(DateTime(timezone=True), server_default=sqlfunc.now(), onupdate=sqlfunc.now())


class EndLabTrial(Base):
    __tablename__ = "end_lab_trial"
    id = Column(Integer, primary_key=True, autoincrement=True)
    user_id = Column(Integer, ForeignKey("users.id"), nullable=False, index=True)
    lab_id = Column(String(100), nullable=True)
    instance_ids = Column(Text, nullable=False)
    started_at = Column(DateTime, default=datetime.utcnow)
    ends_at = Column(DateTime, nullable=True)


# ============================================================================
# SECURITY — password, JWT, auth dependencies, audit
# ============================================================================
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="api/auth/signin")


def verify_password(plain: str, hashed: str) -> bool:
    try:
        return pwd_context.verify(plain, hashed)
    except Exception:
        return False


def get_password_hash(password: str) -> str:
    return pwd_context.hash(password)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None) -> str:
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(hours=USER_TOKEN_DEFAULT_HOURS))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)


def create_admin_token(admin_id: int, role: str, email: str) -> str:
    payload = {
        "scope": "admin",
        "admin_id": admin_id,
        "role": role,
        "email": email,
        "exp": datetime.utcnow() + timedelta(hours=ADMIN_TOKEN_EXPIRE_HOURS),
    }
    return jwt.encode(payload, SECRET_KEY, algorithm=ALGORITHM)


def get_current_user(token: str = Depends(oauth2_scheme), db: Session = Depends(get_db)) -> User:
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        if payload.get("scope") == "admin":
            raise HTTPException(status_code=403, detail="User token required, not admin")
        user_id = payload.get("id")
        if user_id is None:
            raise HTTPException(status_code=401, detail="Invalid token")
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    return user


def get_current_admin(token: str = Depends(oauth2_scheme), db: Session = Depends(get_db)) -> Admin:
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

    if payload.get("scope") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")

    admin = db.query(Admin).filter(Admin.id == payload.get("admin_id"), Admin.is_active == True).first()
    if not admin:
        raise HTTPException(status_code=401, detail="Admin not found or disabled")
    return admin


def require_superadmin(admin: Admin = Depends(get_current_admin)) -> Admin:
    if admin.role != "superadmin":
        raise HTTPException(status_code=403, detail="Superadmin only")
    return admin


def write_audit(db: Session, admin_id: int, action: str, entity_type: str, entity_id, before=None, after=None):
    entry = AdminAuditLog(
        admin_id=admin_id,
        action=action,
        entity_type=entity_type,
        entity_id=str(entity_id) if entity_id is not None else None,
        before_payload=before,
        after_payload=after,
    )
    db.add(entry)
    db.commit()


# ============================================================================
# APP SETUP
# ============================================================================
app = FastAPI(title="CyberDojo API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # TODO: restrict to production domain
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Auto-create tables in dev only. In production, run migrations.sql instead.
# Uncomment the next line if you want SQLAlchemy to create missing tables on startup:
# Base.metadata.create_all(bind=engine)

razorpay_client = razorpay.Client(
    auth=(os.getenv("RAZORPAY_KEY_ID", ""), os.getenv("RAZORPAY_KEY_SECRET", ""))
)

# Razorpay Subscriptions API config.
# Six plans must be created in the Razorpay dashboard (Plans → Create), one per
# (plan_name, billing_cycle) combo. Capture each plan_XXX and set the matching
# env var on the server before activating recurring flow.
RAZORPAY_WEBHOOK_SECRET = os.getenv("RAZORPAY_WEBHOOK_SECRET", "")
RAZORPAY_PLAN_IDS = {
    ("Basic",   "monthly"): os.getenv("RZP_PLAN_BASIC_MONTHLY", ""),
    ("Basic",   "annual"):  os.getenv("RZP_PLAN_BASIC_ANNUAL", ""),
    ("Pro",     "monthly"): os.getenv("RZP_PLAN_PRO_MONTHLY", ""),
    ("Pro",     "annual"):  os.getenv("RZP_PLAN_PRO_ANNUAL", ""),
    ("Premium", "monthly"): os.getenv("RZP_PLAN_PREMIUM_MONTHLY", ""),
    ("Premium", "annual"):  os.getenv("RZP_PLAN_PREMIUM_ANNUAL", ""),
}
# Razorpay requires a finite max charge count on subscriptions.
# 60 monthly cycles = 5 yrs; 5 annual cycles = 5 yrs. Tune later if needed.
RAZORPAY_TOTAL_COUNT = {"monthly": 60, "annual": 5}

# GST applied to every customer-facing rupee amount. Mirrors GST_RATE in
# main-web/src/utils/pricing.js — keep both in sync. Same rounding rule
# (half-away-from-zero) on both sides so server validation never rejects a
# correctly-built request because of a paise off-by-one.
GST_RATE = float(os.getenv("GST_RATE", "0.18"))


def with_gst(base) -> int:
    """Base rupees -> inclusive rupees, integer. Mirrors withGst() in pricing.js."""
    try:
        n = float(base)
    except (TypeError, ValueError):
        return 0
    if n <= 0:
        return 0
    return int(round(n * (1 + GST_RATE)))


def _lookup_coupon(db: Session, code: str, user_id: int):
    """Resolve a coupon code to a row, or raise HTTPException with a user-friendly reason.

    Caller passes user_id so we can also reject codes the same user has already
    redeemed (enforced separately by the UNIQUE (coupon_id, user_id) constraint
    on coupon_redemptions, but this preflight gives a nicer error message).
    """
    if not code:
        raise HTTPException(status_code=400, detail="Coupon code required")
    coupon = (
        db.query(Coupon)
        .filter(Coupon.code == code.strip().upper(), Coupon.is_active == True)
        .first()
    )
    if not coupon:
        raise HTTPException(status_code=404, detail="Invalid or inactive coupon")
    if coupon.expires_at and coupon.expires_at < datetime.now(timezone.utc):
        raise HTTPException(status_code=400, detail="Coupon expired")
    if coupon.max_uses is not None and coupon.uses_count >= coupon.max_uses:
        raise HTTPException(status_code=400, detail="Coupon usage limit reached")
    already = (
        db.query(CouponRedemption)
        .filter(
            CouponRedemption.coupon_id == coupon.id,
            CouponRedemption.user_id == user_id,
        )
        .first()
    )
    if already:
        raise HTTPException(status_code=400, detail="You have already used this coupon")
    return coupon


def _apply_coupon_to_base(base_rupees: float, discount_percent: int) -> float:
    """Returns discounted base rupees (not rounded). GST is computed on this."""
    if discount_percent <= 0:
        return float(base_rupees)
    return float(base_rupees) * (1.0 - (discount_percent / 100.0))


# ============================================================================
# PYDANTIC SCHEMAS — user-facing
# ============================================================================
class UserCreate(BaseModel):
    email: EmailStr
    password: str
    newsletter: bool
    terms: bool
    full_name: Optional[str] = None
    country: Optional[str] = None
    phone: Optional[str] = None
    referred_by_code: Optional[str] = None


class UserLogin(BaseModel):
    email: EmailStr
    password: str
    rememberMe: bool = False


class CreateOrderRequest(BaseModel):
    user_id: int
    plan_name: str
    credits: int
    amount: int  # paise (post-discount, GST-inclusive)
    coupon_code: Optional[str] = None


class VerifyPaymentRequest(BaseModel):
    razorpay_order_id: str
    razorpay_payment_id: str
    razorpay_signature: str
    user_id: int
    plan_name: str
    credits: int
    amount: float
    coupon_code: Optional[str] = None


class SubscribePlan(BaseModel):
    user_id: int
    plan_name: str
    billing_cycle: str


class SubCreateOrder(BaseModel):
    user_id: int
    plan_name: str
    billing_cycle: str
    amount: int  # paise


class SubVerifyPayment(BaseModel):
    razorpay_order_id: str
    razorpay_payment_id: str
    razorpay_signature: str
    user_id: int
    plan_name: str
    billing_cycle: str
    amount: float


# Razorpay Subscriptions API (recurring) schemas
class SubCreateRecurring(BaseModel):
    user_id: int
    plan_name: str
    billing_cycle: str


class SubVerifyInitial(BaseModel):
    razorpay_payment_id: str
    razorpay_subscription_id: str
    razorpay_signature: str
    user_id: int


class SubUpgradeRequest(BaseModel):
    new_plan: str
    billing_cycle: str


# Credit pack admin schemas
class CreditPackPayload(BaseModel):
    pack_name: str
    credits: int
    price: float
    is_active: Optional[bool] = True
    display_order: Optional[int] = 0
    description: Optional[str] = None


class CreditPackUpdate(BaseModel):
    pack_name: Optional[str] = None
    credits: Optional[int] = None
    price: Optional[float] = None
    is_active: Optional[bool] = None
    display_order: Optional[int] = None
    description: Optional[str] = None


# Course catalog admin schemas (Part 8)
class CoursePayload(BaseModel):
    slug: str
    title: str
    tagline: Optional[str] = None
    description: Optional[str] = None
    category: Optional[str] = None
    difficulty: Optional[str] = None       # 'Beginner' | 'Intermediate' | 'Advanced'
    modules_count: Optional[int] = None
    labs_count: Optional[int] = None
    duration_hours: Optional[int] = None
    price_inr: Optional[float] = None
    currency: Optional[str] = "INR"
    billing_label: Optional[str] = None
    hero_image_url: Optional[str] = None
    accent_color: Optional[str] = "red"
    audience: Optional[List[Any]] = None
    benefits: Optional[List[Any]] = None
    syllabus: Optional[List[Dict[str, Any]]] = None
    is_active: Optional[bool] = True
    display_order: Optional[int] = 0


class CourseUpdate(BaseModel):
    slug: Optional[str] = None
    title: Optional[str] = None
    tagline: Optional[str] = None
    description: Optional[str] = None
    category: Optional[str] = None
    difficulty: Optional[str] = None
    modules_count: Optional[int] = None
    labs_count: Optional[int] = None
    duration_hours: Optional[int] = None
    price_inr: Optional[float] = None
    currency: Optional[str] = None
    billing_label: Optional[str] = None
    hero_image_url: Optional[str] = None
    accent_color: Optional[str] = None
    audience: Optional[List[Any]] = None
    benefits: Optional[List[Any]] = None
    syllabus: Optional[List[Dict[str, Any]]] = None
    is_active: Optional[bool] = None
    display_order: Optional[int] = None


# Course purchase schemas (Part 8b)
class CourseVerifyRequest(BaseModel):
    razorpay_order_id: str
    razorpay_payment_id: str
    razorpay_signature: str
    slug: str

# ============================================================================
# SCHEMAS — Bootcamps
# ============================================================================
class BootcampPayload(BaseModel):
    slug: str
    title: str
    tagline: Optional[str] = None
    description: Optional[str] = None
    category: Optional[str] = None
    difficulty: Optional[str] = None
    modules_count: Optional[int] = None
    labs_count: Optional[int] = None
    duration_hours: Optional[int] = None
    price_inr: Optional[float] = None
    currency: Optional[str] = "INR"
    billing_label: Optional[str] = None
    hero_image_url: Optional[str] = None
    accent_color: Optional[str] = "red"
    audience: Optional[List[Any]] = None
    benefits: Optional[List[Any]] = None
    syllabus: Optional[List[Dict[str, Any]]] = None
    is_active: Optional[bool] = True
    display_order: Optional[int] = 0

class BootcampUpdate(BaseModel):
    slug: Optional[str] = None
    title: Optional[str] = None
    tagline: Optional[str] = None
    description: Optional[str] = None
    category: Optional[str] = None
    difficulty: Optional[str] = None
    modules_count: Optional[int] = None
    labs_count: Optional[int] = None
    duration_hours: Optional[int] = None
    price_inr: Optional[float] = None
    currency: Optional[str] = None
    billing_label: Optional[str] = None
    hero_image_url: Optional[str] = None
    accent_color: Optional[str] = None
    audience: Optional[List[Any]] = None
    benefits: Optional[List[Any]] = None
    syllabus: Optional[List[Dict[str, Any]]] = None
    is_active: Optional[bool] = None
    display_order: Optional[int] = None

class BootcampVerifyRequest(BaseModel):
    razorpay_order_id: str
    razorpay_payment_id: str
    razorpay_signature: str
    slug: str

# Coupon schemas
class CouponPayload(BaseModel):
    code: str
    discount_percent: int
    description: Optional[str] = None
    max_uses: Optional[int] = 1
    expires_at: Optional[str] = None          # ISO datetime; None = never
    is_active: Optional[bool] = True


class CouponUpdate(BaseModel):
    code: Optional[str] = None
    discount_percent: Optional[int] = None
    description: Optional[str] = None
    max_uses: Optional[int] = None
    expires_at: Optional[str] = None
    is_active: Optional[bool] = None


class CouponValidateRequest(BaseModel):
    code: str
    pack_name: Optional[str] = None           # for credit pack validation
    plan_name: Optional[str] = None           # subscription (future)
    billing_cycle: Optional[str] = None       # subscription (future)


class ContactForm(BaseModel):
    name: str
    email: EmailStr
    message: str


class ProgressUpdate(BaseModel):
    user_id: int
    course_name: str
    module_name: str
    
class BootcampProgressUpdate(BaseModel):
    user_id: int
    bootcamp_name: str
    module_name: str


class UpdateProfileRequest(BaseModel):
    full_name: Optional[str] = None
    country: Optional[str] = None
    phone: Optional[str] = None
    newsletter_opt_in: Optional[bool] = None


# ============================================================================
# PYDANTIC SCHEMAS — admin
# ============================================================================
class AdminSignInRequest(BaseModel):
    email: EmailStr
    password: str


class CreateAdminRequest(BaseModel):
    email: EmailStr
    password: str
    full_name: Optional[str] = None
    role: str = "admin"


class LabPayload(BaseModel):
    lab_id: str
    slug: str
    name: str
    category: str
    difficulty: str = "Medium"
    credits_cost: int = 1
    duration_minutes: int = 60
    os_type: str = "linux"
    description: Optional[str] = None
    is_active: bool = True


class LabUpdatePayload(BaseModel):
    slug: Optional[str] = None
    name: Optional[str] = None
    category: Optional[str] = None
    difficulty: Optional[str] = None
    credits_cost: Optional[int] = None
    duration_minutes: Optional[int] = None
    os_type: Optional[str] = None
    description: Optional[str] = None
    is_active: Optional[bool] = None


class CreditAdjustRequest(BaseModel):
    delta: int
    reason: str


class UserStatusRequest(BaseModel):
    is_active: bool
    reason: str


class ContactStatusRequest(BaseModel):
    status: str
    note: Optional[str] = None


VALID_CONTACT_STATUSES = {"unread", "read", "replied", "dismissed"}


class AssessmentAnswerSubmit(BaseModel):
    answers: dict  # {str(question_id): list[int]}


class AssessmentViolationLog(BaseModel):
    attempt_id: int
    violation_type: str = "tab_switch"


# ============================================================================
# HELPERS
# ============================================================================
def generate_referral_code() -> str:
    return "".join(random.choices(string.ascii_uppercase + string.digits, k=8))

# Helper to format bootcamp row
def _bootcamp_row(b: BootcampCatalog) -> dict:
    return {
        "id": b.id,
        "slug": b.slug,
        "title": b.title,
        "tagline": b.tagline,
        "description": b.description,
        "category": b.category,
        "difficulty": b.difficulty,
        "modules_count": b.modules_count,
        "labs_count": b.labs_count,
        "duration_hours": b.duration_hours,
        "price_inr": float(b.price_inr) if b.price_inr is not None else None,
        "currency": b.currency,
        "billing_label": b.billing_label,
        "hero_image_url": b.hero_image_url,
        "accent_color": b.accent_color,
        "audience": b.audience or [],
        "benefits": b.benefits or [],
        "syllabus": b.syllabus or [],
        "is_active": b.is_active,
        "display_order": b.display_order,
        "updated_at": _iso_z(b.updated_at),
    }

# Unified Catalog Route
@app.get("/api/catalog")
def get_unified_catalog(db: Session = Depends(get_db)):
    """Fetches both active courses and bootcamps for the main browse page."""
    courses = db.query(CourseCatalog).filter(CourseCatalog.is_active == True).order_by(CourseCatalog.display_order, CourseCatalog.id).all()
    bootcamps = db.query(BootcampCatalog).filter(BootcampCatalog.is_active == True).order_by(BootcampCatalog.display_order, BootcampCatalog.id).all()
    
    course_data = [{**_course_row(c), "type": "course"} for c in courses]
    bootcamp_data = [{**_bootcamp_row(b), "type": "bootcamp"} for b in bootcamps]
    
    return {
        "courses": course_data,
        "bootcamps": bootcamp_data
    }

# ============================================================================
# USER-FACING ROUTES: AUTH
# ============================================================================
@app.post("/api/auth/signup")
def create_user(user: UserCreate, db: Session = Depends(get_db)):
    if not user.terms:
        raise HTTPException(status_code=400, detail="You must agree to terms")
    if db.query(User).filter(User.email == user.email).first():
        raise HTTPException(status_code=400, detail="Email already registered")

    ref_code = generate_referral_code()
    while db.query(User).filter(User.referral_code == ref_code).first():
        ref_code = generate_referral_code()

    new_user = User(
        email=user.email,
        full_name=user.full_name,
        country=user.country,
        phone=user.phone,
        hashed_password=get_password_hash(user.password),
        newsletter_opt_in=user.newsletter,
        terms_agreed=user.terms,
        referral_code=ref_code,
    )
    db.add(new_user)
    db.commit()
    db.refresh(new_user)

    if user.referred_by_code:
        referrer = db.query(User).filter(User.referral_code == user.referred_by_code).first()
        if referrer:
            # reward_credits=0 means "pending — will be granted on referred user's first purchase"
            referral = Referral(referrer_id=referrer.id, referred_user_id=new_user.id, reward_credits=0)
            db.add(referral)
            db.commit()

    return {"message": "User created successfully", "user_id": new_user.id}


@app.post("/api/auth/signin")
def login(user: UserLogin, db: Session = Depends(get_db)):
    db_user = db.query(User).filter(User.email == user.email).first()
    if not db_user or not verify_password(user.password, db_user.hashed_password):
        raise HTTPException(status_code=401, detail="Invalid email or password")
    if not db_user.is_active:
        raise HTTPException(status_code=403, detail="Account is disabled")

    token_expiry = timedelta(days=30) if user.rememberMe else timedelta(hours=24)
    access_token = create_access_token(
        data={"sub": db_user.email, "id": db_user.id},
        expires_delta=token_expiry,
    )
    return {
        "access_token": access_token,
        "token_type": "bearer",
        "user": {"id": db_user.id, "email": db_user.email},
    }


# ============================================================================
# USER-FACING ROUTES: PROFILE
# ============================================================================
@app.get("/api/users/{user_id}/profile")
def get_profile(user_id: int, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    sub = db.query(Subscription).filter(
        Subscription.user_id == user_id,
        Subscription.is_active == True,
        Subscription.end_date >= datetime.utcnow(),
    ).first()
    return {
        "id": user.id,
        "email": user.email,
        "full_name": user.full_name,
        "country": user.country,
        "phone": user.phone,
        "credits": user.available_credits,
        "referral_code": user.referral_code,
        "subscription_plan": sub.plan_name if sub else None,
        "subscription_billing_cycle": sub.billing_cycle if sub else None,
        "is_subscribed": sub is not None,
        "subscription_end_date": _iso_z(sub.end_date) if sub else None,
        "subscription_auto_renew_status": sub.auto_renew_status if sub else None,
        "subscription_next_charge_at": _iso_z(sub.next_charge_at) if sub else None,
        "subscription_razorpay_id": sub.razorpay_subscription_id if sub else None,
        "newsletter_opt_in": user.newsletter_opt_in,
        "created_at": user.created_at.isoformat() if user.created_at else None,
    }


@app.put("/api/users/{user_id}/profile")
def update_profile(user_id: int, payload: UpdateProfileRequest, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    if payload.full_name is not None:
        user.full_name = payload.full_name
    if payload.country is not None:
        user.country = payload.country
    if payload.phone is not None:
        user.phone = payload.phone
    if payload.newsletter_opt_in is not None:
        user.newsletter_opt_in = payload.newsletter_opt_in
    db.commit()
    return {"message": "Profile updated"}


def _grant_referral_reward_on_first_purchase(user_id: int, db: Session):
    """
    Grant the referrer 10 credits when the referred user makes their first purchase.
    Call this BEFORE adding the new purchase/transaction so the count is accurate.
    reward_credits=0 on the Referral row means the reward is still pending.
    """
    # Only fires on the very first purchase across all payment types
    prior_courses = db.query(CoursePurchase).filter(CoursePurchase.user_id == user_id).count()
    prior_credits = db.query(CreditTransaction).filter(CreditTransaction.user_id == user_id).count()
    prior_bootcamps = db.query(BootcampPurchase).filter(BootcampPurchase.user_id == user_id).count()
    if prior_courses + prior_credits + prior_bootcamps > 0:
        return  # not the first purchase

    referral = (
        db.query(Referral)
        .filter(Referral.referred_user_id == user_id, Referral.reward_credits == 0)
        .first()
    )
    if not referral:
        return  # user wasn't referred, or reward already granted

    referrer = db.query(User).filter(User.id == referral.referrer_id).first()
    if referrer:
        referrer.available_credits += 10
        referral.reward_credits = 10  # mark as rewarded


# ============================================================================
# USER-FACING ROUTES: CREDITS (Razorpay)
# ============================================================================
@app.get("/api/credit-packs")
def list_credit_packs(db: Session = Depends(get_db)):
    """Public catalog — replaces the hardcoded `plans` array in Credit.jsx."""
    packs = (
        db.query(CreditPack)
        .filter(CreditPack.is_active == True)
        .order_by(CreditPack.display_order, CreditPack.id)
        .all()
    )
    return [
        {
            "id": p.id,
            "pack_name": p.pack_name,
            "credits": p.credits,
            "price": float(p.price),
            "display_order": p.display_order,
            "description": p.description,
        }
        for p in packs
    ]


@app.post("/api/credits/create-order")
def create_order(req: CreateOrderRequest, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.id == req.user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    # Server-side defence: a tampered client can't request 100 credits at the 25-credit price.
    # `pack.price` is stored as BASE (excl. GST); customer-facing charge is base × 1.18.
    pack = (
        db.query(CreditPack)
        .filter(CreditPack.pack_name == req.plan_name, CreditPack.is_active == True)
        .first()
    )
    if not pack or pack.credits != req.credits:
        raise HTTPException(status_code=400, detail="Pack mismatch")

    # Apply coupon if supplied. Discount is on base; GST is recomputed.
    discount_percent = 0
    if req.coupon_code:
        coupon = _lookup_coupon(db, req.coupon_code, req.user_id)
        discount_percent = coupon.discount_percent
    discounted_base = _apply_coupon_to_base(float(pack.price), discount_percent)
    expected_paise = with_gst(discounted_base) * 100
    if expected_paise != req.amount:
        raise HTTPException(status_code=400, detail="Pack mismatch")

    notes = {
        "user_id": str(req.user_id),
        "plan_name": req.plan_name,
        "credits": str(req.credits),
    }
    if req.coupon_code:
        notes["coupon_code"] = req.coupon_code.strip().upper()
    order = razorpay_client.order.create({
        "amount": req.amount,
        "currency": "INR",
        "receipt": f"user_{req.user_id}_{req.plan_name}",
        "notes": notes,
    })
    return {"order_id": order["id"], "amount": req.amount, "currency": "INR"}


@app.post("/api/coupons/validate")
def validate_coupon(
    req: CouponValidateRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Returns the priced-out preview for a coupon + pack combination.

    Used by the /credit checkout to show 'after discount' totals before the
    user clicks Purchase. Pure read; no state change. Same validation rules
    are re-applied at order-create time, so a stale preview can't be exploited.
    """
    coupon = _lookup_coupon(db, req.code, current_user.id)

    # Pack-level preview (subscription support is a future follow-up).
    if req.pack_name:
        pack = (
            db.query(CreditPack)
            .filter(CreditPack.pack_name == req.pack_name, CreditPack.is_active == True)
            .first()
        )
        if not pack:
            raise HTTPException(status_code=404, detail="Pack not found")
        base = float(pack.price)
        discounted_base = _apply_coupon_to_base(base, coupon.discount_percent)
        original_total = with_gst(base)
        new_total = with_gst(discounted_base)
        return {
            "valid": True,
            "coupon": {
                "code": coupon.code,
                "discount_percent": coupon.discount_percent,
                "description": coupon.description,
            },
            "pack": {
                "pack_name": pack.pack_name,
                "credits": pack.credits,
                "base": base,
                "original_total": original_total,
            },
            "preview": {
                "discounted_base": round(discounted_base, 2),
                "new_total": new_total,
                "discount_amount": original_total - new_total,
            },
        }

    # Generic validation (no target pack/plan supplied) — useful when frontend
    # only wants to confirm the code exists and grab the discount %.
    return {
        "valid": True,
        "coupon": {
            "code": coupon.code,
            "discount_percent": coupon.discount_percent,
            "description": coupon.description,
        },
    }


@app.get("/api/credits/history")
def get_credit_history(
    type: str = Query("all", regex="^(all|credit|subscription)$"),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """User-visible purchase history. `type=credit` excludes subscription receipts."""
    q = db.query(CreditTransaction).filter(CreditTransaction.user_id == current_user.id)
    if type == "credit":
        q = q.filter(~CreditTransaction.plan_name.like("sub:%"))
    elif type == "subscription":
        q = q.filter(CreditTransaction.plan_name.like("sub:%"))
    rows = q.order_by(CreditTransaction.purchased_at.desc()).limit(200).all()
    return [
        {
            "id": r.id,
            "plan_name": r.plan_name,
            "credits_added": r.credits_added,
            "amount_paid": float(r.amount_paid),
            "purchased_at": _iso_z(r.purchased_at),
            "type": "subscription" if (r.plan_name or "").startswith("sub:") else "credit",
        }
        for r in rows
    ]


@app.post("/api/credits/verify-payment")
def verify_and_credit(req: VerifyPaymentRequest, db: Session = Depends(get_db)):
    body = f"{req.razorpay_order_id}|{req.razorpay_payment_id}"
    expected_signature = hmac.new(
        key=os.getenv("RAZORPAY_KEY_SECRET", "").encode("utf-8"),
        msg=body.encode("utf-8"),
        digestmod=hashlib.sha256,
    ).hexdigest()
    if expected_signature != req.razorpay_signature:
        raise HTTPException(status_code=400, detail="Invalid payment signature")

    user = db.query(User).filter(User.id == req.user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    _grant_referral_reward_on_first_purchase(user.id, db)
    user.available_credits += req.credits
    transaction = CreditTransaction(
        user_id=req.user_id,
        plan_name=req.plan_name,
        credits_added=req.credits,
        amount_paid=req.amount,
    )
    db.add(transaction)
    db.flush()  # need transaction.id for the redemption FK

    # Record coupon redemption if the order was placed with one.
    # We trust the verified signature: the order_id was signed by Razorpay against
    # the amount we created the order with, so coupon_code arriving here matches
    # the discount that was actually applied to the charge.
    if req.coupon_code:
        coupon = (
            db.query(Coupon)
            .filter(Coupon.code == req.coupon_code.strip().upper())
            .first()
        )
        if coupon:
            already = (
                db.query(CouponRedemption)
                .filter(
                    CouponRedemption.coupon_id == coupon.id,
                    CouponRedemption.user_id == req.user_id,
                )
                .first()
            )
            if not already:
                # Discount value = (full inclusive price for this pack) - (paid inclusive amount).
                pack = (
                    db.query(CreditPack)
                    .filter(CreditPack.pack_name == req.plan_name)
                    .first()
                )
                full_total = with_gst(float(pack.price)) if pack else 0
                discount_amount = max(0.0, float(full_total) - float(req.amount))
                db.add(CouponRedemption(
                    coupon_id=coupon.id,
                    user_id=req.user_id,
                    credit_transaction_id=transaction.id,
                    discount_amount=discount_amount,
                ))
                coupon.uses_count = (coupon.uses_count or 0) + 1
    db.commit()
    return {"message": f"{req.credits} credits added", "new_balance": user.available_credits}


# ============================================================================
# USER-FACING ROUTES: SUBSCRIPTIONS
# ============================================================================
@app.post("/api/subscriptions/create-order")
def create_subscription_order(req: SubCreateOrder, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.id == req.user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    order = razorpay_client.order.create({
        "amount": req.amount,
        "currency": "INR",
        "receipt": f"sub_{req.user_id}_{req.plan_name}_{req.billing_cycle}",
        "notes": {
            "user_id": str(req.user_id),
            "plan_name": req.plan_name,
            "billing_cycle": req.billing_cycle,
            "type": "subscription",
        },
    })
    return {"order_id": order["id"], "amount": req.amount, "currency": "INR"}


@app.post("/api/subscriptions/verify-payment")
def verify_subscription_payment(req: SubVerifyPayment, db: Session = Depends(get_db)):
    body = f"{req.razorpay_order_id}|{req.razorpay_payment_id}"
    expected_signature = hmac.new(
        key=os.getenv("RAZORPAY_KEY_SECRET", "").encode("utf-8"),
        msg=body.encode("utf-8"),
        digestmod=hashlib.sha256,
    ).hexdigest()
    if expected_signature != req.razorpay_signature:
        raise HTTPException(status_code=400, detail="Invalid payment signature")

    user = db.query(User).filter(User.id == req.user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    db.query(Subscription).filter(
        Subscription.user_id == req.user_id, Subscription.is_active == True
    ).update({"is_active": False})

    start_date = datetime.utcnow()
    end_date = start_date + (
        timedelta(days=365) if req.billing_cycle == "annual" else timedelta(days=30)
    )
    new_sub = Subscription(
        user_id=user.id,
        plan_name=req.plan_name,
        billing_cycle=req.billing_cycle,
        start_date=start_date,
        end_date=end_date,
    )
    db.add(new_sub)

    txn = CreditTransaction(
        user_id=req.user_id,
        plan_name=f"sub:{req.plan_name}",
        credits_added=0,
        amount_paid=req.amount,
    )
    db.add(txn)
    db.commit()
    return {"message": "Subscription activated successfully", "expires_on": end_date.isoformat()}


@app.post("/api/subscriptions/subscribe")
def create_subscription_legacy(req: SubscribePlan, db: Session = Depends(get_db)):
    """Legacy free-tier path (no payment). Kept for backwards compat."""
    user = db.query(User).filter(User.id == req.user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    start_date = datetime.utcnow()
    end_date = start_date + (
        timedelta(days=365) if req.billing_cycle == "annual" else timedelta(days=30)
    )
    db.query(Subscription).filter(
        Subscription.user_id == req.user_id, Subscription.is_active == True
    ).update({"is_active": False})
    new_sub = Subscription(
        user_id=user.id, plan_name=req.plan_name,
        billing_cycle=req.billing_cycle, start_date=start_date, end_date=end_date,
    )
    db.add(new_sub)
    db.commit()
    return {"message": "Subscription activated successfully", "expires_on": end_date.isoformat()}


# ----------------------------------------------------------------------------
# RECURRING SUBSCRIPTIONS (Razorpay Subscriptions API)
# ----------------------------------------------------------------------------
def _from_rzp_ts(ts):
    """Razorpay returns epoch seconds (int) in webhook payloads."""
    if ts in (None, 0, ""):
        return None
    try:
        return datetime.fromtimestamp(int(ts), tz=timezone.utc)
    except (ValueError, TypeError, OSError):
        return None


@app.post("/api/subscriptions/create-subscription")
def create_recurring_subscription(req: SubCreateRecurring, db: Session = Depends(get_db)):
    """Create a Razorpay subscription and return its short_url for the JS SDK."""
    plan_id = RAZORPAY_PLAN_IDS.get((req.plan_name, req.billing_cycle))
    if not plan_id:
        raise HTTPException(status_code=400, detail="Unknown plan or billing cycle")
    user = db.query(User).filter(User.id == req.user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    total_count = RAZORPAY_TOTAL_COUNT.get(req.billing_cycle, 60)
    sub = razorpay_client.subscription.create({
        "plan_id": plan_id,
        "total_count": total_count,
        "customer_notify": 1,
        "notes": {
            "user_id": str(req.user_id),
            "plan_name": req.plan_name,
            "billing_cycle": req.billing_cycle,
        },
    })
    # Pending row — fills in on subscription.activated webhook OR verify-initial-payment callback.
    pending = Subscription(
        user_id=user.id,
        plan_name=req.plan_name,
        billing_cycle=req.billing_cycle,
        start_date=datetime.utcnow(),
        end_date=datetime.utcnow() + timedelta(days=1),  # placeholder; activated webhook overwrites
        is_active=False,
        auto_renew_status="pending",
        razorpay_subscription_id=sub["id"],
    )
    db.add(pending)
    db.commit()
    return {
        "subscription_id": sub["id"],
        "short_url": sub.get("short_url"),
        "status": sub.get("status"),
    }


@app.post("/api/subscriptions/verify-initial-payment")
def verify_initial_subscription_payment(req: SubVerifyInitial, db: Session = Depends(get_db)):
    """JS SDK callback after the user authorises the mandate / first charge.

    NOTE: signature format for Subscriptions API is `payment_id|subscription_id`,
    NOT `order_id|payment_id` like one-shot orders. Wrong concat = signature mismatch.
    """
    body = f"{req.razorpay_payment_id}|{req.razorpay_subscription_id}"
    expected_signature = hmac.new(
        key=os.getenv("RAZORPAY_KEY_SECRET", "").encode("utf-8"),
        msg=body.encode("utf-8"),
        digestmod=hashlib.sha256,
    ).hexdigest()
    if expected_signature != req.razorpay_signature:
        raise HTTPException(status_code=400, detail="Invalid payment signature")

    row = (
        db.query(Subscription)
        .filter(Subscription.razorpay_subscription_id == req.razorpay_subscription_id)
        .first()
    )
    if not row or row.user_id != req.user_id:
        raise HTTPException(status_code=404, detail="Subscription not found")

    # Pull live state from Razorpay (authoritative for dates).
    try:
        live = razorpay_client.subscription.fetch(req.razorpay_subscription_id)
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Razorpay fetch failed: {e}")

    current_end = _from_rzp_ts(live.get("current_end"))
    charge_at = _from_rzp_ts(live.get("charge_at"))

    # Deactivate any prior active subs for this user (mirrors existing pattern).
    db.query(Subscription).filter(
        Subscription.user_id == req.user_id,
        Subscription.is_active == True,
        Subscription.id != row.id,
    ).update({"is_active": False, "auto_renew_status": "cancelled"})

    row.is_active = True
    row.auto_renew_status = "active"
    row.start_date = datetime.utcnow()
    row.current_period_end = current_end
    row.end_date = current_end or (datetime.utcnow() + timedelta(days=30))
    row.next_charge_at = charge_at
    row.razorpay_customer_id = live.get("customer_id")

    # Match existing admin-payments accounting: every sub event gets a CreditTransaction.
    txn = CreditTransaction(
        user_id=req.user_id,
        plan_name=f"sub:{row.plan_name}",
        credits_added=0,
        amount_paid=0,  # initial mandate may be a ₹0 token; real charges write via webhook.
    )
    db.add(txn)
    db.commit()
    return {
        "message": "Subscription activated",
        "expires_on": _iso_z(row.end_date),
        "auto_renew_status": row.auto_renew_status,
    }


@app.post("/api/subscriptions/cancel")
def cancel_subscription(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Manual cancel — Razorpay stops charging at cycle end. We keep access until end_date."""
    row = (
        db.query(Subscription)
        .filter(
            Subscription.user_id == current_user.id,
            Subscription.is_active == True,
            Subscription.auto_renew_status.in_(["active", "halted"]),
        )
        .order_by(Subscription.id.desc())
        .first()
    )
    if not row or not row.razorpay_subscription_id:
        raise HTTPException(status_code=404, detail="No cancellable subscription found")
    try:
        razorpay_client.subscription.cancel(
            row.razorpay_subscription_id, {"cancel_at_cycle_end": 1}
        )
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Razorpay cancel failed: {e}")
    row.auto_renew_status = "cancelled_pending"
    db.commit()
    return {
        "cancelled": True,
        "access_until": _iso_z(row.end_date),
        "auto_renew_status": row.auto_renew_status,
    }


@app.post("/api/subscriptions/upgrade")
def upgrade_subscription(
    req: SubUpgradeRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Instant switch: cancel current Razorpay sub immediately, create new one, charge full new plan."""
    new_plan_id = RAZORPAY_PLAN_IDS.get((req.new_plan, req.billing_cycle))
    if not new_plan_id:
        raise HTTPException(status_code=400, detail="Unknown new plan or billing cycle")
    current = (
        db.query(Subscription)
        .filter(
            Subscription.user_id == current_user.id,
            Subscription.is_active == True,
        )
        .order_by(Subscription.id.desc())
        .first()
    )
    if not current:
        raise HTTPException(status_code=404, detail="No active subscription to upgrade")
    if current.plan_name == req.new_plan and current.billing_cycle == req.billing_cycle:
        raise HTTPException(status_code=400, detail="Already on this plan and cycle")

    # Best-effort cancel old sub immediately. Don't block on Razorpay errors (some legacy
    # rows may not have a razorpay_subscription_id — those are pure one-shot subs).
    if current.razorpay_subscription_id:
        try:
            razorpay_client.subscription.cancel(
                current.razorpay_subscription_id, {"cancel_at_cycle_end": 0}
            )
        except Exception:
            pass
    current.is_active = False
    current.auto_renew_status = "cancelled"
    current.end_date = datetime.utcnow()
    db.commit()

    # Create the new recurring subscription (same path as create-recurring).
    total_count = RAZORPAY_TOTAL_COUNT.get(req.billing_cycle, 60)
    new_sub = razorpay_client.subscription.create({
        "plan_id": new_plan_id,
        "total_count": total_count,
        "customer_notify": 1,
        "notes": {
            "user_id": str(current_user.id),
            "plan_name": req.new_plan,
            "billing_cycle": req.billing_cycle,
            "upgrade_from": current.plan_name,
        },
    })
    pending = Subscription(
        user_id=current_user.id,
        plan_name=req.new_plan,
        billing_cycle=req.billing_cycle,
        start_date=datetime.utcnow(),
        end_date=datetime.utcnow() + timedelta(days=1),
        is_active=False,
        auto_renew_status="pending",
        razorpay_subscription_id=new_sub["id"],
    )
    db.add(pending)
    db.commit()
    return {
        "subscription_id": new_sub["id"],
        "short_url": new_sub.get("short_url"),
        "upgraded_from": current.plan_name,
    }


# ----------------------------------------------------------------------------
# RAZORPAY WEBHOOK
# ----------------------------------------------------------------------------
def _apply_subscription_event(db: Session, evt: dict) -> None:
    """Dispatch a verified webhook event to subscription state changes."""
    event_type = evt.get("event", "")
    payload = evt.get("payload", {}) or {}
    sub_entity = (payload.get("subscription") or {}).get("entity") or {}
    sub_id = sub_entity.get("id")
    if not sub_id:
        return
    row = (
        db.query(Subscription)
        .filter(Subscription.razorpay_subscription_id == sub_id)
        .first()
    )
    if not row:
        return  # Webhook for a sub we never created — ignore (test events, leaked plan, etc.).

    current_end = _from_rzp_ts(sub_entity.get("current_end"))
    charge_at = _from_rzp_ts(sub_entity.get("charge_at"))

    if event_type in ("subscription.activated", "subscription.charged"):
        row.is_active = True
        row.auto_renew_status = "active"
        if current_end:
            row.current_period_end = current_end
            row.end_date = current_end
        if charge_at:
            row.next_charge_at = charge_at
        # On a recurring charge, log the payment for the admin payments view.
        if event_type == "subscription.charged":
            payment_entity = (payload.get("payment") or {}).get("entity") or {}
            amount_paise = payment_entity.get("amount") or 0
            txn = CreditTransaction(
                user_id=row.user_id,
                plan_name=f"sub:{row.plan_name}",
                credits_added=0,
                amount_paid=float(amount_paise) / 100.0,
            )
            db.add(txn)
    elif event_type == "subscription.halted":
        # Recurring charge failed past Razorpay's retry window. Keep access until end_date;
        # the daily reconcile sweep deactivates when end_date passes.
        row.auto_renew_status = "halted"
    elif event_type == "subscription.cancelled":
        row.auto_renew_status = "cancelled"
        # `is_active=False` happens automatically when reconcile_labs.py sweeps past end_date.
    elif event_type == "subscription.completed":
        row.auto_renew_status = "completed"
    # subscription.pending and payment.failed are logged via the event row but require no state change.


@app.post("/api/webhooks/razorpay")
async def razorpay_webhook(request: Request, db: Session = Depends(get_db)):
    """Razorpay subscription event firehose. Webhook secret comes from env."""
    raw = await request.body()
    sig = request.headers.get("X-Razorpay-Signature", "")
    if not RAZORPAY_WEBHOOK_SECRET:
        raise HTTPException(status_code=500, detail="Webhook secret not configured")
    expected = hmac.new(
        RAZORPAY_WEBHOOK_SECRET.encode("utf-8"), raw, hashlib.sha256
    ).hexdigest()
    if not hmac.compare_digest(expected, sig):
        raise HTTPException(status_code=400, detail="Invalid webhook signature")

    try:
        evt = _json.loads(raw.decode("utf-8"))
    except Exception:
        raise HTTPException(status_code=400, detail="Malformed JSON")

    # Build an event-id from whatever Razorpay sends. Their docs say events carry an `id`
    # field at the top level; fall back to `{subscription_id}:{event}:{created_at}` so a
    # missing id still hits the unique constraint reliably.
    sub_entity = ((evt.get("payload") or {}).get("subscription") or {}).get("entity") or {}
    sub_id_for_evt = sub_entity.get("id") or "unknown"
    event_id = evt.get("id") or f"{sub_id_for_evt}:{evt.get('event','')}:{evt.get('created_at','')}"

    try:
        db.add(SubscriptionEvent(
            razorpay_event_id=event_id,
            razorpay_subscription_id=sub_id_for_evt,
            event_type=evt.get("event", "unknown"),
            payload=evt,
        ))
        db.commit()
    except IntegrityError:
        # Duplicate POST — Razorpay retries on non-2xx. Already processed.
        db.rollback()
        return {"ok": True, "dup": True}

    try:
        _apply_subscription_event(db, evt)
        db.query(SubscriptionEvent).filter(
            SubscriptionEvent.razorpay_event_id == event_id
        ).update({"processed": True})
        db.commit()
    except Exception:
        # Don't fail the webhook — Razorpay would retry and re-trigger side effects.
        # The unprocessed row is picked up by the reconcile sweep as a manual review hint.
        db.rollback()
    return {"ok": True}


# ============================================================================
# USER-FACING ROUTES: CONTACT
# ============================================================================
@app.post("/api/contact")
def submit_contact_form(form: ContactForm, db: Session = Depends(get_db)):
    new_message = ContactMessage(name=form.name, email=form.email, message=form.message)
    db.add(new_message)
    db.commit()
    return {"message": "Thank you! We have received your message."}


# ============================================================================
# USER-FACING ROUTES: REFERRALS
# ============================================================================
@app.get("/api/referrals/{user_id}")
def get_referral_stats(user_id: int, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    total_referrals = db.query(Referral).filter(Referral.referrer_id == user_id).count()
    rewards = (
        db.query(Referral)
        .filter(Referral.referrer_id == user_id)
        .with_entities(Referral.reward_credits)
        .all()
    )
    return {
        "referral_code": user.referral_code,
        "total_successful_referrals": total_referrals,
        "total_credits_earned": sum(c[0] or 0 for c in rewards),
    }


# ============================================================================
# USER-FACING ROUTES: LEARNING
# ============================================================================
# All four progress endpoints require a valid user JWT and derive the user_id
# from the token rather than trusting the request body / path param. The
# user_id field on the request payload (and path) is kept for backwards-compat
# with the existing frontend but is cross-checked against current_user.id and
# rejected on mismatch — prevents IDOR (writing/reading another user's progress).

@app.post("/api/learning/progress")
def mark_lesson_complete(
    req: ProgressUpdate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if req.user_id != current_user.id:
        raise HTTPException(status_code=403, detail="Cannot record progress for another user")
    existing = db.query(UserProgress).filter_by(
        user_id=current_user.id, course_name=req.course_name, module_name=req.module_name
    ).first()
    if not existing:
        progress = UserProgress(
            user_id=current_user.id, course_name=req.course_name, module_name=req.module_name
        )
        db.add(progress)
        db.commit()
    return {"message": "Progress saved"}


@app.get("/api/learning/progress/{user_id}")
def get_user_progress(
    user_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if user_id != current_user.id:
        raise HTTPException(status_code=403, detail="Cannot view another user's progress")
    progress = db.query(UserProgress).filter(UserProgress.user_id == current_user.id).all()
    return {
        "completed_modules": [
            {"course": p.course_name, "module": p.module_name} for p in progress
        ]
    }


@app.post("/api/learning/bootcamp-progress")
def mark_bootcamp_lesson_complete(
    req: BootcampProgressUpdate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if req.user_id != current_user.id:
        raise HTTPException(status_code=403, detail="Cannot record progress for another user")
    existing = db.query(BootcampProgress).filter_by(
        user_id=current_user.id, bootcamp_name=req.bootcamp_name, module_name=req.module_name
    ).first()
    if not existing:
        progress = BootcampProgress(
            user_id=current_user.id, bootcamp_name=req.bootcamp_name, module_name=req.module_name
        )
        db.add(progress)
        db.commit()
    return {"message": "Bootcamp progress saved"}


@app.get("/api/learning/bootcamp-progress/{user_id}")
def get_user_bootcamp_progress(
    user_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if user_id != current_user.id:
        raise HTTPException(status_code=403, detail="Cannot view another user's progress")
    progress = db.query(BootcampProgress).filter(BootcampProgress.user_id == current_user.id).all()
    return {
        "completed_modules": [{"bootcamp": p.bootcamp_name, "module": p.module_name} for p in progress]
    }


# ============================================================================
# ADMIN ROUTES: AUTH
# ============================================================================
def _admin_to_dict(a: Admin) -> dict:
    return {
        "id": a.id,
        "email": a.email,
        "full_name": a.full_name,
        "role": a.role,
        "is_active": a.is_active,
        "created_at": a.created_at.isoformat() if a.created_at else None,
        "last_login_at": a.last_login_at.isoformat() if a.last_login_at else None,
    }


@app.post("/api/admin/auth/signin")
def admin_signin(req: AdminSignInRequest, db: Session = Depends(get_db)):
    admin = db.query(Admin).filter(Admin.email == req.email).first()
    if not admin or not verify_password(req.password, admin.hashed_password):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    if not admin.is_active:
        raise HTTPException(status_code=403, detail="Account is disabled")

    admin.last_login_at = datetime.utcnow()
    db.commit()
    db.refresh(admin)

    token = create_admin_token(admin_id=admin.id, role=admin.role, email=admin.email)
    return {"access_token": token, "admin": _admin_to_dict(admin)}


@app.get("/api/admin/auth/me")
def admin_me(admin: Admin = Depends(get_current_admin)):
    return _admin_to_dict(admin)


@app.get("/api/admin/auth/admins")
def list_admins(_: Admin = Depends(require_superadmin), db: Session = Depends(get_db)):
    admins = db.query(Admin).order_by(Admin.created_at.desc()).all()
    return [_admin_to_dict(a) for a in admins]


@app.post("/api/admin/auth/admins")
def create_admin(
    req: CreateAdminRequest,
    superadmin: Admin = Depends(require_superadmin),
    db: Session = Depends(get_db),
):
    if req.role not in ("admin", "superadmin"):
        raise HTTPException(status_code=400, detail="Invalid role")
    if len(req.password) < 8:
        raise HTTPException(status_code=400, detail="Password must be at least 8 characters")
    if db.query(Admin).filter(Admin.email == req.email).first():
        raise HTTPException(status_code=409, detail="Email already in use")

    new_admin = Admin(
        email=req.email,
        hashed_password=get_password_hash(req.password),
        full_name=req.full_name,
        role=req.role,
    )
    db.add(new_admin)
    db.commit()
    db.refresh(new_admin)

    write_audit(
        db=db,
        admin_id=superadmin.id,
        action="create",
        entity_type="admin",
        entity_id=new_admin.id,
        after={"email": req.email, "full_name": req.full_name, "role": req.role},
    )
    return _admin_to_dict(new_admin)


# ============================================================================
# ADMIN ROUTES: LABS
# ============================================================================
def _lab_to_dict(lab: LabCatalog) -> dict:
    return {
        "lab_id": lab.lab_id,
        "slug": lab.slug,
        "name": lab.name,
        "category": lab.category,
        "difficulty": lab.difficulty,
        "credits_cost": lab.credits_cost,
        "duration_minutes": lab.duration_minutes,
        "os_type": lab.os_type,
        "description": lab.description,
        "is_active": lab.is_active,
        "created_at": lab.created_at.isoformat() if lab.created_at else None,
        "updated_at": lab.updated_at.isoformat() if lab.updated_at else None,
    }


@app.get("/api/admin/labs")
def admin_list_labs(_: Admin = Depends(get_current_admin), db: Session = Depends(get_db)):
    labs = db.query(LabCatalog).order_by(LabCatalog.category, LabCatalog.lab_id).all()
    return [_lab_to_dict(l) for l in labs]


@app.get("/api/admin/labs/{lab_id}")
def admin_get_lab(lab_id: str, _: Admin = Depends(get_current_admin), db: Session = Depends(get_db)):
    lab = db.query(LabCatalog).filter(LabCatalog.lab_id == lab_id).first()
    if not lab:
        raise HTTPException(status_code=404, detail="Lab not found")
    return _lab_to_dict(lab)


@app.post("/api/admin/labs")
def admin_create_lab(
    payload: LabPayload,
    admin: Admin = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    if db.query(LabCatalog).filter(LabCatalog.lab_id == payload.lab_id).first():
        raise HTTPException(status_code=409, detail="lab_id already exists")
    if payload.difficulty not in ("Easy", "Medium", "Hard"):
        raise HTTPException(status_code=400, detail="Invalid difficulty")
    if payload.os_type not in ("windows", "linux", "mixed"):
        raise HTTPException(status_code=400, detail="Invalid os_type")

    lab = LabCatalog(**payload.dict())
    db.add(lab)
    db.commit()
    db.refresh(lab)

    write_audit(
        db=db, admin_id=admin.id, action="create", entity_type="lab",
        entity_id=payload.lab_id, after=payload.dict(),
    )
    return _lab_to_dict(lab)


@app.put("/api/admin/labs/{lab_id}")
def admin_update_lab(
    lab_id: str,
    payload: LabUpdatePayload,
    admin: Admin = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    lab = db.query(LabCatalog).filter(LabCatalog.lab_id == lab_id).first()
    if not lab:
        raise HTTPException(status_code=404, detail="Lab not found")

    before = _lab_to_dict(lab)
    updates = {k: v for k, v in payload.dict().items() if v is not None}

    if "difficulty" in updates and updates["difficulty"] not in ("Easy", "Medium", "Hard"):
        raise HTTPException(status_code=400, detail="Invalid difficulty")
    if "os_type" in updates and updates["os_type"] not in ("windows", "linux", "mixed"):
        raise HTTPException(status_code=400, detail="Invalid os_type")

    for k, v in updates.items():
        setattr(lab, k, v)
    db.commit()
    db.refresh(lab)

    after = _lab_to_dict(lab)
    write_audit(
        db=db, admin_id=admin.id, action="update", entity_type="lab",
        entity_id=lab_id, before=before, after=after,
    )
    return after


@app.delete("/api/admin/labs/{lab_id}")
def admin_delete_lab(
    lab_id: str,
    admin: Admin = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    lab = db.query(LabCatalog).filter(LabCatalog.lab_id == lab_id).first()
    if not lab:
        raise HTTPException(status_code=404, detail="Lab not found")
    before = {"is_active": lab.is_active}
    lab.is_active = False
    db.commit()

    write_audit(
        db=db, admin_id=admin.id, action="delete", entity_type="lab",
        entity_id=lab_id, before=before, after={"is_active": False},
    )
    return {"success": True}


# ============================================================================
# ADMIN ROUTES: CREDIT PACKS
# ============================================================================
def _pack_row(p: CreditPack) -> dict:
    return {
        "id": p.id,
        "pack_name": p.pack_name,
        "credits": p.credits,
        "price": float(p.price),
        "is_active": p.is_active,
        "display_order": p.display_order,
        "description": p.description,
        "updated_at": _iso_z(p.updated_at),
    }


@app.get("/api/admin/credit-packs")
def admin_list_credit_packs(
    admin: Admin = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    """Admin sees ALL packs (active + soft-deleted)."""
    rows = db.query(CreditPack).order_by(CreditPack.display_order, CreditPack.id).all()
    return [_pack_row(p) for p in rows]


@app.post("/api/admin/credit-packs")
def admin_create_credit_pack(
    payload: CreditPackPayload,
    admin: Admin = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    if db.query(CreditPack).filter(CreditPack.pack_name == payload.pack_name).first():
        raise HTTPException(status_code=409, detail="Pack name already exists")
    pack = CreditPack(
        pack_name=payload.pack_name,
        credits=payload.credits,
        price=payload.price,
        is_active=payload.is_active if payload.is_active is not None else True,
        display_order=payload.display_order or 0,
        description=payload.description,
    )
    db.add(pack)
    db.commit()
    db.refresh(pack)
    after = _pack_row(pack)
    write_audit(
        db=db, admin_id=admin.id, action="create", entity_type="credit_pack",
        entity_id=pack.id, before=None, after=after,
    )
    return after


@app.put("/api/admin/credit-packs/{pack_id}")
def admin_update_credit_pack(
    pack_id: int,
    payload: CreditPackUpdate,
    admin: Admin = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    pack = db.query(CreditPack).filter(CreditPack.id == pack_id).first()
    if not pack:
        raise HTTPException(status_code=404, detail="Pack not found")
    before = _pack_row(pack)
    data = payload.dict(exclude_unset=True)
    # Guard against renaming into a collision.
    new_name = data.get("pack_name")
    if new_name and new_name != pack.pack_name:
        if db.query(CreditPack).filter(CreditPack.pack_name == new_name).first():
            raise HTTPException(status_code=409, detail="Pack name already exists")
    for field, value in data.items():
        setattr(pack, field, value)
    db.commit()
    after = _pack_row(pack)
    write_audit(
        db=db, admin_id=admin.id, action="update", entity_type="credit_pack",
        entity_id=pack.id, before=before, after=after,
    )
    return after


@app.delete("/api/admin/credit-packs/{pack_id}")
def admin_delete_credit_pack(
    pack_id: int,
    admin: Admin = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    """Soft delete — keeps historical CreditTransaction.plan_name references readable."""
    pack = db.query(CreditPack).filter(CreditPack.id == pack_id).first()
    if not pack:
        raise HTTPException(status_code=404, detail="Pack not found")
    before = _pack_row(pack)
    pack.is_active = False
    db.commit()
    write_audit(
        db=db, admin_id=admin.id, action="delete", entity_type="credit_pack",
        entity_id=pack.id, before=before, after={"is_active": False},
    )
    return {"success": True}


# ============================================================================
# ADMIN ROUTES: COURSE CATALOG (Part 8)
# ============================================================================
# Mirrors the credit_packs CRUD pattern: soft-delete via is_active, audit-logged,
# slug-collision guard on create + rename. Public reads go through /api/courses
# (active rows only); admin reads return everything including soft-deleted.

@app.get("/api/admin/courses")
def admin_list_courses(
    admin: Admin = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    """Admin sees ALL courses (active + soft-deleted)."""
    rows = (
        db.query(CourseCatalog)
        .order_by(CourseCatalog.display_order, CourseCatalog.id)
        .all()
    )
    return [_course_row(c) for c in rows]


@app.get("/api/admin/courses/{course_id}")
def admin_get_course(
    course_id: int,
    admin: Admin = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    course = db.query(CourseCatalog).filter(CourseCatalog.id == course_id).first()
    if not course:
        raise HTTPException(status_code=404, detail="Course not found")
    return _course_row(course)


@app.post("/api/admin/courses")
def admin_create_course(
    payload: CoursePayload,
    admin: Admin = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    if db.query(CourseCatalog).filter(CourseCatalog.slug == payload.slug).first():
        raise HTTPException(status_code=409, detail="Course slug already exists")
    course = CourseCatalog(
        slug=payload.slug,
        title=payload.title,
        tagline=payload.tagline,
        description=payload.description,
        category=payload.category,
        difficulty=payload.difficulty,
        modules_count=payload.modules_count,
        labs_count=payload.labs_count,
        duration_hours=payload.duration_hours,
        price_inr=payload.price_inr,
        currency=payload.currency or "INR",
        billing_label=payload.billing_label,
        hero_image_url=payload.hero_image_url,
        accent_color=payload.accent_color or "red",
        audience=payload.audience or [],
        benefits=payload.benefits or [],
        syllabus=payload.syllabus or [],
        is_active=payload.is_active if payload.is_active is not None else True,
        display_order=payload.display_order or 0,
    )
    db.add(course)
    db.commit()
    db.refresh(course)
    after = _course_row(course)
    write_audit(
        db=db, admin_id=admin.id, action="create", entity_type="course",
        entity_id=course.id, before=None, after=after,
    )
    return after


@app.put("/api/admin/courses/{course_id}")
def admin_update_course(
    course_id: int,
    payload: CourseUpdate,
    admin: Admin = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    course = db.query(CourseCatalog).filter(CourseCatalog.id == course_id).first()
    if not course:
        raise HTTPException(status_code=404, detail="Course not found")
    before = _course_row(course)
    data = payload.dict(exclude_unset=True)
    # Guard against renaming a slug into a collision.
    new_slug = data.get("slug")
    if new_slug and new_slug != course.slug:
        if db.query(CourseCatalog).filter(CourseCatalog.slug == new_slug).first():
            raise HTTPException(status_code=409, detail="Course slug already exists")
    for field, value in data.items():
        setattr(course, field, value)
    db.commit()
    after = _course_row(course)
    write_audit(
        db=db, admin_id=admin.id, action="update", entity_type="course",
        entity_id=course.id, before=before, after=after,
    )
    return after


@app.delete("/api/admin/courses/{course_id}")
def admin_delete_course(
    course_id: int,
    admin: Admin = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    """Soft delete so audit-log references stay readable."""
    course = db.query(CourseCatalog).filter(CourseCatalog.id == course_id).first()
    if not course:
        raise HTTPException(status_code=404, detail="Course not found")
    before = _course_row(course)
    course.is_active = False
    db.commit()
    write_audit(
        db=db, admin_id=admin.id, action="delete", entity_type="course",
        entity_id=course.id, before=before, after={"is_active": False},
    )
    return {"success": True}

#ADMIN ROUTES: BOOTCAMPS CATALOG
# ============================================================================
# ADMIN ROUTES: BOOTCAMP CATALOG
# ============================================================================
@app.get("/api/admin/bootcamps")
def admin_list_bootcamps(admin: Admin = Depends(get_current_admin), db: Session = Depends(get_db)):
    rows = db.query(BootcampCatalog).order_by(BootcampCatalog.display_order, BootcampCatalog.id).all()
    return [_bootcamp_row(b) for b in rows]

@app.get("/api/admin/bootcamps/{bootcamp_id}")
def admin_get_bootcamp(bootcamp_id: int, admin: Admin = Depends(get_current_admin), db: Session = Depends(get_db)):
    bootcamp = db.query(BootcampCatalog).filter(BootcampCatalog.id == bootcamp_id).first()
    if not bootcamp:
        raise HTTPException(status_code=404, detail="Bootcamp not found")
    return _bootcamp_row(bootcamp)

@app.post("/api/admin/bootcamps")
def admin_create_bootcamp(payload: BootcampPayload, admin: Admin = Depends(get_current_admin), db: Session = Depends(get_db)):
    if db.query(BootcampCatalog).filter(BootcampCatalog.slug == payload.slug).first():
        raise HTTPException(status_code=409, detail="Bootcamp slug already exists")
    
    bootcamp = BootcampCatalog(
        slug=payload.slug, title=payload.title, tagline=payload.tagline, description=payload.description,
        category=payload.category, difficulty=payload.difficulty, modules_count=payload.modules_count,
        labs_count=payload.labs_count, duration_hours=payload.duration_hours, price_inr=payload.price_inr,
        currency=payload.currency or "INR", billing_label=payload.billing_label, hero_image_url=payload.hero_image_url,
        accent_color=payload.accent_color or "red", audience=payload.audience or [], benefits=payload.benefits or [],
        syllabus=payload.syllabus or [], is_active=payload.is_active if payload.is_active is not None else True,
        display_order=payload.display_order or 0,
    )
    db.add(bootcamp)
    db.commit()
    db.refresh(bootcamp)
    after = _bootcamp_row(bootcamp)
    write_audit(db=db, admin_id=admin.id, action="create", entity_type="bootcamp", entity_id=bootcamp.id, before=None, after=after)
    return after

@app.put("/api/admin/bootcamps/{bootcamp_id}")
def admin_update_bootcamp(bootcamp_id: int, payload: BootcampUpdate, admin: Admin = Depends(get_current_admin), db: Session = Depends(get_db)):
    bootcamp = db.query(BootcampCatalog).filter(BootcampCatalog.id == bootcamp_id).first()
    if not bootcamp:
        raise HTTPException(status_code=404, detail="Bootcamp not found")
    before = _bootcamp_row(bootcamp)
    data = payload.dict(exclude_unset=True)
    new_slug = data.get("slug")
    if new_slug and new_slug != bootcamp.slug:
        if db.query(BootcampCatalog).filter(BootcampCatalog.slug == new_slug).first():
            raise HTTPException(status_code=409, detail="Bootcamp slug already exists")
    for field, value in data.items():
        setattr(bootcamp, field, value)
    db.commit()
    after = _bootcamp_row(bootcamp)
    write_audit(db=db, admin_id=admin.id, action="update", entity_type="bootcamp", entity_id=bootcamp.id, before=before, after=after)
    return after

@app.delete("/api/admin/bootcamps/{bootcamp_id}")
def admin_delete_bootcamp(bootcamp_id: int, admin: Admin = Depends(get_current_admin), db: Session = Depends(get_db)):
    bootcamp = db.query(BootcampCatalog).filter(BootcampCatalog.id == bootcamp_id).first()
    if not bootcamp:
        raise HTTPException(status_code=404, detail="Bootcamp not found")
    before = _bootcamp_row(bootcamp)
    bootcamp.is_active = False
    db.commit()
    write_audit(db=db, admin_id=admin.id, action="delete", entity_type="bootcamp", entity_id=bootcamp.id, before=before, after={"is_active": False})
    return {"success": True}


# ============================================================================
# ADMIN ROUTES: COUPONS
# ============================================================================
def _coupon_row(c: Coupon) -> dict:
    return {
        "id": c.id,
        "code": c.code,
        "discount_percent": c.discount_percent,
        "description": c.description,
        "max_uses": c.max_uses,
        "uses_count": c.uses_count,
        "expires_at": _iso_z(c.expires_at),
        "is_active": c.is_active,
        "created_at": _iso_z(c.created_at),
        "updated_at": _iso_z(c.updated_at),
    }


def _parse_expires_at(raw):
    """Accept ISO 8601 datetime (with or without TZ) or date; None passes through."""
    if raw in (None, "", "null"):
        return None
    try:
        # datetime.fromisoformat handles '2026-12-31', '2026-12-31T23:59:59', and '...+00:00'
        dt = datetime.fromisoformat(str(raw).replace("Z", "+00:00"))
    except ValueError:
        raise HTTPException(status_code=400, detail="expires_at must be ISO 8601 (e.g. 2026-12-31 or 2026-12-31T23:59:59+05:30)")
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt


@app.get("/api/admin/coupons")
def admin_list_coupons(
    admin: Admin = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    rows = db.query(Coupon).order_by(Coupon.created_at.desc()).all()
    return [_coupon_row(c) for c in rows]


@app.post("/api/admin/coupons")
def admin_create_coupon(
    payload: CouponPayload,
    admin: Admin = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    code = (payload.code or "").strip().upper()
    if not code:
        raise HTTPException(status_code=400, detail="Code required")
    if not (1 <= int(payload.discount_percent) <= 100):
        raise HTTPException(status_code=400, detail="discount_percent must be 1..100")
    if db.query(Coupon).filter(Coupon.code == code).first():
        raise HTTPException(status_code=409, detail="Coupon code already exists")
    coupon = Coupon(
        code=code,
        discount_percent=int(payload.discount_percent),
        description=payload.description,
        max_uses=payload.max_uses if payload.max_uses is not None and payload.max_uses > 0 else payload.max_uses,
        expires_at=_parse_expires_at(payload.expires_at),
        is_active=bool(payload.is_active) if payload.is_active is not None else True,
    )
    db.add(coupon)
    db.commit()
    db.refresh(coupon)
    after = _coupon_row(coupon)
    write_audit(
        db=db, admin_id=admin.id, action="create", entity_type="coupon",
        entity_id=coupon.id, before=None, after=after,
    )
    return after


@app.put("/api/admin/coupons/{coupon_id}")
def admin_update_coupon(
    coupon_id: int,
    payload: CouponUpdate,
    admin: Admin = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    coupon = db.query(Coupon).filter(Coupon.id == coupon_id).first()
    if not coupon:
        raise HTTPException(status_code=404, detail="Coupon not found")
    before = _coupon_row(coupon)
    data = payload.dict(exclude_unset=True)
    if "code" in data and data["code"] is not None:
        new_code = data["code"].strip().upper()
        if new_code != coupon.code and db.query(Coupon).filter(Coupon.code == new_code).first():
            raise HTTPException(status_code=409, detail="Coupon code already exists")
        coupon.code = new_code
    if "discount_percent" in data and data["discount_percent"] is not None:
        if not (1 <= int(data["discount_percent"]) <= 100):
            raise HTTPException(status_code=400, detail="discount_percent must be 1..100")
        coupon.discount_percent = int(data["discount_percent"])
    if "description" in data:
        coupon.description = data["description"]
    if "max_uses" in data:
        coupon.max_uses = data["max_uses"]
    if "expires_at" in data:
        coupon.expires_at = _parse_expires_at(data["expires_at"])
    if "is_active" in data and data["is_active"] is not None:
        coupon.is_active = bool(data["is_active"])
    db.commit()
    after = _coupon_row(coupon)
    write_audit(
        db=db, admin_id=admin.id, action="update", entity_type="coupon",
        entity_id=coupon.id, before=before, after=after,
    )
    return after


@app.delete("/api/admin/coupons/{coupon_id}")
def admin_delete_coupon(
    coupon_id: int,
    admin: Admin = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    """Soft delete (is_active=false) so the redemption history stays attached."""
    coupon = db.query(Coupon).filter(Coupon.id == coupon_id).first()
    if not coupon:
        raise HTTPException(status_code=404, detail="Coupon not found")
    before = _coupon_row(coupon)
    coupon.is_active = False
    db.commit()
    write_audit(
        db=db, admin_id=admin.id, action="delete", entity_type="coupon",
        entity_id=coupon.id, before=before, after={"is_active": False},
    )
    return {"success": True}


@app.get("/api/admin/coupons/{coupon_id}/redemptions")
def admin_list_coupon_redemptions(
    coupon_id: int,
    admin: Admin = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    if not db.query(Coupon).filter(Coupon.id == coupon_id).first():
        raise HTTPException(status_code=404, detail="Coupon not found")
    rows = (
        db.query(CouponRedemption, User.email)
        .join(User, User.id == CouponRedemption.user_id)
        .filter(CouponRedemption.coupon_id == coupon_id)
        .order_by(CouponRedemption.redeemed_at.desc())
        .all()
    )
    return [
        {
            "id": r.id,
            "user_id": r.user_id,
            "user_email": email,
            "credit_transaction_id": r.credit_transaction_id,
            "discount_amount": float(r.discount_amount),
            "redeemed_at": _iso_z(r.redeemed_at),
        }
        for r, email in rows
    ]


# ============================================================================
# ADMIN ROUTES: USERS
# ============================================================================
def _user_row(u: User, plan: Optional[str]) -> dict:
    return {
        "id": u.id,
        "email": u.email,
        "full_name": u.full_name,
        "credits": u.available_credits,
        "referral_code": u.referral_code,
        "is_active": u.is_active,
        "subscription_plan": plan,
        "created_at": u.created_at.isoformat() if u.created_at else None,
    }


@app.get("/api/admin/users")
def admin_list_users(
    search: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    size: int = Query(50, ge=1, le=200),
    _: Admin = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    offset = (page - 1) * size
    q = db.query(User)
    if search:
        like = f"%{search}%"
        q = q.filter(or_(User.email.ilike(like), User.full_name.ilike(like)))
    total = q.count()
    users = q.order_by(User.id.desc()).offset(offset).limit(size).all()

    user_ids = [u.id for u in users]
    subs = (
        db.query(Subscription)
        .filter(Subscription.user_id.in_(user_ids), Subscription.is_active == True)
        .all()
    )
    plan_by_user = {s.user_id: s.plan_name for s in subs}

    return {
        "users": [_user_row(u, plan_by_user.get(u.id)) for u in users],
        "total": total, "page": page, "size": size,
    }


@app.get("/api/admin/users/{user_id}")
def admin_get_user_detail(
    user_id: int,
    _: Admin = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    u = db.query(User).filter(User.id == user_id).first()
    if not u:
        raise HTTPException(status_code=404, detail="User not found")

    sub = (
        db.query(Subscription)
        .filter(Subscription.user_id == user_id, Subscription.is_active == True)
        .order_by(Subscription.start_date.desc())
        .first()
    )
    subscription = None
    if sub:
        subscription = {
            "plan_name": sub.plan_name,
            "billing_cycle": sub.billing_cycle,
            "start_date": sub.start_date.isoformat() if sub.start_date else None,
            "end_date": sub.end_date.isoformat() if sub.end_date else None,
        }

    recent_payments = (
        db.query(CreditTransaction)
        .filter(CreditTransaction.user_id == user_id)
        .order_by(CreditTransaction.purchased_at.desc())
        .limit(10)
        .all()
    )

    return {
        "user": {
            "id": u.id,
            "email": u.email,
            "full_name": u.full_name,
            "credits": u.available_credits,
            "referral_code": u.referral_code,
            "newsletter_opt_in": u.newsletter_opt_in,
            "is_active": u.is_active,
            "created_at": u.created_at.isoformat() if u.created_at else None,
        },
        "subscription": subscription,
        "active_lab": None,  # populated in Part 2 when lab integration ships
        "recent_payments": [
            {
                "id": p.id,
                "plan_name": p.plan_name,
                "amount": float(p.amount_paid) if p.amount_paid is not None else 0,
                "credits_added": p.credits_added,
                "type": "subscription" if (p.plan_name or "").startswith("sub:") else "credit",
                "status": "success",
                "created_at": p.purchased_at.isoformat() if p.purchased_at else None,
            }
            for p in recent_payments
        ],
    }


@app.put("/api/admin/users/{user_id}/credits")
def admin_adjust_user_credits(
    user_id: int,
    req: CreditAdjustRequest,
    admin: Admin = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    if not req.reason or not req.reason.strip():
        raise HTTPException(status_code=400, detail="Reason is required")

    u = db.query(User).filter(User.id == user_id).first()
    if not u:
        raise HTTPException(status_code=404, detail="User not found")

    before_credits = u.available_credits or 0
    after_credits = before_credits + req.delta
    if after_credits < 0:
        raise HTTPException(status_code=400, detail="Cannot reduce credits below zero")

    u.available_credits = after_credits
    db.commit()

    write_audit(
        db=db, admin_id=admin.id, action="credit_adjust", entity_type="user",
        entity_id=user_id,
        before={"credits": before_credits},
        after={"credits": after_credits, "delta": req.delta, "reason": req.reason},
    )
    return {"success": True, "new_balance": after_credits}


@app.put("/api/admin/users/{user_id}/status")
def admin_set_user_status(
    user_id: int,
    req: UserStatusRequest,
    admin: Admin = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    if not req.reason or not req.reason.strip():
        raise HTTPException(status_code=400, detail="Reason is required")

    u = db.query(User).filter(User.id == user_id).first()
    if not u:
        raise HTTPException(status_code=404, detail="User not found")

    before_active = u.is_active
    u.is_active = req.is_active
    db.commit()

    write_audit(
        db=db, admin_id=admin.id, action="status_change", entity_type="user",
        entity_id=user_id,
        before={"is_active": before_active},
        after={"is_active": req.is_active, "reason": req.reason},
    )
    return {"success": True, "is_active": req.is_active}


# ============================================================================
# ADMIN ROUTES: CONTACTS
# ============================================================================
def _contact_to_dict(c: ContactMessage) -> dict:
    return {
        "id": c.id,
        "name": c.name,
        "email": c.email,
        "message": c.message,
        "status": c.status or "unread",
        "admin_note": c.admin_note,
        "created_at": c.created_at.isoformat() if c.created_at else None,
    }


@app.get("/api/admin/contacts")
def admin_list_contacts(
    status: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    size: int = Query(25, ge=1, le=200),
    _: Admin = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    if status and status not in VALID_CONTACT_STATUSES:
        raise HTTPException(status_code=400, detail="Invalid status filter")

    offset = (page - 1) * size
    q = db.query(ContactMessage)
    if status:
        q = q.filter(func.coalesce(ContactMessage.status, "unread") == status)
    total = q.count()
    rows = q.order_by(ContactMessage.created_at.desc()).offset(offset).limit(size).all()

    return {
        "contacts": [_contact_to_dict(r) for r in rows],
        "total": total, "page": page, "size": size,
    }


@app.put("/api/admin/contacts/{contact_id}/status")
def admin_update_contact_status(
    contact_id: int,
    req: ContactStatusRequest,
    admin: Admin = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    if req.status not in VALID_CONTACT_STATUSES:
        raise HTTPException(status_code=400, detail="Invalid status")

    c = db.query(ContactMessage).filter(ContactMessage.id == contact_id).first()
    if not c:
        raise HTTPException(status_code=404, detail="Contact message not found")

    before = {"status": c.status or "unread", "admin_note": c.admin_note}
    c.status = req.status
    c.admin_note = req.note
    db.commit()

    write_audit(
        db=db, admin_id=admin.id, action="update", entity_type="contact",
        entity_id=contact_id, before=before,
        after={"status": req.status, "admin_note": req.note},
    )
    return {"success": True}


# ============================================================================
# ADMIN ROUTES: PAYMENTS
# ============================================================================
@app.get("/api/admin/payments")
def admin_list_payments(
    user_id: Optional[int] = Query(None),
    from_: Optional[str] = Query(None, alias="from"),
    to: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    size: int = Query(50, ge=1, le=200),
    _: Admin = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    payments = []

    # --- credit pack + subscription transactions ---
    cq = db.query(CreditTransaction, User.email).outerjoin(User, User.id == CreditTransaction.user_id)
    if user_id is not None:
        cq = cq.filter(CreditTransaction.user_id == user_id)
    if from_:
        cq = cq.filter(cast(CreditTransaction.purchased_at, Date) >= from_)
    if to:
        cq = cq.filter(cast(CreditTransaction.purchased_at, Date) <= to)
    for txn, email in cq.all():
        payments.append({
            "id": f"cr-{txn.id}",
            "user_id": txn.user_id,
            "user_email": email,
            "plan_name": txn.plan_name,
            "amount": float(txn.amount_paid) if txn.amount_paid is not None else 0,
            "credits_added": txn.credits_added,
            "type": "subscription" if (txn.plan_name or "").startswith("sub:") else "credit",
            "status": "success",
            "razorpay_payment_id": None,
            "razorpay_order_id": None,
            "created_at": txn.purchased_at.isoformat() if txn.purchased_at else None,
        })

    # --- course purchases ---
    cpq = db.query(CoursePurchase, User.email).outerjoin(User, User.id == CoursePurchase.user_id)
    if user_id is not None:
        cpq = cpq.filter(CoursePurchase.user_id == user_id)
    if from_:
        cpq = cpq.filter(cast(CoursePurchase.purchased_at, Date) >= from_)
    if to:
        cpq = cpq.filter(cast(CoursePurchase.purchased_at, Date) <= to)
    for cp, email in cpq.all():
        payments.append({
            "id": f"course-{cp.id}",
            "user_id": cp.user_id,
            "user_email": email,
            "plan_name": cp.course_title or cp.course_slug,
            "amount": float(cp.total_amount) if cp.total_amount is not None else 0,
            "credits_added": 0,
            "type": "course",
            "status": "success",
            "razorpay_payment_id": cp.razorpay_payment_id,
            "razorpay_order_id": cp.razorpay_order_id,
            "created_at": cp.purchased_at.isoformat() if cp.purchased_at else None,
        })

    # --- bootcamp purchases ---
    bpq = db.query(BootcampPurchase, User.email).outerjoin(User, User.id == BootcampPurchase.user_id)
    if user_id is not None:
        bpq = bpq.filter(BootcampPurchase.user_id == user_id)
    if from_:
        bpq = bpq.filter(cast(BootcampPurchase.purchased_at, Date) >= from_)
    if to:
        bpq = bpq.filter(cast(BootcampPurchase.purchased_at, Date) <= to)
    for bp, email in bpq.all():
        payments.append({
            "id": f"bootcamp-{bp.id}",
            "user_id": bp.user_id,
            "user_email": email,
            "plan_name": bp.bootcamp_title or bp.bootcamp_slug,
            "amount": float(bp.total_amount) if bp.total_amount is not None else 0,
            "credits_added": 0,
            "type": "bootcamp",
            "status": "success",
            "razorpay_payment_id": bp.razorpay_payment_id,
            "razorpay_order_id": bp.razorpay_order_id,
            "created_at": bp.purchased_at.isoformat() if bp.purchased_at else None,
        })

    # sort all sources newest-first then paginate
    payments.sort(key=lambda x: x["created_at"] or "", reverse=True)
    total = len(payments)
    offset = (page - 1) * size
    return {"payments": payments[offset: offset + size], "total": total, "page": page, "size": size}


@app.get("/api/admin/payments/summary")
def admin_payments_summary(
    from_: Optional[str] = Query(None, alias="from"),
    to: Optional[str] = Query(None),
    _: Admin = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    from collections import defaultdict

    daily = defaultdict(float)
    by_plan_map = defaultdict(lambda: {"total": 0.0, "count": 0})

    def _apply(rows, date_field, amount_field, label_field):
        for r in rows:
            d = str(getattr(r, date_field).date()) if getattr(r, date_field) else None
            amt = float(getattr(r, amount_field) or 0)
            lbl = getattr(r, label_field) or "unknown"
            if d:
                daily[d] += amt
            by_plan_map[lbl]["total"] += amt
            by_plan_map[lbl]["count"] += 1

    # credit/subscription transactions
    cq = db.query(CreditTransaction)
    if from_:
        cq = cq.filter(cast(CreditTransaction.purchased_at, Date) >= from_)
    if to:
        cq = cq.filter(cast(CreditTransaction.purchased_at, Date) <= to)
    _apply(cq.all(), "purchased_at", "amount_paid", "plan_name")

    # course purchases
    cpq = db.query(CoursePurchase)
    if from_:
        cpq = cpq.filter(cast(CoursePurchase.purchased_at, Date) >= from_)
    if to:
        cpq = cpq.filter(cast(CoursePurchase.purchased_at, Date) <= to)
    _apply(cpq.all(), "purchased_at", "total_amount", "course_title")

    # bootcamp purchases
    bpq = db.query(BootcampPurchase)
    if from_:
        bpq = bpq.filter(cast(BootcampPurchase.purchased_at, Date) >= from_)
    if to:
        bpq = bpq.filter(cast(BootcampPurchase.purchased_at, Date) <= to)
    _apply(bpq.all(), "purchased_at", "total_amount", "bootcamp_title")

    by_day = [{"date": d, "total": daily[d]} for d in sorted(daily)]
    by_plan = [{"plan_name": k, "total": v["total"], "count": v["count"]} for k, v in by_plan_map.items()]
    return {"by_day": by_day, "by_plan": by_plan}


# ============================================================================
# ADMIN ROUTES: UNIVERSITIES (LMS)
# ============================================================================
class UniCreateRequest(BaseModel):
    name: str
    slug: str
    domain: Optional[str] = None
    enforce_domain: bool = False
    logo_url: Optional[str] = None
    description: Optional[str] = None
    credits_per_student: int = 0
    display_order: int = 0

class UniUpdateRequest(BaseModel):
    name: Optional[str] = None
    slug: Optional[str] = None
    domain: Optional[str] = None
    enforce_domain: Optional[bool] = None
    logo_url: Optional[str] = None
    description: Optional[str] = None
    credits_per_student: Optional[int] = None
    display_order: Optional[int] = None
    is_active: Optional[bool] = None

class ProgramCreateRequest(BaseModel):
    name: str
    slug: str
    description: Optional[str] = None
    credits_per_student: int = 0
    display_order: int = 0
    stat_duration: Optional[str] = None
    stat_labs: Optional[str] = None
    stat_modules: Optional[str] = None

class ProgramUpdateRequest(BaseModel):
    name: Optional[str] = None
    slug: Optional[str] = None
    description: Optional[str] = None
    credits_per_student: Optional[int] = None
    display_order: Optional[int] = None
    is_active: Optional[bool] = None
    stat_duration: Optional[str] = None
    stat_labs: Optional[str] = None
    stat_modules: Optional[str] = None

class SemesterSaveRequest(BaseModel):
    semester_number: int
    name: str
    content: list = []
    credits_grant: int = 0

# Sentinel written to razorpay_payment_id when an admin grants LMS access
# manually, so it is distinguishable from a genuine Razorpay payment.
_MANUAL_LMS_GRANT = "manual_admin_grant"


class LMSAccessRequest(BaseModel):
    lms_paid: bool


class AddStudentRequest(BaseModel):
    email: str
    program_id: Optional[int] = None

def _uni_row(u: University):
    return {
        "id": u.id, "name": u.name, "slug": u.slug, "domain": u.domain,
        "enforce_domain": u.enforce_domain,
        "logo_url": u.logo_url, "description": u.description,
        "credits_per_student": u.credits_per_student,
        "is_active": u.is_active, "display_order": u.display_order,
        "created_at": u.created_at.isoformat() if u.created_at else None,
    }

def _prog_row(p: UniversityProgram):
    return {
        "id": p.id, "university_id": p.university_id, "name": p.name,
        "slug": p.slug, "description": p.description,
        "credits_per_student": p.credits_per_student,
        "is_active": p.is_active, "display_order": p.display_order,
        "stat_duration": p.stat_duration, "stat_labs": p.stat_labs,
        "stat_modules": p.stat_modules,
        "created_at": p.created_at.isoformat() if p.created_at else None,
    }

def _sem_row(s: UniversitySemester):
    return {
        "id": s.id, "university_id": s.university_id, "program_id": s.program_id,
        "semester_number": s.semester_number, "name": s.name,
        "content": s.content or [], "credits_grant": s.credits_grant,
        "links_total": s.links_total, "links_ok": s.links_ok,
        "links_warn": s.links_warn,
        "links_checked_at": s.links_checked_at.isoformat() if s.links_checked_at else None,
    }

def _ustudent_row(us: UniversityStudent, email: str = None, program_name: str = None):
    return {
        "id": us.id, "university_id": us.university_id, "user_id": us.user_id,
        "program_id": us.program_id, "program_name": program_name,
        "email": email, "current_semester": us.current_semester,
        "status": us.status, "lms_paid": us.lms_paid,
        "manual_grant": us.razorpay_payment_id == _MANUAL_LMS_GRANT,
        "enrolled_at": us.enrolled_at.isoformat() if us.enrolled_at else None,
        "promoted_at": us.promoted_at.isoformat() if us.promoted_at else None,
    }


def _assessment_row(a: Assessment):
    return {
        "id": a.id, "title": a.title, "slug": a.slug,
        "description": a.description, "topic": a.topic,
        "difficulty": a.difficulty, "time_limit_minutes": a.time_limit_minutes,
        "is_active": a.is_active,
    }


def _question_row(q: AssessmentQuestion, include_answers: bool = False):
    row = {
        "id": q.id, "assessment_id": q.assessment_id,
        "question_text": q.question_text, "options": q.options,
        "multi_select": q.multi_select, "order_num": q.order_num,
    }
    if include_answers:
        row["correct_answers"] = q.correct_answers
    return row


@app.get("/api/admin/universities")
def admin_list_universities(_: Admin = Depends(get_current_admin), db: Session = Depends(get_db)):
    rows = db.query(University).order_by(University.display_order, University.id).all()
    return [_uni_row(u) for u in rows]

@app.post("/api/admin/universities")
def admin_create_university(req: UniCreateRequest, admin: Admin = Depends(get_current_admin), db: Session = Depends(get_db)):
    u = University(**req.dict())
    db.add(u)
    db.commit()
    db.refresh(u)
    write_audit(db, admin.id, "university", u.id, "create", {}, _uni_row(u))
    return _uni_row(u)

@app.put("/api/admin/universities/{uid}")
def admin_update_university(uid: int, req: UniUpdateRequest, admin: Admin = Depends(get_current_admin), db: Session = Depends(get_db)):
    u = db.query(University).filter(University.id == uid).first()
    if not u:
        raise HTTPException(status_code=404, detail="University not found")
    before = _uni_row(u)
    for k, v in req.dict(exclude_none=True).items():
        setattr(u, k, v)
    db.commit()
    write_audit(db, admin.id, "university", u.id, "update", before, _uni_row(u))
    return _uni_row(u)

@app.delete("/api/admin/universities/{uid}")
def admin_delete_university(uid: int, admin: Admin = Depends(get_current_admin), db: Session = Depends(get_db)):
    u = db.query(University).filter(University.id == uid).first()
    if not u:
        raise HTTPException(status_code=404, detail="University not found")
    u.is_active = False
    db.commit()
    write_audit(db, admin.id, "university", uid, "delete", {}, {"is_active": False})
    return {"success": True}

@app.get("/api/admin/universities/{uid}/programs")
def admin_list_programs(uid: int, _: Admin = Depends(get_current_admin), db: Session = Depends(get_db)):
    rows = db.query(UniversityProgram).filter(UniversityProgram.university_id == uid).order_by(UniversityProgram.display_order, UniversityProgram.id).all()

    # Roll the cached per-semester link counts up to the program, so the list
    # shows a reachability tag before anything is expanded. Semesters that were
    # never checked hold NULLs and are ignored by SUM.
    stats = {}
    for pid_, total, ok, warn, checked in (
        db.query(
            UniversitySemester.program_id,
            sqlfunc.sum(UniversitySemester.links_total),
            sqlfunc.sum(UniversitySemester.links_ok),
            sqlfunc.sum(UniversitySemester.links_warn),
            sqlfunc.max(UniversitySemester.links_checked_at),
        )
        .filter(UniversitySemester.university_id == uid)
        .group_by(UniversitySemester.program_id)
        .all()
    ):
        if total is None:
            continue
        stats[pid_] = {
            "links_total": int(total),
            "links_ok": int(ok or 0),
            "links_warn": int(warn or 0),
            "links_checked_at": checked.isoformat() if checked else None,
        }

    out = []
    for p in rows:
        row = _prog_row(p)
        row["link_stats"] = stats.get(p.id)
        out.append(row)
    return out

@app.post("/api/admin/universities/{uid}/programs")
def admin_create_program(uid: int, req: ProgramCreateRequest, admin: Admin = Depends(get_current_admin), db: Session = Depends(get_db)):
    p = UniversityProgram(university_id=uid, **req.dict())
    db.add(p)
    db.commit()
    db.refresh(p)
    write_audit(db, admin.id, "university", uid, "create_program", {}, _prog_row(p))
    return _prog_row(p)

@app.put("/api/admin/universities/{uid}/programs/{pid}")
def admin_update_program(uid: int, pid: int, req: ProgramUpdateRequest, admin: Admin = Depends(get_current_admin), db: Session = Depends(get_db)):
    p = db.query(UniversityProgram).filter(UniversityProgram.id == pid, UniversityProgram.university_id == uid).first()
    if not p:
        raise HTTPException(status_code=404, detail="Program not found")
    before = _prog_row(p)
    for k, v in req.dict(exclude_none=True).items():
        setattr(p, k, v)
    db.commit()
    write_audit(db, admin.id, "university", uid, "update_program", before, _prog_row(p))
    return _prog_row(p)

@app.delete("/api/admin/universities/{uid}/programs/{pid}")
def admin_delete_program(uid: int, pid: int, admin: Admin = Depends(get_current_admin), db: Session = Depends(get_db)):
    p = db.query(UniversityProgram).filter(UniversityProgram.id == pid, UniversityProgram.university_id == uid).first()
    if not p:
        raise HTTPException(status_code=404, detail="Program not found")
    db.delete(p)
    db.commit()
    write_audit(db, admin.id, "university", uid, "delete_program", {"program_id": pid}, {})
    return {"success": True}

@app.get("/api/admin/universities/{uid}/programs/{pid}/semesters")
def admin_list_semesters(uid: int, pid: int, _: Admin = Depends(get_current_admin), db: Session = Depends(get_db)):
    rows = db.query(UniversitySemester).filter(UniversitySemester.program_id == pid).order_by(UniversitySemester.semester_number).all()
    return [_sem_row(s) for s in rows]

@app.post("/api/admin/universities/{uid}/programs/{pid}/semesters")
def admin_save_semester(uid: int, pid: int, req: SemesterSaveRequest, admin: Admin = Depends(get_current_admin), db: Session = Depends(get_db)):
    existing = db.query(UniversitySemester).filter(
        UniversitySemester.program_id == pid,
        UniversitySemester.semester_number == req.semester_number,
    ).first()
    if existing:
        content_changed = (existing.content or []) != (req.content or [])
        existing.name = req.name
        existing.content = req.content
        existing.credits_grant = req.credits_grant
        # The cached reachability summary describes the old URL set, so drop it
        # rather than leave a stale "links OK" tag against edited content.
        if content_changed:
            existing.links_total = None
            existing.links_ok = None
            existing.links_warn = None
            existing.links_checked_at = None
        db.commit()
        return _sem_row(existing)
    s = UniversitySemester(university_id=uid, program_id=pid, **req.dict())
    db.add(s)
    db.commit()
    db.refresh(s)
    write_audit(db, admin.id, "university", uid, "add_semester", {}, _sem_row(s))
    return _sem_row(s)

@app.delete("/api/admin/universities/{uid}/programs/{pid}/semesters/{sem_id}")
def admin_delete_semester(uid: int, pid: int, sem_id: int, admin: Admin = Depends(get_current_admin), db: Session = Depends(get_db)):
    s = db.query(UniversitySemester).filter(UniversitySemester.id == sem_id, UniversitySemester.program_id == pid).first()
    if not s:
        raise HTTPException(status_code=404, detail="Semester not found")
    db.delete(s)
    db.commit()
    write_audit(db, admin.id, "university", uid, "delete_semester", {"semester_number": s.semester_number}, {})
    return {"success": True}

@app.get("/api/admin/universities/{uid}/students")
def admin_list_uni_students(uid: int, _: Admin = Depends(get_current_admin), db: Session = Depends(get_db)):
    rows = (
        db.query(UniversityStudent, User.email, UniversityProgram.name)
        .outerjoin(User, User.id == UniversityStudent.user_id)
        .outerjoin(UniversityProgram, UniversityProgram.id == UniversityStudent.program_id)
        .filter(UniversityStudent.university_id == uid)
        .order_by(UniversityStudent.enrolled_at.desc())
        .all()
    )
    return [_ustudent_row(us, email, prog_name) for us, email, prog_name in rows]

@app.post("/api/admin/universities/{uid}/students")
def admin_add_uni_student(uid: int, req: AddStudentRequest, admin: Admin = Depends(get_current_admin), db: Session = Depends(get_db)):
    uni = db.query(University).filter(University.id == uid).first()
    if not uni:
        raise HTTPException(status_code=404, detail="University not found")
    email = req.email.strip().lower()
    if uni.enforce_domain and uni.domain:
        if not email.endswith("@" + uni.domain.lower().lstrip("@")):
            raise HTTPException(status_code=400, detail=f"Email must be from @{uni.domain}")
    user = db.query(User).filter(User.email == email).first()
    if not user:
        raise HTTPException(status_code=404, detail="No user found with that email")
    existing = db.query(UniversityStudent).filter(
        UniversityStudent.university_id == uid, UniversityStudent.user_id == user.id
    ).first()
    if existing:
        raise HTTPException(status_code=409, detail="Student already enrolled in this university")
    us = UniversityStudent(university_id=uid, program_id=req.program_id, user_id=user.id, status="pending")
    db.add(us)
    db.commit()
    db.refresh(us)
    write_audit(db, admin.id, "university", uid, "add_student", {}, {"user_email": email, "program_id": req.program_id})
    prog = db.query(UniversityProgram).filter(UniversityProgram.id == req.program_id).first() if req.program_id else None
    return _ustudent_row(us, user.email, prog.name if prog else None)

@app.post("/api/admin/universities/{uid}/students/upload-csv")
async def admin_upload_students_csv(uid: int, file: UploadFile, admin: Admin = Depends(get_current_admin), db: Session = Depends(get_db)):
    import csv, io
    uni = db.query(University).filter(University.id == uid).first()
    if not uni:
        raise HTTPException(status_code=404, detail="University not found")
    content = await file.read()
    try:
        text = content.decode("utf-8-sig")
    except Exception:
        raise HTTPException(status_code=400, detail="Could not decode CSV — ensure it is UTF-8")
    reader = csv.DictReader(io.StringIO(text))
    # accept "email", "Email", "EMAIL", or first column
    fieldnames = reader.fieldnames or []
    email_col = next((f for f in fieldnames if f.strip().lower() == "email"), fieldnames[0] if fieldnames else None)
    if not email_col:
        raise HTTPException(status_code=400, detail="CSV must have an 'email' column")
    enrolled, skipped, not_found, domain_blocked = [], [], [], []
    for row in reader:
        raw = row.get(email_col, "").strip().lower()
        if not raw:
            continue
        if uni.enforce_domain and uni.domain:
            if not raw.endswith("@" + uni.domain.lower().lstrip("@")):
                domain_blocked.append(raw)
                continue
        user = db.query(User).filter(User.email == raw).first()
        if not user:
            not_found.append(raw)
            continue
        existing = db.query(UniversityStudent).filter(
            UniversityStudent.university_id == uid, UniversityStudent.user_id == user.id
        ).first()
        if existing:
            skipped.append(raw)
            continue
        db.add(UniversityStudent(university_id=uid, program_id=None, user_id=user.id, status="pending"))
        enrolled.append(raw)
    db.commit()
    write_audit(db, admin.id, "university", uid, "bulk_csv_enroll", {}, {"enrolled": len(enrolled)})
    return {"enrolled": enrolled, "skipped": skipped, "not_found": not_found, "domain_blocked": domain_blocked}

@app.delete("/api/admin/universities/{uid}/students/{user_id}")
def admin_remove_uni_student(uid: int, user_id: int, admin: Admin = Depends(get_current_admin), db: Session = Depends(get_db)):
    us = db.query(UniversityStudent).filter(
        UniversityStudent.university_id == uid, UniversityStudent.user_id == user_id
    ).first()
    if not us:
        raise HTTPException(status_code=404, detail="Student not found")
    db.delete(us)
    db.commit()
    write_audit(db, admin.id, "university", uid, "remove_student", {"user_id": user_id}, {})
    return {"success": True}

@app.put("/api/admin/universities/{uid}/students/{user_id}/lms-access")
def admin_set_lms_access(
    uid: int, user_id: int, req: LMSAccessRequest,
    admin: Admin = Depends(get_current_admin), db: Session = Depends(get_db),
):
    """Grant or revoke LMS access without a Razorpay payment.

    Needed because Razorpay Checkout only runs on domains registered with the
    account, so it cannot be exercised on a bare IP. Also covers students whose
    university pays offline. Mirrors lms_verify_payment, but stamps the grant
    as manual so a revoke can tell an admin comp apart from a real purchase.
    """
    us = db.query(UniversityStudent).filter(
        UniversityStudent.university_id == uid,
        UniversityStudent.user_id == user_id,
    ).first()
    if not us:
        raise HTTPException(status_code=404, detail="Student not enrolled in this university")
    if us.lms_paid == req.lms_paid:
        return _ustudent_row(us)

    before = _ustudent_row(us)
    user = db.query(User).filter(User.id == user_id).first()
    uni = db.query(University).filter(University.id == uid).first()
    grant = (uni.credits_per_student or 0) if uni else 0

    if req.lms_paid:
        us.lms_paid = True
        us.status = "active"
        us.razorpay_payment_id = _MANUAL_LMS_GRANT
        if user and grant > 0:
            user.available_credits += grant
    else:
        us.lms_paid = False
        us.status = "pending"
        # Only claw back credits this endpoint handed out — a real payment's
        # credits stay with the student.
        if us.razorpay_payment_id == _MANUAL_LMS_GRANT:
            if user and grant > 0:
                user.available_credits = max(0, user.available_credits - grant)
            us.razorpay_payment_id = None

    db.commit()
    write_audit(db, admin.id, "university_student", us.id,
                "grant_lms_access" if req.lms_paid else "revoke_lms_access",
                before, _ustudent_row(us))
    return _ustudent_row(us)


@app.post("/api/admin/universities/{uid}/students/{user_id}/promote")
def admin_promote_student(uid: int, user_id: int, admin: Admin = Depends(get_current_admin), db: Session = Depends(get_db)):
    us = db.query(UniversityStudent).filter(
        UniversityStudent.university_id == uid, UniversityStudent.user_id == user_id
    ).first()
    if not us:
        raise HTTPException(status_code=404, detail="Student not found")
    next_sem = us.current_semester + 1
    sem_q = db.query(UniversitySemester).filter(UniversitySemester.semester_number == next_sem)
    if us.program_id:
        sem_q = sem_q.filter(UniversitySemester.program_id == us.program_id)
    else:
        sem_q = sem_q.filter(UniversitySemester.university_id == uid)
    sem = sem_q.first()
    if sem and sem.credits_grant > 0:
        user = db.query(User).filter(User.id == user_id).first()
        if user:
            user.available_credits += sem.credits_grant
    us.current_semester = next_sem
    us.status = "active"
    us.promoted_at = datetime.utcnow()
    db.commit()
    write_audit(db, admin.id, "university", uid, "promote_student", {"user_id": user_id, "from_sem": next_sem - 1}, {"to_sem": next_sem})
    return {"success": True, "current_semester": next_sem}

@app.post("/api/admin/universities/{uid}/students/bulk-promote")
def admin_bulk_promote(uid: int, admin: Admin = Depends(get_current_admin), db: Session = Depends(get_db)):
    students = db.query(UniversityStudent).filter(
        UniversityStudent.university_id == uid,
        UniversityStudent.status == "active",
        UniversityStudent.lms_paid == True,
    ).all()
    promoted = 0
    for us in students:
        next_sem = us.current_semester + 1
        sem_q = db.query(UniversitySemester).filter(UniversitySemester.semester_number == next_sem)
        if us.program_id:
            sem_q = sem_q.filter(UniversitySemester.program_id == us.program_id)
        else:
            sem_q = sem_q.filter(UniversitySemester.university_id == uid)
        sem = sem_q.first()
        if sem and sem.credits_grant > 0:
            user = db.query(User).filter(User.id == us.user_id).first()
            if user:
                user.available_credits += sem.credits_grant
        us.current_semester = next_sem
        us.promoted_at = datetime.utcnow()
        promoted += 1
    db.commit()
    write_audit(db, admin.id, "university", uid, "bulk_promote", {}, {"promoted_count": promoted})
    return {"success": True, "promoted": promoted}


# ============================================================================
# STUDENT-FACING ROUTES: LMS
# ============================================================================
class LMSOrderRequest(BaseModel):
    pass

class LMSVerifyRequest(BaseModel):
    razorpay_order_id: str
    razorpay_payment_id: str
    razorpay_signature: str

LMS_PRICE_INR = 1

@app.get("/api/lms/my")
def get_my_lms(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    us = db.query(UniversityStudent).filter(UniversityStudent.user_id == current_user.id).first()
    if not us:
        return {"enrolled": False}
    uni = db.query(University).filter(University.id == us.university_id).first()
    prog = db.query(UniversityProgram).filter(UniversityProgram.id == us.program_id).first() if us.program_id else None
    sem_q = db.query(UniversitySemester).filter(UniversitySemester.semester_number == us.current_semester)
    if us.program_id:
        sem_q = sem_q.filter(UniversitySemester.program_id == us.program_id)
    else:
        sem_q = sem_q.filter(UniversitySemester.university_id == us.university_id)
    sem = sem_q.first()
    return {
        "enrolled": True,
        "lms_paid": us.lms_paid,
        "status": us.status,
        "current_semester": us.current_semester,
        "university": _uni_row(uni) if uni else None,
        "program": _prog_row(prog) if prog else None,
        "semester": _sem_row(sem) if sem else None,
    }

@app.post("/api/lms/create-order")
def lms_create_order(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    us = db.query(UniversityStudent).filter(UniversityStudent.user_id == current_user.id).first()
    if not us:
        raise HTTPException(status_code=403, detail="Not enrolled in any university LMS")
    if us.lms_paid:
        raise HTTPException(status_code=400, detail="LMS access already unlocked")
    amount_paise = with_gst(LMS_PRICE_INR) * 100
    order = razorpay_client.order.create({
        "amount": amount_paise,
        "currency": "INR",
        "receipt": f"lms_{current_user.id}",
        "notes": {"user_id": str(current_user.id), "type": "lms"},
    })
    return {"order_id": order["id"], "amount": amount_paise, "currency": "INR", "key": os.getenv("RAZORPAY_KEY_ID", "")}

@app.post("/api/lms/verify-payment")
def lms_verify_payment(req: LMSVerifyRequest, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    body = f"{req.razorpay_order_id}|{req.razorpay_payment_id}"
    expected = hmac.new(
        key=os.getenv("RAZORPAY_KEY_SECRET", "").encode("utf-8"),
        msg=body.encode("utf-8"),
        digestmod=hashlib.sha256,
    ).hexdigest()
    if expected != req.razorpay_signature:
        raise HTTPException(status_code=400, detail="Invalid payment signature")
    us = db.query(UniversityStudent).filter(UniversityStudent.user_id == current_user.id).first()
    if not us:
        raise HTTPException(status_code=403, detail="Not enrolled in any university LMS")
    if us.lms_paid:
        return {"success": True, "duplicate": True}
    uni = db.query(University).filter(University.id == us.university_id).first()
    us.lms_paid = True
    us.status = "active"
    us.razorpay_order_id = req.razorpay_order_id
    us.razorpay_payment_id = req.razorpay_payment_id
    if uni and uni.credits_per_student > 0:
        current_user.available_credits += uni.credits_per_student
    db.commit()
    return {"success": True}


# ============================================================================
# ADMIN ROUTES: REFERRALS
# ============================================================================
@app.get("/api/admin/referrals")
def admin_list_referrals(
    page: int = Query(1, ge=1),
    size: int = Query(50, ge=1, le=200),
    _: Admin = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    Referrer = aliased(User)
    Referred = aliased(User)
    offset = (page - 1) * size

    total = db.query(Referral).count()
    rows = (
        db.query(Referral, Referrer.email, Referred.email)
        .outerjoin(Referrer, Referrer.id == Referral.referrer_id)
        .outerjoin(Referred, Referred.id == Referral.referred_user_id)
        .order_by(Referral.created_at.desc())
        .offset(offset)
        .limit(size)
        .all()
    )

    referrals = []
    for r, referrer_email, referred_email in rows:
        referrals.append({
            "id": r.id,
            "referrer_id": r.referrer_id,
            "referrer_email": referrer_email,
            "referred_id": r.referred_user_id,
            "referred_email": referred_email,
            "referrer_reward": r.reward_credits or 0,
            "status": r.status or "completed",
            "created_at": r.created_at.isoformat() if r.created_at else None,
        })
    return {"referrals": referrals, "total": total, "page": page, "size": size}


@app.get("/api/admin/referrals/{user_id}")
def admin_get_referral_detail(
    user_id: int,
    _: Admin = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    u = db.query(User).filter(User.id == user_id).first()
    if not u:
        raise HTTPException(status_code=404, detail="User not found")

    Referrer = aliased(User)
    inbound = (
        db.query(Referral, Referrer.email)
        .outerjoin(Referrer, Referrer.id == Referral.referrer_id)
        .filter(Referral.referred_user_id == user_id)
        .first()
    )
    referred_by = None
    if inbound:
        ref, ref_email = inbound
        referred_by = {"id": ref.referrer_id, "email": ref_email}

    Referred = aliased(User)
    outbound = (
        db.query(Referral, Referred.email)
        .outerjoin(Referred, Referred.id == Referral.referred_user_id)
        .filter(Referral.referrer_id == user_id)
        .order_by(Referral.created_at.desc())
        .all()
    )
    referrals_made = []
    for r, referred_email in outbound:
        referrals_made.append({
            "id": r.id,
            "referred_id": r.referred_user_id,
            "referred_email": referred_email,
            "referrer_reward": r.reward_credits or 0,
            "status": r.status or "completed",
            "created_at": r.created_at.isoformat() if r.created_at else None,
        })
    total_reward = sum(r["referrer_reward"] for r in referrals_made)

    return {
        "user": {"id": u.id, "email": u.email, "full_name": u.full_name},
        "referred_by": referred_by,
        "referrals_made": referrals_made,
        "total_reward": total_reward,
    }


# ============================================================================
# ADMIN ROUTES: AUDIT LOG
# ============================================================================
@app.get("/api/admin/audit-log")
def admin_list_audit_log(
    admin_id: Optional[int] = Query(None),
    entity_type: Optional[str] = Query(None),
    from_: Optional[str] = Query(None, alias="from"),
    to: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    size: int = Query(50, ge=1, le=200),
    _: Admin = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    offset = (page - 1) * size
    q = db.query(AdminAuditLog, Admin.email).outerjoin(Admin, Admin.id == AdminAuditLog.admin_id)
    if admin_id is not None:
        q = q.filter(AdminAuditLog.admin_id == admin_id)
    if entity_type:
        q = q.filter(AdminAuditLog.entity_type == entity_type)
    if from_:
        q = q.filter(cast(AdminAuditLog.created_at, Date) >= from_)
    if to:
        q = q.filter(cast(AdminAuditLog.created_at, Date) <= to)

    total = q.count()
    rows = q.order_by(AdminAuditLog.created_at.desc()).offset(offset).limit(size).all()

    entries = []
    for entry, admin_email in rows:
        entries.append({
            "id": entry.id,
            "admin_id": entry.admin_id,
            "admin_email": admin_email,
            "action": entry.action,
            "entity_type": entry.entity_type,
            "entity_id": entry.entity_id,
            "before_payload": entry.before_payload,
            "after_payload": entry.after_payload,
            "created_at": entry.created_at.isoformat() if entry.created_at else None,
        })
    return {"entries": entries, "total": total, "page": page, "size": size}


# ============================================================================
# ADMIN ROUTES: DASHBOARD
# ============================================================================
@app.get("/api/admin/dashboard")
def admin_dashboard(_: Admin = Depends(get_current_admin), db: Session = Depends(get_db)):
    total_users = db.query(func.count(User.id)).scalar() or 0
    active_subscriptions = (
        db.query(func.count(Subscription.id))
        .filter(Subscription.is_active == True)
        .scalar() or 0
    )
    todays_revenue = (
        db.query(func.coalesce(func.sum(CreditTransaction.amount_paid), 0))
        .filter(cast(CreditTransaction.purchased_at, Date) == date.today())
        .scalar() or 0
    )
    unread_contacts = (
        db.query(func.count(ContactMessage.id))
        .filter(func.coalesce(ContactMessage.status, "unread") == "unread")
        .scalar() or 0
    )
    total_labs = (
        db.query(func.count(LabCatalog.lab_id))
        .filter(LabCatalog.is_active == True)
        .scalar() or 0
    )

    return {
        "total_users": total_users,
        "active_subscriptions": active_subscriptions,
        "todays_revenue": float(todays_revenue),
        "unread_contacts": unread_contacts,
        "running_labs": 0,  # populated in Part 2 when lab integration ships
        "total_labs": total_labs,
    }


# ============================================================================
# BULK EMAILER — Gmail SMTP (mirrors backend/cydo-market/server_v2.js setup)
# ============================================================================
import csv
import io
import re
import smtplib
import time
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

SMTP_HOST = os.getenv("SMTP_HOST", "smtp.gmail.com")
SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))
SMTP_USER = os.getenv("SMTP_USER", "")
SMTP_PASSWORD = os.getenv("SMTP_PASSWORD", "")
SMTP_FROM = os.getenv("SMTP_FROM", 'Cyber Dojo <cyberdojo.02@gmail.com>')

# Gmail free-tier limit is ~500/day. Per-blast cap + inter-send delay keep us
# well under the threshold and reduce the chance of a temporary lock that would
# also break OTP signups (same sender address).
BULK_EMAIL_MAX_RECIPIENTS = 400
BULK_EMAIL_DELAY_SECONDS = 1.0

_EMAIL_RE = re.compile(r"^[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}$")


def _parse_recipients_from_upload(filename: str, content: bytes) -> list[str]:
    """Extract a deduped list of valid email addresses from an uploaded CSV or XLSX.

    Rules: looks for a header cell whose lowercased value is 'email' (or
    'e-mail' / 'email_address'). If none is found and the sheet has exactly
    one column, that column is treated as emails. Otherwise raises 400.
    """
    name = (filename or "").lower()
    rows: list[list[str]] = []

    if name.endswith(".csv"):
        text = content.decode("utf-8-sig", errors="replace")
        reader = csv.reader(io.StringIO(text))
        rows = [[(c or "").strip() for c in r] for r in reader if any((c or "").strip() for c in r)]
    elif name.endswith(".xlsx"):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise HTTPException(500, "openpyxl is not installed on the server")
        try:
            wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        except Exception as e:
            raise HTTPException(400, f"Could not read xlsx file: {e}")
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            cells = [str(c).strip() if c is not None else "" for c in row]
            if any(cells):
                rows.append(cells)
    else:
        raise HTTPException(400, "Audience file must be .csv or .xlsx")

    if not rows:
        raise HTTPException(400, "Audience file is empty")

    header = [c.lower() for c in rows[0]]
    email_col = None
    for i, h in enumerate(header):
        if h in ("email", "e-mail", "email_address", "email address"):
            email_col = i
            break

    if email_col is None:
        # No header match. If single-column file, treat first column as emails
        # (including row 0 — there was no header).
        if len(header) == 1:
            candidates = [r[0] for r in rows if r]
        else:
            raise HTTPException(
                400,
                "Could not find an 'email' column. Add a header row with an 'email' column.",
            )
    else:
        candidates = [r[email_col] for r in rows[1:] if len(r) > email_col]

    seen: set = set()
    emails: list[str] = []
    for raw in candidates:
        addr = (raw or "").strip().lower()
        if not addr or addr in seen:
            continue
        if not _EMAIL_RE.match(addr):
            continue
        seen.add(addr)
        emails.append(addr)

    if not emails:
        raise HTTPException(400, "No valid email addresses found in the file")

    return emails


def _send_bulk_emails(recipients: list[str], subject: str, body: str, is_html: bool):
    """Open one SMTP session, loop with a small delay, log per recipient."""
    sent = 0
    failed = 0
    try:
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=30) as server:
            server.ehlo()
            server.starttls()
            server.login(SMTP_USER, SMTP_PASSWORD)
            for addr in recipients:
                try:
                    msg = MIMEMultipart("alternative")
                    msg["Subject"] = subject
                    msg["From"] = SMTP_FROM
                    msg["To"] = addr
                    msg.attach(MIMEText(body, "html" if is_html else "plain", "utf-8"))
                    server.sendmail(SMTP_USER, [addr], msg.as_string())
                    sent += 1
                    print(f"[bulk-email] sent to {addr}")
                except Exception as e:
                    failed += 1
                    print(f"[bulk-email] FAILED {addr}: {e}")
                time.sleep(BULK_EMAIL_DELAY_SECONDS)
    except Exception as e:
        print(f"[bulk-email] SMTP session error: {e}")
        failed = len(recipients) - sent
    print(f"[bulk-email] DONE — sent={sent} failed={failed}")


@app.post("/api/admin/send-bulk")
async def admin_send_bulk_email(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    subject: str = Form(...),
    body: str = Form(""),
    template_file: UploadFile | None = File(None),
    admin: Admin = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    subject = (subject or "").strip()
    if not subject:
        raise HTTPException(400, "Subject is required")

    is_html = False
    final_body = (body or "").strip()
    if template_file is not None:
        tpl_bytes = await template_file.read()
        try:
            final_body = tpl_bytes.decode("utf-8", errors="replace")
        except Exception:
            raise HTTPException(400, "HTML template must be UTF-8 text")
        is_html = True

    if not final_body:
        raise HTTPException(400, "Email body (or HTML template) is required")

    audience_bytes = await file.read()
    recipients = _parse_recipients_from_upload(file.filename or "", audience_bytes)

    truncated = False
    if len(recipients) > BULK_EMAIL_MAX_RECIPIENTS:
        recipients = recipients[:BULK_EMAIL_MAX_RECIPIENTS]
        truncated = True

    background_tasks.add_task(_send_bulk_emails, recipients, subject, final_body, is_html)

    write_audit(
        db, admin.id, "bulk_email_send", "bulk_email", None,
        after={
            "subject": subject,
            "format": "html" if is_html else "text",
            "recipients_queued": len(recipients),
            "truncated_from": (len(recipients) if not truncated else None),
        },
    )

    msg = f"Queued {len(recipients)} emails."
    if truncated:
        msg += f" (truncated to per-blast cap of {BULK_EMAIL_MAX_RECIPIENTS})"

    return {
        "success": True,
        "total_contacts": len(recipients),
        "queued": len(recipients),
        "truncated": truncated,
        "max_per_blast": BULK_EMAIL_MAX_RECIPIENTS,
        "message": msg,
    }



# ============================================================================
# LAB INTEGRATION — PART 2
# ============================================================================

# AWS credentials - sourced from the environment; see .env
_AWS_KEY_ID = os.getenv("AWS_ACCESS_KEY_ID", "")
_AWS_SECRET = os.getenv("AWS_SECRET_ACCESS_KEY", "")
_DEFAULT_SG = 'sg-0fdf682a4133bb715'
_PEM_FILE = 'app/kp22.pem'

# --- Browser-desktop (noVNC) lab -------------------------------------------
# ami-0cd43456cbca48405 "Browser_ready_ubuntu" ships Ubuntu 24.04 with a VNC
# server bound to localhost and websockify/noVNC exposed on 6080. Verified by
# live probe: 6080 open ~45s after boot, /vnc.html serves the noVNC client,
# RFB security type 2 (password auth). 3389/5900/5901 are closed, so this lab
# is browser-only — there is no RDP or native-VNC path.
#
# NOTE: the AMI is owned by account 380414079427 and only *shared* with us
# (isPublic=false). If that share is revoked this lab stops launching. Copy it
# into our own account and swap the id here to remove that dependency.
_GUI_AMI = 'ami-0cd43456cbca48405'
_GUI_VNC_PORT = 6080

# The VNC password is baked into the image, so every instance shares it.
# Verified by a live RFB auth handshake against the baked-in credential.
# (Same credential the DPI Ubuntu lab uses — see _USER_DATA_UBUNTU_DPI above.)
# Overridable via env if the image is ever rebuilt with a different password.
_GUI_VNC_PASSWORD = os.getenv("GUI_LAB_VNC_PASSWORD", "")

# In-memory timer registry
active_lab_timers: dict = {}

# ============================================================================
# USER DATA SCRIPTS
# ============================================================================
_USER_DATA_AMAZON_LINUX = r"""#!/bin/bash
yum update -y
yum install -y vim
cat <<'SSHEOF' > /etc/ssh/sshd_config
Include /etc/ssh/sshd_config.d/*.conf
PermitRootLogin yes
PasswordAuthentication yes
PermitEmptyPasswords no
AuthorizedKeysFile .ssh/authorized_keys
Subsystem sftp /usr/libexec/openssh/sftp-server
AuthorizedKeysCommand /opt/aws/bin/eic_run_authorized_keys %u %f
AuthorizedKeysCommandUser ec2-instance-connect
ChallengeResponseAuthentication no
SSHEOF
useradd -d /home/ubuntu2 -m ubuntu2
echo "ubuntu2:ubuntu2" | chpasswd
echo "root:root" | chpasswd
systemctl restart sshd.service
"""

_USER_DATA_KALI_LINUX = r"""#!/bin/bash
exec > /var/log/user-data.log 2>&1
apt-get update -y
apt-get install -y vim
cat <<'SSHEOF' > /etc/ssh/sshd_config
PermitRootLogin yes
PasswordAuthentication yes
PermitEmptyPasswords no
ChallengeResponseAuthentication no
SSHEOF
useradd -m ubuntu2
echo "ubuntu2:ubuntu2" | chpasswd
echo "root:root" | chpasswd
systemctl restart sshd.service
"""

_USER_DATA_UBUNTU_DPI = r"""#!/bin/bash
set -e
sudo apt update -y
sudo apt install -y expect
sudo debconf-set-selections <<< "keyboard-configuration keyboard-configuration/layout select English (US)"
sudo debconf-set-selections <<< "keyboard-configuration keyboard-configuration/variant select English (US)"
sudo debconf-set-selections <<< "keyboard-configuration keyboard-configuration/model select Generic 105-key PC"
sudo debconf-set-selections <<< "libc6 libraries/restart-without-asking boolean true"
sudo apt install -y xfce4 xfce4-goodies
sudo apt install -y xrdp
sudo systemctl enable xrdp
sudo systemctl start xrdp
username="ubuntu"
password="__LAB_PASSWORD__"
if id "$username" &>/dev/null; then
    echo "[INFO] User '$username' exists."
else
    sudo useradd -m -s /bin/bash "$username"
fi
echo "startxfce4" > /home/$username/.xsession
sudo chown $username:$username /home/$username/.xsession
sudo chmod +x /home/$username/.xsession
sudo bash -c 'echo "startxfce4" > /etc/skel/.xsession'
sudo sed -i.bak '/^test -x \/etc\/X11\/Xsession && exec \/etc\/X11\/Xsession$/c\startxfce4' /etc/xrdp/startwm.sh
echo "$username:$password" | sudo chpasswd
sudo usermod -aG ssl-cert $username
sudo ufw allow 3389
sudo ufw reload
sudo systemctl restart xrdp
echo "root:toor" | sudo chpasswd
"""
_USER_DATA_UBUNTU_DPI = _USER_DATA_UBUNTU_DPI.replace(
    "__LAB_PASSWORD__", _GUI_VNC_PASSWORD
)


# ============================================================================
# REQUEST MODEL
# ============================================================================
class StartLabRequest(BaseModel):
    userID: int


# ============================================================================
# LAB HELPER FUNCTIONS
# ============================================================================

def _get_ec2_session():
    return boto3.Session(
        aws_access_key_id=_AWS_KEY_ID,
        aws_secret_access_key=_AWS_SECRET,
        region_name='ap-south-1',
    )


def deduct_credits_or_raise(user_id: int, lab_id: str, db: Session):
    lab = db.query(LabCatalog).filter(
        LabCatalog.lab_id == lab_id, LabCatalog.is_active == True
    ).first()
    cost = lab.credits_cost if lab else 1
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    if user.available_credits < cost:
        raise HTTPException(
            status_code=402,
            detail=f"Insufficient credits. Need {cost}, have {user.available_credits}."
        )
    user.available_credits -= cost
    db.commit()
    return cost, user.available_credits


def register_lab_start(
    user_id: int, lab_id: str, instance_ids: list,
    duration_minutes: int, ends_at: datetime,
    db: Session, background_tasks: BackgroundTasks,
):
    row = EndLabTrial(
        user_id=user_id,
        lab_id=lab_id,
        instance_ids=_json.dumps(instance_ids),
        started_at=datetime.utcnow(),
        ends_at=ends_at,
    )
    db.add(row)
    db.commit()
    background_tasks.add_task(auto_end_lab_task, user_id, float(duration_minutes * 60))


async def auto_end_lab_task(user_id: int, delay_seconds: float):
    await asyncio.sleep(delay_seconds)
    db = SessionLocal()
    try:
        row = db.query(EndLabTrial).filter(EndLabTrial.user_id == user_id).first()
        if row:
            ids = _json.loads(row.instance_ids)
            _terminate_ec2(ids)
            db.delete(row)
            db.commit()
    finally:
        db.close()


def _terminate_ec2(instance_ids: list):
    if not instance_ids:
        return
    try:
        sess = _get_ec2_session()
        ec2 = sess.resource('ec2')
        ec2.instances.filter(InstanceIds=instance_ids).terminate()
    except Exception as exc:
        print(f"EC2 termination error: {exc}")


def _decrypt_rdp_password(session, instance_id: str) -> str:
    ec2_client = session.client('ec2')
    encrypted = ''
    for _ in range(60):
        resp = ec2_client.get_password_data(InstanceId=instance_id)
        encrypted = resp.get('PasswordData', '')
        if encrypted:
            break
        import time; time.sleep(10)
    with open(_PEM_FILE, 'r') as f:
        private_key = RSA.import_key(f.read())
    cipher = PKCS1_v1_5.new(private_key)
    return cipher.decrypt(base64.b64decode(encrypted), None).decode('utf-8')


def _wait_windows(session, instance) -> dict:
    instance.wait_until_running()
    ec2_client = session.client('ec2')
    ec2_client.get_waiter('instance_status_ok').wait(InstanceIds=[instance.id])
    password = _decrypt_rdp_password(session, instance.id)
    instance.load()
    return {
        'instance_id': instance.id,
        'public_ip': instance.public_ip_address,
        'private_ip': instance.private_ip_address,
        'os_type': 'windows',
        'username': 'Administrator',
        'password': password,
        'ssh_command': None,
    }


def _wait_linux(session, instance, username='ubuntu2', password='ubuntu2') -> dict:
    instance.wait_until_running()
    ec2_client = session.client('ec2')
    ec2_client.get_waiter('instance_status_ok').wait(InstanceIds=[instance.id])
    instance.load()
    public_ip = instance.public_ip_address
    return {
        'instance_id': instance.id,
        'public_ip': public_ip,
        'private_ip': instance.private_ip_address,
        'os_type': 'linux',
        'username': username,
        'password': password,
        'ssh_command': f'ssh {username}@{public_ip}',
    }


def _wait_novnc(session, instance) -> dict:
    """Wait for a browser-desktop instance and return its noVNC connect info.

    Same wait as _wait_linux (running + instance_status_ok), but the useful
    credential is a URL rather than an SSH command: the desktop is reached at
    http://<ip>:6080/vnc.html in any browser. websockify starts within ~45s of
    boot, comfortably inside the status_ok wait.
    """
    instance.wait_until_running()
    ec2_client = session.client('ec2')
    ec2_client.get_waiter('instance_status_ok').wait(InstanceIds=[instance.id])
    instance.load()
    public_ip = instance.public_ip_address
    return {
        'instance_id': instance.id,
        'public_ip': public_ip,
        'private_ip': instance.private_ip_address,
        'os_type': 'linux',
        'access_type': 'novnc',
        'gui_url': f'http://{public_ip}:{_GUI_VNC_PORT}/vnc.html',
        'username': None,
        'password': _GUI_VNC_PASSWORD or None,
        'ssh_command': None,
    }


def _iso_z(dt):
    """Serialize a datetime as a valid UTC ISO string ending in 'Z'.

    Handles both naive (datetime.utcnow()) and timezone-aware values (the
    timestamptz columns return aware datetimes). Without this, doing
    `.isoformat() + "Z"` on an aware value yields an invalid `...+00:00Z`
    string that JS `new Date()` parses as Invalid Date.
    """
    if dt is None:
        return None
    if dt.tzinfo is not None:
        dt = dt.astimezone(timezone.utc).replace(tzinfo=None)
    return dt.isoformat() + "Z"


def _build_response(lab_id, started_at, ends_at, cost, remaining, instances_list):
    return {
        "success": True,
        "lab_id": lab_id,
        "started_at": _iso_z(started_at),
        "ends_at": _iso_z(ends_at),
        "credits_deducted": cost,
        "credits_remaining": remaining,
        "instances": instances_list,
    }


# ============================================================================
# LAB CATALOG + SESSION API ROUTES
# ============================================================================

@app.get("/api/labs")
def get_lab_catalog(db: Session = Depends(get_db)):
    labs = db.query(LabCatalog).filter(LabCatalog.is_active == True).all()
    return [{
        "lab_id": l.lab_id,
        "slug": l.slug,
        "name": l.name,
        "category": l.category,
        "difficulty": l.difficulty,
        "credits_cost": l.credits_cost,
        "duration_minutes": l.duration_minutes,
        "os_type": l.os_type,
        "description": l.description,
    } for l in labs]


def _course_row(c: CourseCatalog) -> dict:
    return {
        "id": c.id,
        "slug": c.slug,
        "title": c.title,
        "tagline": c.tagline,
        "description": c.description,
        "category": c.category,
        "difficulty": c.difficulty,
        "modules_count": c.modules_count,
        "labs_count": c.labs_count,
        "duration_hours": c.duration_hours,
        "price_inr": float(c.price_inr) if c.price_inr is not None else None,
        "currency": c.currency,
        "billing_label": c.billing_label,
        "hero_image_url": c.hero_image_url,
        "accent_color": c.accent_color,
        "audience": c.audience or [],
        "benefits": c.benefits or [],
        "syllabus": c.syllabus or [],
        "is_active": c.is_active,
        "display_order": c.display_order,
        "updated_at": _iso_z(c.updated_at),
    }


@app.get("/api/courses")
def get_courses(db: Session = Depends(get_db)):
    """Public course catalog. Active rows only, ordered for display."""
    rows = (
        db.query(CourseCatalog)
        .filter(CourseCatalog.is_active == True)
        .order_by(CourseCatalog.display_order, CourseCatalog.id)
        .all()
    )
    return [_course_row(c) for c in rows]


@app.get("/api/courses/{slug}")
def get_course_by_slug(slug: str, db: Session = Depends(get_db)):
    course = (
        db.query(CourseCatalog)
        .filter(CourseCatalog.slug == slug, CourseCatalog.is_active == True)
        .first()
    )
    if not course:
        raise HTTPException(status_code=404, detail="Course not found")
    return _course_row(course)


# --- Course purchases (Part 8b) ---------------------------------------------
# Replaces the previous Descp -> /subscription redirect. User clicks Enroll on
# a course card; frontend asks for an order, Razorpay popup opens with the
# GST-inclusive amount, verify endpoint signs off and records the purchase.

@app.post("/api/courses/{slug}/create-order")
def create_course_order(
    slug: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    course = (
        db.query(CourseCatalog)
        .filter(CourseCatalog.slug == slug, CourseCatalog.is_active == True)
        .first()
    )
    if not course:
        raise HTTPException(status_code=404, detail="Course not found")
    if course.price_inr is None or float(course.price_inr) <= 0:
        raise HTTPException(status_code=400, detail="This course has no purchasable price set")

    base = float(course.price_inr)
    total_inclusive = with_gst(base)               # rupees, rounded — matches frontend display
    amount_paise = total_inclusive * 100

    order = razorpay_client.order.create({
        "amount": amount_paise,
        "currency": (course.currency or "INR"),
        "receipt": f"course_{current_user.id}_{slug}",
        "notes": {
            "user_id": str(current_user.id),
            "course_slug": slug,
            "course_id": str(course.id),
        },
    })
    return {
        "order_id": order["id"],
        "amount": amount_paise,            # paise — matches what Razorpay charges
        "currency": (course.currency or "INR"),
        "course": {
            "slug": course.slug,
            "title": course.title,
            "base_price": base,
            "gst_amount": total_inclusive - base,
            "total_amount": total_inclusive,
        },
    }


@app.post("/api/courses/verify-payment")
def verify_course_payment(
    req: CourseVerifyRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    body = f"{req.razorpay_order_id}|{req.razorpay_payment_id}"
    expected_signature = hmac.new(
        key=os.getenv("RAZORPAY_KEY_SECRET", "").encode("utf-8"),
        msg=body.encode("utf-8"),
        digestmod=hashlib.sha256,
    ).hexdigest()
    if expected_signature != req.razorpay_signature:
        raise HTTPException(status_code=400, detail="Invalid payment signature")

    course = (
        db.query(CourseCatalog)
        .filter(CourseCatalog.slug == req.slug)
        .first()
    )
    if not course:
        raise HTTPException(status_code=404, detail="Course not found")

    # Idempotency: if the same order has already been recorded, return success
    # without inserting a duplicate row. Razorpay can retry the handler if the
    # popup is closed mid-callback.
    existing = (
        db.query(CoursePurchase)
        .filter(CoursePurchase.razorpay_order_id == req.razorpay_order_id)
        .first()
    )
    if existing:
        return {
            "success": True,
            "duplicate": True,
            "purchase_id": existing.id,
            "course_slug": existing.course_slug,
        }

    _grant_referral_reward_on_first_purchase(current_user.id, db)
    base = float(course.price_inr or 0)
    total_inclusive = with_gst(base)
    purchase = CoursePurchase(
        user_id=current_user.id,
        course_id=course.id,
        course_slug=course.slug,
        course_title=course.title,
        base_price=base,
        gst_amount=total_inclusive - base,
        total_amount=total_inclusive,
        razorpay_order_id=req.razorpay_order_id,
        razorpay_payment_id=req.razorpay_payment_id,
    )
    db.add(purchase)
    try:
        db.commit()
    except IntegrityError:
        # Concurrent verify-payment race: another request already inserted by the
        # time we did. UNIQUE(razorpay_order_id) caught it — recover gracefully.
        db.rollback()
        existing = db.query(CoursePurchase).filter(
            CoursePurchase.razorpay_order_id == req.razorpay_order_id
        ).first()
        if existing:
            return {
                "success": True,
                "duplicate": True,
                "purchase_id": existing.id,
                "course_slug": existing.course_slug,
            }
        raise HTTPException(status_code=409, detail="Conflicting purchase record")
    db.refresh(purchase)
    return {
        "success": True,
        "purchase_id": purchase.id,
        "course_slug": course.slug,
        "total_amount": total_inclusive,
    }
    
@app.get("/api/bootcamps/{slug}")
def get_bootcamp_by_slug(slug: str, db: Session = Depends(get_db)):
    bootcamp = db.query(BootcampCatalog).filter(BootcampCatalog.slug == slug, BootcampCatalog.is_active == True).first()
    if not bootcamp:
        raise HTTPException(status_code=404, detail="Bootcamp not found")
    return _bootcamp_row(bootcamp)

@app.post("/api/bootcamps/{slug}/create-order")
def create_bootcamp_order(slug: str, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    bootcamp = db.query(BootcampCatalog).filter(BootcampCatalog.slug == slug, BootcampCatalog.is_active == True).first()
    if not bootcamp:
        raise HTTPException(status_code=404, detail="Bootcamp not found")
    if bootcamp.price_inr is None or float(bootcamp.price_inr) <= 0:
        raise HTTPException(status_code=400, detail="This bootcamp has no purchasable price set")

    base = float(bootcamp.price_inr)
    total_inclusive = with_gst(base)               
    amount_paise = total_inclusive * 100

    order = razorpay_client.order.create({
        "amount": amount_paise,
        "currency": (bootcamp.currency or "INR"),
        "receipt": f"bootcamp_{current_user.id}_{slug}",
        "notes": {
            "user_id": str(current_user.id),
            "bootcamp_slug": slug,
            "bootcamp_id": str(bootcamp.id),
        },
    })
    return {
        "order_id": order["id"],
        "amount": amount_paise,
        "currency": (bootcamp.currency or "INR"),
        "bootcamp": {
            "slug": bootcamp.slug, "title": bootcamp.title, "base_price": base,
            "gst_amount": total_inclusive - base, "total_amount": total_inclusive,
        },
    }

@app.post("/api/bootcamps/verify-payment")
def verify_bootcamp_payment(req: BootcampVerifyRequest, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    body = f"{req.razorpay_order_id}|{req.razorpay_payment_id}"
    expected_signature = hmac.new(
        key=os.getenv("RAZORPAY_KEY_SECRET", "").encode("utf-8"),
        msg=body.encode("utf-8"),
        digestmod=hashlib.sha256,
    ).hexdigest()
    
    if expected_signature != req.razorpay_signature:
        raise HTTPException(status_code=400, detail="Invalid payment signature")

    bootcamp = db.query(BootcampCatalog).filter(BootcampCatalog.slug == req.slug).first()
    if not bootcamp:
        raise HTTPException(status_code=404, detail="Bootcamp not found")

    existing = db.query(BootcampPurchase).filter(BootcampPurchase.razorpay_order_id == req.razorpay_order_id).first()
    if existing:
        return {"success": True, "duplicate": True, "purchase_id": existing.id, "bootcamp_slug": existing.bootcamp_slug}

    _grant_referral_reward_on_first_purchase(current_user.id, db)
    base = float(bootcamp.price_inr or 0)
    total_inclusive = with_gst(base)
    purchase = BootcampPurchase(
        user_id=current_user.id, bootcamp_id=bootcamp.id, bootcamp_slug=bootcamp.slug,
        bootcamp_title=bootcamp.title, base_price=base, gst_amount=total_inclusive - base, total_amount=total_inclusive,
        razorpay_order_id=req.razorpay_order_id, razorpay_payment_id=req.razorpay_payment_id,
    )
    db.add(purchase)
    try:
        db.commit()
    except IntegrityError:
        # Two concurrent verify-payment calls both passed the existence check
        # and raced the INSERT. UNIQUE(razorpay_order_id) wins; recover by
        # returning the existing row instead of bubbling a 500.
        db.rollback()
        existing = db.query(BootcampPurchase).filter(
            BootcampPurchase.razorpay_order_id == req.razorpay_order_id
        ).first()
        if existing:
            return {"success": True, "duplicate": True, "purchase_id": existing.id, "bootcamp_slug": existing.bootcamp_slug}
        raise HTTPException(status_code=409, detail="Conflicting purchase record")
    db.refresh(purchase)
    return {"success": True, "purchase_id": purchase.id, "bootcamp_slug": bootcamp.slug, "total_amount": total_inclusive}


@app.get("/api/active-lab/{user_id}")
def get_active_lab(
    user_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    row = db.query(EndLabTrial).filter(EndLabTrial.user_id == user_id).first()
    if not row:
        return None
    lab = db.query(LabCatalog).filter(LabCatalog.lab_id == row.lab_id).first()
    return {
        "lab_id": row.lab_id,
        "lab_name": lab.name if lab else row.lab_id,
        "started_at": _iso_z(row.started_at),
        "ends_at": _iso_z(row.ends_at),
        "instance_ids": _json.loads(row.instance_ids),
        "instances": [],
    }


@app.post("/api/end-lab")
def end_lab_route(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    user_id = current_user.id
    row = db.query(EndLabTrial).filter(EndLabTrial.user_id == user_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="No active lab.")
    ids = _json.loads(row.instance_ids)
    _terminate_ec2(ids)
    db.delete(row)
    db.commit()
    return {"success": True, "message": "Lab ended."}


@app.post("/api/extend-lab")
def extend_lab_route(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    user_id = current_user.id
    row = db.query(EndLabTrial).filter(EndLabTrial.user_id == user_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="No active lab.")
    if current_user.available_credits < 1:
        raise HTTPException(status_code=402, detail="Insufficient credits to extend.")
    current_user.available_credits -= 1
    row.ends_at = (row.ends_at or datetime.utcnow()) + timedelta(minutes=60)
    db.commit()
    return {
        "success": True,
        "ends_at": _iso_z(row.ends_at),
        "credits_remaining": current_user.available_credits,
    }


# ============================================================================
# START-LAB ENDPOINTS — 38 labs
# ============================================================================

# --- osf1: Windows, t2.micro, ami-05f035b6ae19778b1, duration=60 ---
@app.post("/api/start-lab-csfc-osf1")
async def start_lab_csfc_osf1(
    req: StartLabRequest, background_tasks: BackgroundTasks,
    db: Session = Depends(get_db), _cu: User = Depends(get_current_user),
):
    uid = req.userID
    if db.query(EndLabTrial).filter(EndLabTrial.user_id == uid).first():
        raise HTTPException(400, "Another lab is currently running.")
    cost, rem = deduct_credits_or_raise(uid, "csfc-osf1", db)
    sess = _get_ec2_session(); ec2 = sess.resource('ec2')
    inst = ec2.create_instances(
        ImageId='ami-05f035b6ae19778b1', MinCount=1, MaxCount=1,
        InstanceType='t2.micro', KeyName='kp22',
        SecurityGroupIds=[_DEFAULT_SG], UserData=_USER_DATA_AMAZON_LINUX,
    )[0]
    creds = _wait_windows(sess, inst)
    now = datetime.utcnow(); ends = now + timedelta(minutes=60)
    register_lab_start(uid, "csfc-osf1", [inst.id], 60, ends, db, background_tasks)
    return _build_response("csfc-osf1", now, ends, cost, rem, [creds])


# --- osf2: Windows, t2.small, ami-05f035b6ae19778b1, duration=60 ---
@app.post("/api/start-lab-csfc-osf2")
async def start_lab_csfc_osf2(
    req: StartLabRequest, background_tasks: BackgroundTasks,
    db: Session = Depends(get_db), _cu: User = Depends(get_current_user),
):
    uid = req.userID
    if db.query(EndLabTrial).filter(EndLabTrial.user_id == uid).first():
        raise HTTPException(400, "Another lab is currently running.")
    cost, rem = deduct_credits_or_raise(uid, "csfc-osf2", db)
    sess = _get_ec2_session(); ec2 = sess.resource('ec2')
    inst = ec2.create_instances(
        ImageId='ami-05f035b6ae19778b1', MinCount=1, MaxCount=1,
        InstanceType='t2.small', KeyName='kp22',
        SecurityGroupIds=[_DEFAULT_SG], UserData=_USER_DATA_AMAZON_LINUX,
    )[0]
    creds = _wait_windows(sess, inst)
    now = datetime.utcnow(); ends = now + timedelta(minutes=60)
    register_lab_start(uid, "csfc-osf2", [inst.id], 60, ends, db, background_tasks)
    return _build_response("csfc-osf2", now, ends, cost, rem, [creds])


# --- osf3: Linux, t2.micro, ami-0e1d06225679bc1c5, 3 instances, duration=60 ---
@app.post("/api/start-lab-csfc-osf3")
async def start_lab_csfc_osf3(
    req: StartLabRequest, background_tasks: BackgroundTasks,
    db: Session = Depends(get_db), _cu: User = Depends(get_current_user),
):
    uid = req.userID
    if db.query(EndLabTrial).filter(EndLabTrial.user_id == uid).first():
        raise HTTPException(400, "Another lab is currently running.")
    cost, rem = deduct_credits_or_raise(uid, "csfc-osf3", db)
    sess = _get_ec2_session(); ec2 = sess.resource('ec2')
    instances = ec2.create_instances(
        ImageId='ami-0e1d06225679bc1c5', MinCount=3, MaxCount=3,
        InstanceType='t2.micro', KeyName='kkp1',
        SecurityGroupIds=[_DEFAULT_SG], UserData=_USER_DATA_AMAZON_LINUX,
    )
    creds_list = [_wait_linux(sess, i) for i in instances]
    ids = [i.id for i in instances]
    now = datetime.utcnow(); ends = now + timedelta(minutes=60)
    register_lab_start(uid, "csfc-osf3", ids, 60, ends, db, background_tasks)
    return _build_response("csfc-osf3", now, ends, cost, rem, creds_list)


# --- osf4-osf9: Linux, t2.micro, ami-0e1d06225679bc1c5, single instance ---
@app.post("/api/start-lab-csfc-osf4")
async def start_lab_csfc_osf4(
    req: StartLabRequest, background_tasks: BackgroundTasks,
    db: Session = Depends(get_db), _cu: User = Depends(get_current_user),
):
    uid = req.userID
    if db.query(EndLabTrial).filter(EndLabTrial.user_id == uid).first():
        raise HTTPException(400, "Another lab is currently running.")
    cost, rem = deduct_credits_or_raise(uid, "csfc-osf4", db)
    sess = _get_ec2_session(); ec2 = sess.resource('ec2')
    inst = ec2.create_instances(
        ImageId='ami-0e1d06225679bc1c5', MinCount=1, MaxCount=1,
        InstanceType='t2.micro', KeyName='kkp1',
        SecurityGroupIds=[_DEFAULT_SG], UserData=_USER_DATA_AMAZON_LINUX,
    )[0]
    creds = _wait_linux(sess, inst)
    now = datetime.utcnow(); ends = now + timedelta(minutes=60)
    register_lab_start(uid, "csfc-osf4", [inst.id], 60, ends, db, background_tasks)
    return _build_response("csfc-osf4", now, ends, cost, rem, [creds])


@app.post("/api/start-lab-csfc-osf5")
async def start_lab_csfc_osf5(
    req: StartLabRequest, background_tasks: BackgroundTasks,
    db: Session = Depends(get_db), _cu: User = Depends(get_current_user),
):
    uid = req.userID
    if db.query(EndLabTrial).filter(EndLabTrial.user_id == uid).first():
        raise HTTPException(400, "Another lab is currently running.")
    cost, rem = deduct_credits_or_raise(uid, "csfc-osf5", db)
    sess = _get_ec2_session(); ec2 = sess.resource('ec2')
    inst = ec2.create_instances(
        ImageId='ami-0e1d06225679bc1c5', MinCount=1, MaxCount=1,
        InstanceType='t2.micro', KeyName='kkp1',
        SecurityGroupIds=[_DEFAULT_SG], UserData=_USER_DATA_AMAZON_LINUX,
    )[0]
    creds = _wait_linux(sess, inst)
    now = datetime.utcnow(); ends = now + timedelta(minutes=60)
    register_lab_start(uid, "csfc-osf5", [inst.id], 60, ends, db, background_tasks)
    return _build_response("csfc-osf5", now, ends, cost, rem, [creds])


@app.post("/api/start-lab-csfc-osf6")
async def start_lab_csfc_osf6(
    req: StartLabRequest, background_tasks: BackgroundTasks,
    db: Session = Depends(get_db), _cu: User = Depends(get_current_user),
):
    uid = req.userID
    if db.query(EndLabTrial).filter(EndLabTrial.user_id == uid).first():
        raise HTTPException(400, "Another lab is currently running.")
    cost, rem = deduct_credits_or_raise(uid, "csfc-osf6", db)
    sess = _get_ec2_session(); ec2 = sess.resource('ec2')
    inst = ec2.create_instances(
        ImageId='ami-0e1d06225679bc1c5', MinCount=1, MaxCount=1,
        InstanceType='t2.micro', KeyName='kkp1',
        SecurityGroupIds=[_DEFAULT_SG], UserData=_USER_DATA_AMAZON_LINUX,
    )[0]
    creds = _wait_linux(sess, inst)
    now = datetime.utcnow(); ends = now + timedelta(minutes=60)
    register_lab_start(uid, "csfc-osf6", [inst.id], 60, ends, db, background_tasks)
    return _build_response("csfc-osf6", now, ends, cost, rem, [creds])


@app.post("/api/start-lab-csfc-osf7")
async def start_lab_csfc_osf7(
    req: StartLabRequest, background_tasks: BackgroundTasks,
    db: Session = Depends(get_db), _cu: User = Depends(get_current_user),
):
    uid = req.userID
    if db.query(EndLabTrial).filter(EndLabTrial.user_id == uid).first():
        raise HTTPException(400, "Another lab is currently running.")
    cost, rem = deduct_credits_or_raise(uid, "csfc-osf7", db)
    sess = _get_ec2_session(); ec2 = sess.resource('ec2')
    inst = ec2.create_instances(
        ImageId='ami-0e1d06225679bc1c5', MinCount=1, MaxCount=1,
        InstanceType='t2.micro', KeyName='kkp1',
        SecurityGroupIds=[_DEFAULT_SG], UserData=_USER_DATA_AMAZON_LINUX,
    )[0]
    creds = _wait_linux(sess, inst)
    now = datetime.utcnow(); ends = now + timedelta(minutes=60)
    register_lab_start(uid, "csfc-osf7", [inst.id], 60, ends, db, background_tasks)
    return _build_response("csfc-osf7", now, ends, cost, rem, [creds])


@app.post("/api/start-lab-csfc-osf8")
async def start_lab_csfc_osf8(
    req: StartLabRequest, background_tasks: BackgroundTasks,
    db: Session = Depends(get_db), _cu: User = Depends(get_current_user),
):
    uid = req.userID
    if db.query(EndLabTrial).filter(EndLabTrial.user_id == uid).first():
        raise HTTPException(400, "Another lab is currently running.")
    cost, rem = deduct_credits_or_raise(uid, "csfc-osf8", db)
    sess = _get_ec2_session(); ec2 = sess.resource('ec2')
    inst = ec2.create_instances(
        ImageId='ami-0e1d06225679bc1c5', MinCount=1, MaxCount=1,
        InstanceType='t2.micro', KeyName='kkp1',
        SecurityGroupIds=[_DEFAULT_SG], UserData=_USER_DATA_AMAZON_LINUX,
    )[0]
    creds = _wait_linux(sess, inst)
    now = datetime.utcnow(); ends = now + timedelta(minutes=60)
    register_lab_start(uid, "csfc-osf8", [inst.id], 60, ends, db, background_tasks)
    return _build_response("csfc-osf8", now, ends, cost, rem, [creds])


@app.post("/api/start-lab-csfc-osf9")
async def start_lab_csfc_osf9(
    req: StartLabRequest, background_tasks: BackgroundTasks,
    db: Session = Depends(get_db), _cu: User = Depends(get_current_user),
):
    uid = req.userID
    if db.query(EndLabTrial).filter(EndLabTrial.user_id == uid).first():
        raise HTTPException(400, "Another lab is currently running.")
    cost, rem = deduct_credits_or_raise(uid, "csfc-osf9", db)
    sess = _get_ec2_session(); ec2 = sess.resource('ec2')
    inst = ec2.create_instances(
        ImageId='ami-0e1d06225679bc1c5', MinCount=1, MaxCount=1,
        InstanceType='t2.micro', KeyName='kkp1',
        SecurityGroupIds=[_DEFAULT_SG], UserData=_USER_DATA_AMAZON_LINUX,
    )[0]
    creds = _wait_linux(sess, inst)
    now = datetime.utcnow(); ends = now + timedelta(minutes=90)
    register_lab_start(uid, "csfc-osf9", [inst.id], 90, ends, db, background_tasks)
    return _build_response("csfc-osf9", now, ends, cost, rem, [creds])


# --- nf1-nf9: Windows, t2.micro, ami-05f035b6ae19778b1 ---
@app.post("/api/start-lab-csfc-nf1")
async def start_lab_csfc_nf1(
    req: StartLabRequest, background_tasks: BackgroundTasks,
    db: Session = Depends(get_db), _cu: User = Depends(get_current_user),
):
    uid = req.userID
    if db.query(EndLabTrial).filter(EndLabTrial.user_id == uid).first():
        raise HTTPException(400, "Another lab is currently running.")
    cost, rem = deduct_credits_or_raise(uid, "csfc-nf1", db)
    sess = _get_ec2_session(); ec2 = sess.resource('ec2')
    inst = ec2.create_instances(
        ImageId='ami-05f035b6ae19778b1', MinCount=1, MaxCount=1,
        InstanceType='t2.micro', KeyName='kp22',
        SecurityGroupIds=[_DEFAULT_SG], UserData=_USER_DATA_AMAZON_LINUX,
    )[0]
    creds = _wait_windows(sess, inst)
    now = datetime.utcnow(); ends = now + timedelta(minutes=90)
    register_lab_start(uid, "csfc-nf1", [inst.id], 90, ends, db, background_tasks)
    return _build_response("csfc-nf1", now, ends, cost, rem, [creds])


@app.post("/api/start-lab-csfc-nf2")
async def start_lab_csfc_nf2(
    req: StartLabRequest, background_tasks: BackgroundTasks,
    db: Session = Depends(get_db), _cu: User = Depends(get_current_user),
):
    uid = req.userID
    if db.query(EndLabTrial).filter(EndLabTrial.user_id == uid).first():
        raise HTTPException(400, "Another lab is currently running.")
    cost, rem = deduct_credits_or_raise(uid, "csfc-nf2", db)
    sess = _get_ec2_session(); ec2 = sess.resource('ec2')
    inst = ec2.create_instances(
        ImageId='ami-05f035b6ae19778b1', MinCount=1, MaxCount=1,
        InstanceType='t2.micro', KeyName='kp22',
        SecurityGroupIds=[_DEFAULT_SG], UserData=_USER_DATA_AMAZON_LINUX,
    )[0]
    creds = _wait_windows(sess, inst)
    now = datetime.utcnow(); ends = now + timedelta(minutes=65)
    register_lab_start(uid, "csfc-nf2", [inst.id], 65, ends, db, background_tasks)
    return _build_response("csfc-nf2", now, ends, cost, rem, [creds])


@app.post("/api/start-lab-csfc-nf3")
async def start_lab_csfc_nf3(
    req: StartLabRequest, background_tasks: BackgroundTasks,
    db: Session = Depends(get_db), _cu: User = Depends(get_current_user),
):
    uid = req.userID
    if db.query(EndLabTrial).filter(EndLabTrial.user_id == uid).first():
        raise HTTPException(400, "Another lab is currently running.")
    cost, rem = deduct_credits_or_raise(uid, "csfc-nf3", db)
    sess = _get_ec2_session(); ec2 = sess.resource('ec2')
    inst = ec2.create_instances(
        ImageId='ami-05f035b6ae19778b1', MinCount=1, MaxCount=1,
        InstanceType='t2.micro', KeyName='kp22',
        SecurityGroupIds=[_DEFAULT_SG], UserData=_USER_DATA_AMAZON_LINUX,
    )[0]
    creds = _wait_windows(sess, inst)
    now = datetime.utcnow(); ends = now + timedelta(minutes=65)
    register_lab_start(uid, "csfc-nf3", [inst.id], 65, ends, db, background_tasks)
    return _build_response("csfc-nf3", now, ends, cost, rem, [creds])


@app.post("/api/start-lab-csfc-nf4")
async def start_lab_csfc_nf4(
    req: StartLabRequest, background_tasks: BackgroundTasks,
    db: Session = Depends(get_db), _cu: User = Depends(get_current_user),
):
    uid = req.userID
    if db.query(EndLabTrial).filter(EndLabTrial.user_id == uid).first():
        raise HTTPException(400, "Another lab is currently running.")
    cost, rem = deduct_credits_or_raise(uid, "csfc-nf4", db)
    sess = _get_ec2_session(); ec2 = sess.resource('ec2')
    inst = ec2.create_instances(
        ImageId='ami-05f035b6ae19778b1', MinCount=1, MaxCount=1,
        InstanceType='t2.micro', KeyName='kp22',
        SecurityGroupIds=[_DEFAULT_SG], UserData=_USER_DATA_AMAZON_LINUX,
    )[0]
    creds = _wait_windows(sess, inst)
    now = datetime.utcnow(); ends = now + timedelta(minutes=110)
    register_lab_start(uid, "csfc-nf4", [inst.id], 110, ends, db, background_tasks)
    return _build_response("csfc-nf4", now, ends, cost, rem, [creds])


@app.post("/api/start-lab-csfc-nf5")
async def start_lab_csfc_nf5(
    req: StartLabRequest, background_tasks: BackgroundTasks,
    db: Session = Depends(get_db), _cu: User = Depends(get_current_user),
):
    uid = req.userID
    if db.query(EndLabTrial).filter(EndLabTrial.user_id == uid).first():
        raise HTTPException(400, "Another lab is currently running.")
    cost, rem = deduct_credits_or_raise(uid, "csfc-nf5", db)
    sess = _get_ec2_session(); ec2 = sess.resource('ec2')
    inst = ec2.create_instances(
        ImageId='ami-05f035b6ae19778b1', MinCount=1, MaxCount=1,
        InstanceType='t2.micro', KeyName='kp22',
        SecurityGroupIds=[_DEFAULT_SG], UserData=_USER_DATA_AMAZON_LINUX,
    )[0]
    creds = _wait_windows(sess, inst)
    now = datetime.utcnow(); ends = now + timedelta(minutes=65)
    register_lab_start(uid, "csfc-nf5", [inst.id], 65, ends, db, background_tasks)
    return _build_response("csfc-nf5", now, ends, cost, rem, [creds])


@app.post("/api/start-lab-csfc-nf6")
async def start_lab_csfc_nf6(
    req: StartLabRequest, background_tasks: BackgroundTasks,
    db: Session = Depends(get_db), _cu: User = Depends(get_current_user),
):
    uid = req.userID
    if db.query(EndLabTrial).filter(EndLabTrial.user_id == uid).first():
        raise HTTPException(400, "Another lab is currently running.")
    cost, rem = deduct_credits_or_raise(uid, "csfc-nf6", db)
    sess = _get_ec2_session(); ec2 = sess.resource('ec2')
    inst = ec2.create_instances(
        ImageId='ami-05f035b6ae19778b1', MinCount=1, MaxCount=1,
        InstanceType='t2.micro', KeyName='kp22',
        SecurityGroupIds=[_DEFAULT_SG], UserData=_USER_DATA_AMAZON_LINUX,
    )[0]
    creds = _wait_windows(sess, inst)
    now = datetime.utcnow(); ends = now + timedelta(minutes=50)
    register_lab_start(uid, "csfc-nf6", [inst.id], 50, ends, db, background_tasks)
    return _build_response("csfc-nf6", now, ends, cost, rem, [creds])


@app.post("/api/start-lab-csfc-nf7")
async def start_lab_csfc_nf7(
    req: StartLabRequest, background_tasks: BackgroundTasks,
    db: Session = Depends(get_db), _cu: User = Depends(get_current_user),
):
    uid = req.userID
    if db.query(EndLabTrial).filter(EndLabTrial.user_id == uid).first():
        raise HTTPException(400, "Another lab is currently running.")
    cost, rem = deduct_credits_or_raise(uid, "csfc-nf7", db)
    sess = _get_ec2_session(); ec2 = sess.resource('ec2')
    inst = ec2.create_instances(
        ImageId='ami-05f035b6ae19778b1', MinCount=1, MaxCount=1,
        InstanceType='t2.micro', KeyName='kp22',
        SecurityGroupIds=[_DEFAULT_SG], UserData=_USER_DATA_AMAZON_LINUX,
    )[0]
    creds = _wait_windows(sess, inst)
    now = datetime.utcnow(); ends = now + timedelta(minutes=80)
    register_lab_start(uid, "csfc-nf7", [inst.id], 80, ends, db, background_tasks)
    return _build_response("csfc-nf7", now, ends, cost, rem, [creds])


@app.post("/api/start-lab-csfc-nf8")
async def start_lab_csfc_nf8(
    req: StartLabRequest, background_tasks: BackgroundTasks,
    db: Session = Depends(get_db), _cu: User = Depends(get_current_user),
):
    uid = req.userID
    if db.query(EndLabTrial).filter(EndLabTrial.user_id == uid).first():
        raise HTTPException(400, "Another lab is currently running.")
    cost, rem = deduct_credits_or_raise(uid, "csfc-nf8", db)
    sess = _get_ec2_session(); ec2 = sess.resource('ec2')
    inst = ec2.create_instances(
        ImageId='ami-05f035b6ae19778b1', MinCount=1, MaxCount=1,
        InstanceType='t2.micro', KeyName='kp22',
        SecurityGroupIds=[_DEFAULT_SG], UserData=_USER_DATA_AMAZON_LINUX,
    )[0]
    creds = _wait_windows(sess, inst)
    now = datetime.utcnow(); ends = now + timedelta(minutes=80)
    register_lab_start(uid, "csfc-nf8", [inst.id], 80, ends, db, background_tasks)
    return _build_response("csfc-nf8", now, ends, cost, rem, [creds])


@app.post("/api/start-lab-csfc-nf9")
async def start_lab_csfc_nf9(
    req: StartLabRequest, background_tasks: BackgroundTasks,
    db: Session = Depends(get_db), _cu: User = Depends(get_current_user),
):
    uid = req.userID
    if db.query(EndLabTrial).filter(EndLabTrial.user_id == uid).first():
        raise HTTPException(400, "Another lab is currently running.")
    cost, rem = deduct_credits_or_raise(uid, "csfc-nf9", db)
    sess = _get_ec2_session(); ec2 = sess.resource('ec2')
    inst = ec2.create_instances(
        ImageId='ami-05f035b6ae19778b1', MinCount=1, MaxCount=1,
        InstanceType='t2.micro', KeyName='kp22',
        SecurityGroupIds=[_DEFAULT_SG], UserData=_USER_DATA_AMAZON_LINUX,
    )[0]
    creds = _wait_windows(sess, inst)
    now = datetime.utcnow(); ends = now + timedelta(minutes=80)
    register_lab_start(uid, "csfc-nf9", [inst.id], 80, ends, db, background_tasks)
    return _build_response("csfc-nf9", now, ends, cost, rem, [creds])


# --- naf1-naf3: Windows, t2.micro, ami-05f035b6ae19778b1 ---
@app.post("/api/start-lab-csfc-naf1")
async def start_lab_csfc_naf1(
    req: StartLabRequest, background_tasks: BackgroundTasks,
    db: Session = Depends(get_db), _cu: User = Depends(get_current_user),
):
    uid = req.userID
    if db.query(EndLabTrial).filter(EndLabTrial.user_id == uid).first():
        raise HTTPException(400, "Another lab is currently running.")
    cost, rem = deduct_credits_or_raise(uid, "csfc-naf1", db)
    sess = _get_ec2_session(); ec2 = sess.resource('ec2')
    inst = ec2.create_instances(
        ImageId='ami-05f035b6ae19778b1', MinCount=1, MaxCount=1,
        InstanceType='t2.micro', KeyName='kp22',
        SecurityGroupIds=[_DEFAULT_SG], UserData=_USER_DATA_AMAZON_LINUX,
    )[0]
    creds = _wait_windows(sess, inst)
    now = datetime.utcnow(); ends = now + timedelta(minutes=140)
    register_lab_start(uid, "csfc-naf1", [inst.id], 140, ends, db, background_tasks)
    return _build_response("csfc-naf1", now, ends, cost, rem, [creds])


@app.post("/api/start-lab-csfc-naf2")
async def start_lab_csfc_naf2(
    req: StartLabRequest, background_tasks: BackgroundTasks,
    db: Session = Depends(get_db), _cu: User = Depends(get_current_user),
):
    uid = req.userID
    if db.query(EndLabTrial).filter(EndLabTrial.user_id == uid).first():
        raise HTTPException(400, "Another lab is currently running.")
    cost, rem = deduct_credits_or_raise(uid, "csfc-naf2", db)
    sess = _get_ec2_session(); ec2 = sess.resource('ec2')
    inst = ec2.create_instances(
        ImageId='ami-05f035b6ae19778b1', MinCount=1, MaxCount=1,
        InstanceType='t2.micro', KeyName='kp22',
        SecurityGroupIds=[_DEFAULT_SG], UserData=_USER_DATA_AMAZON_LINUX,
    )[0]
    creds = _wait_windows(sess, inst)
    now = datetime.utcnow(); ends = now + timedelta(minutes=90)
    register_lab_start(uid, "csfc-naf2", [inst.id], 90, ends, db, background_tasks)
    return _build_response("csfc-naf2", now, ends, cost, rem, [creds])


@app.post("/api/start-lab-csfc-naf3")
async def start_lab_csfc_naf3(
    req: StartLabRequest, background_tasks: BackgroundTasks,
    db: Session = Depends(get_db), _cu: User = Depends(get_current_user),
):
    uid = req.userID
    if db.query(EndLabTrial).filter(EndLabTrial.user_id == uid).first():
        raise HTTPException(400, "Another lab is currently running.")
    cost, rem = deduct_credits_or_raise(uid, "csfc-naf3", db)
    sess = _get_ec2_session(); ec2 = sess.resource('ec2')
    inst = ec2.create_instances(
        ImageId='ami-05f035b6ae19778b1', MinCount=1, MaxCount=1,
        InstanceType='t2.micro', KeyName='kp22',
        SecurityGroupIds=[_DEFAULT_SG], UserData=_USER_DATA_AMAZON_LINUX,
    )[0]
    creds = _wait_windows(sess, inst)
    now = datetime.utcnow(); ends = now + timedelta(minutes=90)
    register_lab_start(uid, "csfc-naf3", [inst.id], 90, ends, db, background_tasks)
    return _build_response("csfc-naf3", now, ends, cost, rem, [creds])


# --- wf3: Kali Linux, t2.medium, ami-05e58d56a542aa2bf, duration=60 ---
@app.post("/api/start-lab-csfc-wf3")
async def start_lab_csfc_wf3(
    req: StartLabRequest, background_tasks: BackgroundTasks,
    db: Session = Depends(get_db), _cu: User = Depends(get_current_user),
):
    uid = req.userID
    if db.query(EndLabTrial).filter(EndLabTrial.user_id == uid).first():
        raise HTTPException(400, "Another lab is currently running.")
    cost, rem = deduct_credits_or_raise(uid, "csfc-wf3", db)
    sess = _get_ec2_session(); ec2 = sess.resource('ec2')
    inst = ec2.create_instances(
        ImageId='ami-05e58d56a542aa2bf', MinCount=1, MaxCount=1,
        InstanceType='t2.medium', KeyName='kkp1',
        SecurityGroupIds=[_DEFAULT_SG], UserData=_USER_DATA_KALI_LINUX,
    )[0]
    creds = _wait_linux(sess, inst)
    now = datetime.utcnow(); ends = now + timedelta(minutes=60)
    register_lab_start(uid, "csfc-wf3", [inst.id], 60, ends, db, background_tasks)
    return _build_response("csfc-wf3", now, ends, cost, rem, [creds])


# --- crf1: Amazon Linux, t2.micro, duration=90 ---
@app.post("/api/start-lab-csfc-crf1")
async def start_lab_csfc_crf1(
    req: StartLabRequest, background_tasks: BackgroundTasks,
    db: Session = Depends(get_db), _cu: User = Depends(get_current_user),
):
    uid = req.userID
    if db.query(EndLabTrial).filter(EndLabTrial.user_id == uid).first():
        raise HTTPException(400, "Another lab is currently running.")
    cost, rem = deduct_credits_or_raise(uid, "csfc-crf1", db)
    sess = _get_ec2_session(); ec2 = sess.resource('ec2')
    inst = ec2.create_instances(
        ImageId='ami-0e1d06225679bc1c5', MinCount=1, MaxCount=1,
        InstanceType='t2.micro', KeyName='kkp1',
        SecurityGroupIds=[_DEFAULT_SG], UserData=_USER_DATA_AMAZON_LINUX,
    )[0]
    creds = _wait_linux(sess, inst)
    now = datetime.utcnow(); ends = now + timedelta(minutes=90)
    register_lab_start(uid, "csfc-crf1", [inst.id], 90, ends, db, background_tasks)
    return _build_response("csfc-crf1", now, ends, cost, rem, [creds])


# --- rtf1: Kali Linux, t2.medium, duration=120 ---
@app.post("/api/start-lab-csfc-rtf1")
async def start_lab_csfc_rtf1(
    req: StartLabRequest, background_tasks: BackgroundTasks,
    db: Session = Depends(get_db), _cu: User = Depends(get_current_user),
):
    uid = req.userID
    if db.query(EndLabTrial).filter(EndLabTrial.user_id == uid).first():
        raise HTTPException(400, "Another lab is currently running.")
    cost, rem = deduct_credits_or_raise(uid, "csfc-rtf1", db)
    sess = _get_ec2_session(); ec2 = sess.resource('ec2')
    inst = ec2.create_instances(
        ImageId='ami-05e58d56a542aa2bf', MinCount=1, MaxCount=1,
        InstanceType='t2.medium', KeyName='kkp1',
        SecurityGroupIds=[_DEFAULT_SG], UserData=_USER_DATA_KALI_LINUX,
    )[0]
    creds = _wait_linux(sess, inst)
    now = datetime.utcnow(); ends = now + timedelta(minutes=120)
    register_lab_start(uid, "csfc-rtf1", [inst.id], 120, ends, db, background_tasks)
    return _build_response("csfc-rtf1", now, ends, cost, rem, [creds])


# --- rtf2: Mixed (Amazon Linux target + Kali attacker), duration=120 ---
@app.post("/api/start-lab-csfc-rtf2")
async def start_lab_csfc_rtf2(
    req: StartLabRequest, background_tasks: BackgroundTasks,
    db: Session = Depends(get_db), _cu: User = Depends(get_current_user),
):
    uid = req.userID
    if db.query(EndLabTrial).filter(EndLabTrial.user_id == uid).first():
        raise HTTPException(400, "Another lab is currently running.")
    cost, rem = deduct_credits_or_raise(uid, "csfc-rtf2", db)
    sess = _get_ec2_session(); ec2 = sess.resource('ec2')
    target = ec2.create_instances(
        ImageId='ami-0e1d06225679bc1c5', MinCount=1, MaxCount=1,
        InstanceType='t2.micro', KeyName='kkp1',
        SecurityGroupIds=[_DEFAULT_SG], UserData=_USER_DATA_AMAZON_LINUX,
    )[0]
    attacker = ec2.create_instances(
        ImageId='ami-05e58d56a542aa2bf', MinCount=1, MaxCount=1,
        InstanceType='t2.medium', KeyName='kkp1',
        SecurityGroupIds=[_DEFAULT_SG], UserData=_USER_DATA_KALI_LINUX,
    )[0]
    creds_list = [_wait_linux(sess, target), _wait_linux(sess, attacker)]
    ids = [target.id, attacker.id]
    now = datetime.utcnow(); ends = now + timedelta(minutes=120)
    register_lab_start(uid, "csfc-rtf2", ids, 120, ends, db, background_tasks)
    return _build_response("csfc-rtf2", now, ends, cost, rem, creds_list)


# --- rtf3: Mixed (Amazon Linux target + Kali attacker), duration=150 ---
@app.post("/api/start-lab-csfc-rtf3")
async def start_lab_csfc_rtf3(
    req: StartLabRequest, background_tasks: BackgroundTasks,
    db: Session = Depends(get_db), _cu: User = Depends(get_current_user),
):
    uid = req.userID
    if db.query(EndLabTrial).filter(EndLabTrial.user_id == uid).first():
        raise HTTPException(400, "Another lab is currently running.")
    cost, rem = deduct_credits_or_raise(uid, "csfc-rtf3", db)
    sess = _get_ec2_session(); ec2 = sess.resource('ec2')
    target = ec2.create_instances(
        ImageId='ami-0e1d06225679bc1c5', MinCount=1, MaxCount=1,
        InstanceType='t2.micro', KeyName='kkp1',
        SecurityGroupIds=[_DEFAULT_SG], UserData=_USER_DATA_AMAZON_LINUX,
    )[0]
    attacker = ec2.create_instances(
        ImageId='ami-05e58d56a542aa2bf', MinCount=1, MaxCount=1,
        InstanceType='t2.medium', KeyName='kkp1',
        SecurityGroupIds=[_DEFAULT_SG], UserData=_USER_DATA_KALI_LINUX,
    )[0]
    creds_list = [_wait_linux(sess, target), _wait_linux(sess, attacker)]
    ids = [target.id, attacker.id]
    now = datetime.utcnow(); ends = now + timedelta(minutes=150)
    register_lab_start(uid, "csfc-rtf3", ids, 150, ends, db, background_tasks)
    return _build_response("csfc-rtf3", now, ends, cost, rem, creds_list)


# --- rtf4: Kali Linux, t2.medium, duration=90 ---
@app.post("/api/start-lab-csfc-rtf4")
async def start_lab_csfc_rtf4(
    req: StartLabRequest, background_tasks: BackgroundTasks,
    db: Session = Depends(get_db), _cu: User = Depends(get_current_user),
):
    uid = req.userID
    if db.query(EndLabTrial).filter(EndLabTrial.user_id == uid).first():
        raise HTTPException(400, "Another lab is currently running.")
    cost, rem = deduct_credits_or_raise(uid, "csfc-rtf4", db)
    sess = _get_ec2_session(); ec2 = sess.resource('ec2')
    inst = ec2.create_instances(
        ImageId='ami-05e58d56a542aa2bf', MinCount=1, MaxCount=1,
        InstanceType='t2.medium', KeyName='kkp1',
        SecurityGroupIds=[_DEFAULT_SG], UserData=_USER_DATA_KALI_LINUX,
    )[0]
    creds = _wait_linux(sess, inst)
    now = datetime.utcnow(); ends = now + timedelta(minutes=90)
    register_lab_start(uid, "csfc-rtf4", [inst.id], 90, ends, db, background_tasks)
    return _build_response("csfc-rtf4", now, ends, cost, rem, [creds])


# --- btf1: Amazon Linux, t2.micro, duration=60 ---
@app.post("/api/start-lab-csfc-btf1")
async def start_lab_csfc_btf1(
    req: StartLabRequest, background_tasks: BackgroundTasks,
    db: Session = Depends(get_db), _cu: User = Depends(get_current_user),
):
    uid = req.userID
    if db.query(EndLabTrial).filter(EndLabTrial.user_id == uid).first():
        raise HTTPException(400, "Another lab is currently running.")
    cost, rem = deduct_credits_or_raise(uid, "csfc-btf1", db)
    sess = _get_ec2_session(); ec2 = sess.resource('ec2')
    inst = ec2.create_instances(
        ImageId='ami-0e1d06225679bc1c5', MinCount=1, MaxCount=1,
        InstanceType='t2.micro', KeyName='kkp1',
        SecurityGroupIds=[_DEFAULT_SG], UserData=_USER_DATA_AMAZON_LINUX,
    )[0]
    creds = _wait_linux(sess, inst)
    now = datetime.utcnow(); ends = now + timedelta(minutes=60)
    register_lab_start(uid, "csfc-btf1", [inst.id], 60, ends, db, background_tasks)
    return _build_response("csfc-btf1", now, ends, cost, rem, [creds])


# --- ptf1: Amazon Linux, t2.micro, duration=120 ---
@app.post("/api/start-lab-csfc-ptf1")
async def start_lab_csfc_ptf1(
    req: StartLabRequest, background_tasks: BackgroundTasks,
    db: Session = Depends(get_db), _cu: User = Depends(get_current_user),
):
    uid = req.userID
    if db.query(EndLabTrial).filter(EndLabTrial.user_id == uid).first():
        raise HTTPException(400, "Another lab is currently running.")
    cost, rem = deduct_credits_or_raise(uid, "csfc-ptf1", db)
    sess = _get_ec2_session(); ec2 = sess.resource('ec2')
    inst = ec2.create_instances(
        ImageId='ami-0e1d06225679bc1c5', MinCount=1, MaxCount=1,
        InstanceType='t2.micro', KeyName='kkp1',
        SecurityGroupIds=[_DEFAULT_SG], UserData=_USER_DATA_AMAZON_LINUX,
    )[0]
    creds = _wait_linux(sess, inst)
    now = datetime.utcnow(); ends = now + timedelta(minutes=120)
    register_lab_start(uid, "csfc-ptf1", [inst.id], 120, ends, db, background_tasks)
    return _build_response("csfc-ptf1", now, ends, cost, rem, [creds])


# --- ptf2: Windows, t2.micro, no UserData, duration=120 ---
@app.post("/api/start-lab-csfc-ptf2")
async def start_lab_csfc_ptf2(
    req: StartLabRequest, background_tasks: BackgroundTasks,
    db: Session = Depends(get_db), _cu: User = Depends(get_current_user),
):
    uid = req.userID
    if db.query(EndLabTrial).filter(EndLabTrial.user_id == uid).first():
        raise HTTPException(400, "Another lab is currently running.")
    cost, rem = deduct_credits_or_raise(uid, "csfc-ptf2", db)
    sess = _get_ec2_session(); ec2 = sess.resource('ec2')
    inst = ec2.create_instances(
        ImageId='ami-05f035b6ae19778b1', MinCount=1, MaxCount=1,
        InstanceType='t2.micro', KeyName='kp22',
        SecurityGroupIds=[_DEFAULT_SG],
    )[0]
    creds = _wait_windows(sess, inst)
    now = datetime.utcnow(); ends = now + timedelta(minutes=120)
    register_lab_start(uid, "csfc-ptf2", [inst.id], 120, ends, db, background_tasks)
    return _build_response("csfc-ptf2", now, ends, cost, rem, [creds])


# --- CUTM group ---

# --- cutm1: Mixed (Amazon Linux + Kali), duration=150 ---
@app.post("/api/start-lab-cutm1")
async def start_lab_cutm1(
    req: StartLabRequest, background_tasks: BackgroundTasks,
    db: Session = Depends(get_db), _cu: User = Depends(get_current_user),
):
    uid = req.userID
    if db.query(EndLabTrial).filter(EndLabTrial.user_id == uid).first():
        raise HTTPException(400, "Another lab is currently running.")
    cost, rem = deduct_credits_or_raise(uid, "cutm1", db)
    sess = _get_ec2_session(); ec2 = sess.resource('ec2')
    target = ec2.create_instances(
        ImageId='ami-0e1d06225679bc1c5', MinCount=1, MaxCount=1,
        InstanceType='t2.micro', KeyName='kkp1',
        SecurityGroupIds=[_DEFAULT_SG], UserData=_USER_DATA_AMAZON_LINUX,
    )[0]
    attacker = ec2.create_instances(
        ImageId='ami-05e58d56a542aa2bf', MinCount=1, MaxCount=1,
        InstanceType='t2.medium', KeyName='kkp1',
        SecurityGroupIds=[_DEFAULT_SG], UserData=_USER_DATA_KALI_LINUX,
    )[0]
    creds_list = [_wait_linux(sess, target), _wait_linux(sess, attacker)]
    ids = [target.id, attacker.id]
    now = datetime.utcnow(); ends = now + timedelta(minutes=150)
    register_lab_start(uid, "cutm1", ids, 150, ends, db, background_tasks)
    return _build_response("cutm1", now, ends, cost, rem, creds_list)


# --- cutm2: Mixed (Windows + Kali), duration=60 ---
@app.post("/api/start-lab-cutm2")
async def start_lab_cutm2(
    req: StartLabRequest, background_tasks: BackgroundTasks,
    db: Session = Depends(get_db), _cu: User = Depends(get_current_user),
):
    uid = req.userID
    if db.query(EndLabTrial).filter(EndLabTrial.user_id == uid).first():
        raise HTTPException(400, "Another lab is currently running.")
    cost, rem = deduct_credits_or_raise(uid, "cutm2", db)
    sess = _get_ec2_session(); ec2 = sess.resource('ec2')
    win_inst = ec2.create_instances(
        ImageId='ami-05f035b6ae19778b1', MinCount=1, MaxCount=1,
        InstanceType='t2.micro', KeyName='kp22',
        SecurityGroupIds=[_DEFAULT_SG],
    )[0]
    kali_inst = ec2.create_instances(
        ImageId='ami-05e58d56a542aa2bf', MinCount=1, MaxCount=1,
        InstanceType='t2.medium', KeyName='kkp1',
        SecurityGroupIds=[_DEFAULT_SG], UserData=_USER_DATA_KALI_LINUX,
    )[0]
    win_creds = _wait_windows(sess, win_inst)
    kali_creds = _wait_linux(sess, kali_inst)
    ids = [win_inst.id, kali_inst.id]
    now = datetime.utcnow(); ends = now + timedelta(minutes=60)
    register_lab_start(uid, "cutm2", ids, 60, ends, db, background_tasks)
    return _build_response("cutm2", now, ends, cost, rem, [win_creds, kali_creds])


# --- cutm3: Mixed (Windows + Kali), duration=150 ---
@app.post("/api/start-lab-cutm3")
async def start_lab_cutm3(
    req: StartLabRequest, background_tasks: BackgroundTasks,
    db: Session = Depends(get_db), _cu: User = Depends(get_current_user),
):
    uid = req.userID
    if db.query(EndLabTrial).filter(EndLabTrial.user_id == uid).first():
        raise HTTPException(400, "Another lab is currently running.")
    cost, rem = deduct_credits_or_raise(uid, "cutm3", db)
    sess = _get_ec2_session(); ec2 = sess.resource('ec2')
    win_inst = ec2.create_instances(
        ImageId='ami-05f035b6ae19778b1', MinCount=1, MaxCount=1,
        InstanceType='t2.micro', KeyName='kp22',
        SecurityGroupIds=[_DEFAULT_SG],
    )[0]
    kali_inst = ec2.create_instances(
        ImageId='ami-05e58d56a542aa2bf', MinCount=1, MaxCount=1,
        InstanceType='t2.medium', KeyName='kkp1',
        SecurityGroupIds=[_DEFAULT_SG], UserData=_USER_DATA_KALI_LINUX,
    )[0]
    win_creds = _wait_windows(sess, win_inst)
    kali_creds = _wait_linux(sess, kali_inst)
    ids = [win_inst.id, kali_inst.id]
    now = datetime.utcnow(); ends = now + timedelta(minutes=150)
    register_lab_start(uid, "cutm3", ids, 150, ends, db, background_tasks)
    return _build_response("cutm3", now, ends, cost, rem, [win_creds, kali_creds])


# --- cutm4: Mixed (Amazon Linux + Kali), duration=150 ---
@app.post("/api/start-lab-cutm4")
async def start_lab_cutm4(
    req: StartLabRequest, background_tasks: BackgroundTasks,
    db: Session = Depends(get_db), _cu: User = Depends(get_current_user),
):
    uid = req.userID
    if db.query(EndLabTrial).filter(EndLabTrial.user_id == uid).first():
        raise HTTPException(400, "Another lab is currently running.")
    cost, rem = deduct_credits_or_raise(uid, "cutm4", db)
    sess = _get_ec2_session(); ec2 = sess.resource('ec2')
    target = ec2.create_instances(
        ImageId='ami-0e1d06225679bc1c5', MinCount=1, MaxCount=1,
        InstanceType='t2.micro', KeyName='kkp1',
        SecurityGroupIds=[_DEFAULT_SG], UserData=_USER_DATA_AMAZON_LINUX,
    )[0]
    attacker = ec2.create_instances(
        ImageId='ami-05e58d56a542aa2bf', MinCount=1, MaxCount=1,
        InstanceType='t2.medium', KeyName='kkp1',
        SecurityGroupIds=[_DEFAULT_SG], UserData=_USER_DATA_KALI_LINUX,
    )[0]
    creds_list = [_wait_linux(sess, target), _wait_linux(sess, attacker)]
    ids = [target.id, attacker.id]
    now = datetime.utcnow(); ends = now + timedelta(minutes=150)
    register_lab_start(uid, "cutm4", ids, 150, ends, db, background_tasks)
    return _build_response("cutm4", now, ends, cost, rem, creds_list)


# --- cutm7: Linux, 3x Amazon Linux, duration=180 ---
@app.post("/api/start-lab-cutm7")
async def start_lab_cutm7(
    req: StartLabRequest, background_tasks: BackgroundTasks,
    db: Session = Depends(get_db), _cu: User = Depends(get_current_user),
):
    uid = req.userID
    if db.query(EndLabTrial).filter(EndLabTrial.user_id == uid).first():
        raise HTTPException(400, "Another lab is currently running.")
    cost, rem = deduct_credits_or_raise(uid, "cutm7", db)
    sess = _get_ec2_session(); ec2 = sess.resource('ec2')
    instances = ec2.create_instances(
        ImageId='ami-0e1d06225679bc1c5', MinCount=3, MaxCount=3,
        InstanceType='t2.micro', KeyName='kkp1',
        SecurityGroupIds=[_DEFAULT_SG], UserData=_USER_DATA_AMAZON_LINUX,
    )
    creds_list = [_wait_linux(sess, i) for i in instances]
    ids = [i.id for i in instances]
    now = datetime.utcnow(); ends = now + timedelta(minutes=180)
    register_lab_start(uid, "cutm7", ids, 180, ends, db, background_tasks)
    return _build_response("cutm7", now, ends, cost, rem, creds_list)


# --- cutm8: Linux, 3x Amazon Linux, duration=180 ---
@app.post("/api/start-lab-cutm8")
async def start_lab_cutm8(
    req: StartLabRequest, background_tasks: BackgroundTasks,
    db: Session = Depends(get_db), _cu: User = Depends(get_current_user),
):
    uid = req.userID
    if db.query(EndLabTrial).filter(EndLabTrial.user_id == uid).first():
        raise HTTPException(400, "Another lab is currently running.")
    cost, rem = deduct_credits_or_raise(uid, "cutm8", db)
    sess = _get_ec2_session(); ec2 = sess.resource('ec2')
    instances = ec2.create_instances(
        ImageId='ami-0e1d06225679bc1c5', MinCount=3, MaxCount=3,
        InstanceType='t2.micro', KeyName='kkp1',
        SecurityGroupIds=[_DEFAULT_SG], UserData=_USER_DATA_AMAZON_LINUX,
    )
    creds_list = [_wait_linux(sess, i) for i in instances]
    ids = [i.id for i in instances]
    now = datetime.utcnow(); ends = now + timedelta(minutes=180)
    register_lab_start(uid, "cutm8", ids, 180, ends, db, background_tasks)
    return _build_response("cutm8", now, ends, cost, rem, creds_list)


# --- ns-lab1: Windows, t2.micro, no UserData, duration=180 ---
@app.post("/api/start-lab-ns-lab1")
async def start_lab_ns_lab1(
    req: StartLabRequest, background_tasks: BackgroundTasks,
    db: Session = Depends(get_db), _cu: User = Depends(get_current_user),
):
    uid = req.userID
    if db.query(EndLabTrial).filter(EndLabTrial.user_id == uid).first():
        raise HTTPException(400, "Another lab is currently running.")
    cost, rem = deduct_credits_or_raise(uid, "ns-lab1", db)
    sess = _get_ec2_session(); ec2 = sess.resource('ec2')
    inst = ec2.create_instances(
        ImageId='ami-05f035b6ae19778b1', MinCount=1, MaxCount=1,
        InstanceType='t2.micro', KeyName='kp22',
        SecurityGroupIds=[_DEFAULT_SG],
    )[0]
    creds = _wait_windows(sess, inst)
    now = datetime.utcnow(); ends = now + timedelta(minutes=180)
    register_lab_start(uid, "ns-lab1", [inst.id], 180, ends, db, background_tasks)
    return _build_response("ns-lab1", now, ends, cost, rem, [creds])


# --- dpi-lab1: Mixed (Windows public + 2x Ubuntu private), special VPC, duration=180 ---
@app.post("/api/start-lab-dpi-lab1")
async def start_lab_dpi_lab1(
    req: StartLabRequest, background_tasks: BackgroundTasks,
    db: Session = Depends(get_db), _cu: User = Depends(get_current_user),
):
    uid = req.userID
    if db.query(EndLabTrial).filter(EndLabTrial.user_id == uid).first():
        raise HTTPException(400, "Another lab is currently running.")
    cost, rem = deduct_credits_or_raise(uid, "dpi-lab1", db)

    _DPI_SG = 'sg-0c070594fa44a4778'
    _PUBLIC_SUBNET = 'subnet-0ef7f69912046d7c9'
    _PRIVATE_SUBNETS = ['subnet-0a6fe54cf263a009b', 'subnet-06913d967217525fc']

    sess = _get_ec2_session()
    ec2 = sess.resource('ec2')
    ec2_client = sess.client('ec2')

    # Pick least-loaded private subnet
    private_subnet_id = _PRIVATE_SUBNETS[0]
    for sid in _PRIVATE_SUBNETS:
        try:
            resp = ec2_client.describe_instances(Filters=[
                {'Name': 'subnet-id', 'Values': [sid]},
                {'Name': 'instance-state-name', 'Values': ['pending', 'running']},
            ])
            count = sum(len(r['Instances']) for r in resp['Reservations'])
            if count < 16000:
                private_subnet_id = sid
                break
        except Exception:
            pass

    win_instances = ec2.create_instances(
        ImageId='ami-05f035b6ae19778b1', MinCount=1, MaxCount=1,
        InstanceType='t3a.micro', KeyName='kp22',
        SecurityGroupIds=[_DPI_SG], SubnetId=_PUBLIC_SUBNET,
    )
    ubuntu_instances = ec2.create_instances(
        ImageId='ami-0e35ddab05955cf57', MinCount=2, MaxCount=2,
        InstanceType='t2.large', KeyName='kp22',
        SecurityGroupIds=[_DPI_SG], SubnetId=private_subnet_id,
        UserData=_USER_DATA_UBUNTU_DPI,
        BlockDeviceMappings=[{
            'DeviceName': '/dev/sda1',
            'Ebs': {'VolumeSize': 70, 'DeleteOnTermination': True, 'VolumeType': 'gp3'},
        }],
    )

    creds_list = []
    all_ids = []
    for inst in win_instances:
        c = _wait_windows(sess, inst)
        creds_list.append(c)
        all_ids.append(inst.id)
    for inst in ubuntu_instances:
        inst.wait_until_running()
        ec2_client.get_waiter('instance_status_ok').wait(InstanceIds=[inst.id])
        inst.load()
        creds_list.append({
            'instance_id': inst.id,
            'public_ip': inst.public_ip_address,
            'private_ip': inst.private_ip_address,
            'os_type': 'linux',
            'username': 'ubuntu',
            'password': _GUI_VNC_PASSWORD,
            'ssh_command': f'ssh ubuntu@{inst.public_ip_address}',
        })
        all_ids.append(inst.id)

    now = datetime.utcnow(); ends = now + timedelta(minutes=180)
    register_lab_start(uid, "dpi-lab1", all_ids, 180, ends, db, background_tasks)
    return _build_response("dpi-lab1", now, ends, cost, rem, creds_list)


# --- gui-lab1: browser desktop, t2.medium, noVNC on 6080, duration=120 ------
# Path must be "/api/start-lab-" + lab_catalog.lab_id — the frontend builds it
# as `/start-lab-${labId}` (main-web/src/services/api.js startLab).
@app.post("/api/start-lab-gui-lab1")
async def start_lab_gui_lab1(
    req: StartLabRequest, background_tasks: BackgroundTasks,
    db: Session = Depends(get_db), _cu: User = Depends(get_current_user),
):
    uid = req.userID
    if db.query(EndLabTrial).filter(EndLabTrial.user_id == uid).first():
        raise HTTPException(400, "Another lab is currently running.")
    cost, rem = deduct_credits_or_raise(uid, "gui-lab1", db)
    sess = _get_ec2_session(); ec2 = sess.resource('ec2')
    inst = ec2.create_instances(
        ImageId=_GUI_AMI, MinCount=1, MaxCount=1,
        InstanceType='t2.medium', KeyName='kp22',
        SecurityGroupIds=[_DEFAULT_SG],
    )[0]
    creds = _wait_novnc(sess, inst)
    now = datetime.utcnow(); ends = now + timedelta(minutes=120)
    register_lab_start(uid, "gui-lab1", [inst.id], 120, ends, db, background_tasks)
    return _build_response("gui-lab1", now, ends, cost, rem, [creds])


# ============================================================================
# LINK REACHABILITY CHECK (admin)
# ============================================================================
# LMS content references PDFs by URL, so a private S3 object or a dead link
# silently renders an empty viewer for students. This lets an admin verify a
# semester's links before publishing it.

class UrlCheckRequest(BaseModel):
    urls: List[str]


_URL_CHECK_MAX = 40
_URL_CHECK_TIMEOUT = 12


def _is_public_host(host: str) -> bool:
    """Reject anything resolving to a private/loopback address so this
    endpoint can't be used to probe the box's own internal network."""
    try:
        infos = socket.getaddrinfo(host, None)
    except Exception:
        return False
    for info in infos:
        try:
            ip = ipaddress.ip_address(info[4][0])
        except ValueError:
            return False
        if (ip.is_private or ip.is_loopback or ip.is_link_local
                or ip.is_reserved or ip.is_multicast or ip.is_unspecified):
            return False
    return True


def _check_one_url(raw: str) -> dict:
    out = {
        "url": raw, "ok": False, "status": None, "content_type": None,
        "size": None, "frame_blocked": False, "error": None,
    }
    url = (raw or "").strip()
    if not url:
        out["error"] = "Empty URL"
        return out

    parsed = urlparse(url)
    if parsed.scheme not in ("http", "https"):
        out["error"] = "URL must start with http:// or https://"
        return out
    if not parsed.hostname:
        out["error"] = "Malformed URL"
        return out
    if not _is_public_host(parsed.hostname):
        out["error"] = "Host does not resolve to a public address"
        return out

    def _fetch(method):
        req = urllib.request.Request(
            url, method=method,
            headers={"User-Agent": "CyberDojo-LinkCheck/1.0"},
        )
        return urllib.request.urlopen(req, timeout=_URL_CHECK_TIMEOUT)

    try:
        try:
            resp = _fetch("HEAD")
        except urllib.error.HTTPError as e:
            # Some hosts reject HEAD but serve GET fine. Retrying costs only
            # the headers — the response is closed before the body is read.
            if e.code in (403, 405, 501):
                resp = _fetch("GET")
            else:
                raise
        with resp:
            out["status"] = resp.status
            out["ok"] = 200 <= resp.status < 300
            out["content_type"] = resp.headers.get("Content-Type")
            length = resp.headers.get("Content-Length")
            out["size"] = int(length) if length and length.isdigit() else None
            # A framing header breaks the in-app PDF viewer even when the
            # file itself is perfectly reachable.
            xfo = (resp.headers.get("X-Frame-Options") or "").strip().upper()
            out["frame_blocked"] = xfo in ("DENY", "SAMEORIGIN")
    except urllib.error.HTTPError as e:
        out["status"] = e.code
        out["error"] = f"HTTP {e.code} {e.reason}"
    except Exception as e:
        out["error"] = (str(e) or "Unreachable")[:200]
    return out


def _collect_content_urls(content) -> List[dict]:
    """Every place an LMS content item can carry a URL, flattened and labelled
    so a failure points at the exact field that needs fixing."""
    out: List[dict] = []
    for item in (content or []):
        if not isinstance(item, dict):
            continue
        name = item.get("title") or item.get("slug") or "untitled"
        if item.get("url"):
            out.append({"url": item["url"], "label": f"{name} — PDF"})
        if item.get("theory_url"):
            out.append({"url": item["theory_url"], "label": f"{name} — Theory Material"})
        for i, lab in enumerate(item.get("labs") or []):
            if isinstance(lab, dict) and lab.get("manual_url"):
                out.append({
                    "url": lab["manual_url"],
                    "label": f"{name} — {lab.get('name') or f'Lab {i + 1}'} manual",
                })
        for i, con in enumerate(item.get("concepts") or []):
            if isinstance(con, dict) and con.get("url"):
                out.append({"url": con["url"], "label": f"{name} — concept {i + 1}"})
    return out


def _run_checks(urls: List[str]) -> List[dict]:
    if not urls:
        return []
    with ThreadPoolExecutor(max_workers=8) as pool:
        return list(pool.map(_check_one_url, urls))


@app.post("/api/admin/check-urls")
def admin_check_urls(req: UrlCheckRequest, _: Admin = Depends(get_current_admin)):
    """Ad-hoc check for URLs not yet saved (the draft JSON in the editor)."""
    urls = [u for u in (req.urls or []) if isinstance(u, str)][:_URL_CHECK_MAX]
    return {"results": _run_checks(urls)}


@app.post("/api/admin/universities/{uid}/programs/{pid}/semesters/{sid}/check-links")
def admin_check_semester_links(
    uid: int, pid: int, sid: int,
    admin: Admin = Depends(get_current_admin), db: Session = Depends(get_db),
):
    """Check the URLs in a saved semester and cache the summary on the row, so
    the admin list can render a reachability tag without re-checking."""
    sem = db.query(UniversitySemester).filter(
        UniversitySemester.id == sid,
        UniversitySemester.program_id == pid,
        UniversitySemester.university_id == uid,
    ).first()
    if not sem:
        raise HTTPException(status_code=404, detail="Semester not found")

    entries = _collect_content_urls(sem.content)[:_URL_CHECK_MAX]
    results = _run_checks([e["url"] for e in entries])

    sem.links_total = len(results)
    sem.links_ok = sum(1 for r in results if r["ok"])
    sem.links_warn = sum(1 for r in results if r["ok"] and r["frame_blocked"])
    sem.links_checked_at = datetime.now(timezone.utc)
    db.commit()

    write_audit(db, admin.id, "university_semester", sem.id, "check_links", None, {
        "total": sem.links_total, "ok": sem.links_ok, "warn": sem.links_warn,
    })

    return {
        "semester": _sem_row(sem),
        "entries": entries,
        "results": results,
    }


# ============================================================================
# ASSESSMENTS
# ============================================================================
@app.get("/api/assessments")
def list_assessments(user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    rows = db.query(Assessment).filter(Assessment.is_active == True).all()
    return [_assessment_row(a) for a in rows]


@app.get("/api/assessments/{slug}")
def get_assessment(slug: str, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    a = db.query(Assessment).filter(Assessment.slug == slug, Assessment.is_active == True).first()
    if not a:
        raise HTTPException(status_code=404, detail="Assessment not found")
    questions = (
        db.query(AssessmentQuestion)
        .filter(AssessmentQuestion.assessment_id == a.id)
        .order_by(AssessmentQuestion.order_num)
        .all()
    )
    attempt = db.query(AssessmentAttempt).filter(
        AssessmentAttempt.assessment_id == a.id,
        AssessmentAttempt.user_id == user.id,
    ).first()
    return {
        **_assessment_row(a),
        "questions": [_question_row(q) for q in questions],
        "attempt": {
            "id": attempt.id,
            "status": attempt.status,
            "score": attempt.score,
            "total_questions": attempt.total_questions,
            "violations": attempt.violations,
        } if attempt else None,
    }


@app.post("/api/assessments/{slug}/start")
def start_assessment(slug: str, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    a = db.query(Assessment).filter(Assessment.slug == slug, Assessment.is_active == True).first()
    if not a:
        raise HTTPException(status_code=404, detail="Assessment not found")
    existing = db.query(AssessmentAttempt).filter(
        AssessmentAttempt.assessment_id == a.id,
        AssessmentAttempt.user_id == user.id,
    ).first()
    if existing:
        return {"attempt_id": existing.id, "status": existing.status, "score": existing.score}
    total = db.query(AssessmentQuestion).filter(AssessmentQuestion.assessment_id == a.id).count()
    attempt = AssessmentAttempt(
        assessment_id=a.id,
        user_id=user.id,
        total_questions=total,
        status="in_progress",
    )
    db.add(attempt)
    db.commit()
    db.refresh(attempt)
    return {"attempt_id": attempt.id, "status": attempt.status, "score": None}


@app.post("/api/assessments/{slug}/submit")
def submit_assessment(slug: str, req: AssessmentAnswerSubmit, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    a = db.query(Assessment).filter(Assessment.slug == slug).first()
    if not a:
        raise HTTPException(status_code=404, detail="Assessment not found")
    attempt = db.query(AssessmentAttempt).filter(
        AssessmentAttempt.assessment_id == a.id,
        AssessmentAttempt.user_id == user.id,
    ).first()
    if not attempt:
        raise HTTPException(status_code=400, detail="No attempt started")
    questions = db.query(AssessmentQuestion).filter(AssessmentQuestion.assessment_id == a.id).all()
    score = 0
    for q in questions:
        submitted = req.answers.get(str(q.id), [])
        if sorted(submitted) == sorted(q.correct_answers):
            score += 1
    attempt.score = score
    attempt.total_questions = len(questions)
    attempt.answers_payload = req.answers
    attempt.status = "completed"
    attempt.completed_at = sqlfunc.now()
    db.commit()
    return {
        "score": score,
        "total": len(questions),
        "percentage": round(score / len(questions) * 100) if questions else 0,
        "violations": attempt.violations,
    }


@app.post("/api/assessments/violations")
def log_violation(req: AssessmentViolationLog, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    attempt = db.query(AssessmentAttempt).filter(
        AssessmentAttempt.id == req.attempt_id,
        AssessmentAttempt.user_id == user.id,
    ).first()
    if not attempt:
        raise HTTPException(status_code=404, detail="Attempt not found")
    v = AssessmentViolation(
        attempt_id=attempt.id,
        user_id=user.id,
        violation_type=req.violation_type,
    )
    db.add(v)
    attempt.violations = (attempt.violations or 0) + 1
    db.commit()
    return {"violations": attempt.violations}


# ============================================================================
# HEALTH
# ============================================================================
@app.get("/")
def health():
    return {"status": "ok", "message": "CyberDojo backend running"}
